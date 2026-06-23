# -*- coding: utf-8 -*-
"""Pulizia + classificazione lista traslochi Easyquote -> prospects.json (contratto kit)."""
import openpyxl, json, re, unicodedata

WB = openpyxl.load_workbook('lista_originale.xlsx')
ws = WB.active

# righe ROSSE (lista Lucca, competitor del pilota) da ESCLUDERE - lette dal colore FFF4CCCC su col A
RED = {99,100,118,119,120,121,135,136,137,138,185,187}

def norm(s):
    s = unicodedata.normalize('NFKD', str(s or '')).encode('ascii','ignore').decode().lower()
    return re.sub(r'\s+',' ', re.sub(r'[^a-z0-9 ]',' ', s)).strip()

def domain(site):
    s = re.sub(r'^https?://','', str(site or '').strip().lower()).rstrip('/')
    s = re.sub(r'^www\.','', s)
    return s.split('/')[0]

# macro-aree
FI = {'firenze','scandicci','sesto fiorentino','campi bisenzio','lastra a signa','signa','calenzano',
      'fiesole','grassina','san donnino','poggio','empoli','luco di mugello','bagno a ripoli'}
PT = {'prato','pistoia','montecatini terme','monsummano terme','borgo a buggiano','pieve a nievole',
      'pescia','campo tizzoro'}
PI = {'pisa','cascina','pontedera','ponsacco','calcinaia','vicopisano','navacchio','san miniato',
      "santa croce sull'arno",'castelfranco di sotto','bientina','casciana terme lari','san giuliano terme',
      'vecchiano','cascine nuove','santa maria ai monti'}
LI = {'livorno','cecina','vada','collesalvetti','castelnuovo della misericordia','piombino'}
LU = {'lucca','capannori','viareggio','altopascio','guamo','ponte a moriano','massarosa','lido di camaiore',
      'montramito','fornaci di barga','turchetto','carrara','marina di carrara'}
AR = {'arezzo','montevarchi','san giovanni valdarno','sansepolcro','siena','cerchiaia','monteriggioni',
      "colle di val d'elsa",'rapolano terme','arbia','torita di siena','chianciano terme',"sant'angelo",
      'grosseto','follonica','arcidosso','gavorrano','porto santo stefano','san mauro'}
def area(zona, nome):
    z = norm(zona)
    if z in {norm(x) for x in FI}: return 'Firenze e provincia'
    if z in {norm(x) for x in PT}: return 'Prato-Pistoia'
    if z in {norm(x) for x in PI}: return 'Pisa e area pisana'
    if z in {norm(x) for x in LI}: return 'Livorno e costa'
    if z in {norm(x) for x in LU}: return 'Lucca-Versilia-Massa'
    if z in {norm(x) for x in AR}: return 'Arezzo-Siena-Grosseto'
    n = norm(nome)
    if 'pontedera' in n or 'pisa' in n or 'san giuliano' in n: return 'Pisa e area pisana'
    if 'versilia' in n or 'viareggio' in n or 'capannori' in n or 'lucca' in n or 'camaiore' in n: return 'Lucca-Versilia-Massa'
    if 'livorno' in n: return 'Livorno e costa'
    if 'firenze' in n: return 'Firenze e provincia'
    return 'Firenze e provincia'  # batch senza zona: in prevalenza Firenze

def split_phones(raw):
    """ritorna (centralino, diretto). Il PRIMO numero va sempre in centralino
    (e il numero che il setter chiama, anche se e un cellulare della ditta);
    un eventuale secondo numero va in diretto."""
    if not raw: return '', ''
    s = str(raw).replace('=','').strip()
    s = re.sub(r'\+39\s*','', s)
    parts = [p.strip() for p in re.split(r'[\n;,/]+', s) if p.strip()]
    nums = []
    for p in parts:
        digits = re.sub(r'\D','', p)
        if digits and fmt(digits) not in nums:
            nums.append(fmt(digits))
    cent = nums[0] if nums else ''
    dire = nums[1] if len(nums) > 1 else ''
    return cent, dire

def fmt(d):
    d = re.sub(r'\D','', d)
    if d.startswith('3') and len(d) >= 9:           # cellulare 3xx xxx xxxx
        return f'{d[:3]} {d[3:6]} {d[6:]}'.strip()
    if d.startswith(('800','803','199','892')):     # numero verde / servizio
        return f'{d[:3]} {d[3:]}'
    if d.startswith('0'):
        if d.startswith('055'):  return f'055 {d[3:]}'
        if d.startswith('0586'): return f'0586 {d[4:]}'
        if d.startswith('0577'): return f'0577 {d[4:]}'
        if d.startswith('0564'): return f'0564 {d[4:]}'
        if d.startswith('0575'): return f'0575 {d[4:]}'
        if d.startswith('0573'): return f'0573 {d[4:]}'
        if d.startswith('0572'): return f'0572 {d[4:]}'
        if d.startswith('0571'): return f'0571 {d[4:]}'
        if d.startswith('0588'): return f'0588 {d[4:]}'
        if d.startswith('0583'): return f'0583 {d[4:]}'
        if d.startswith('0584'): return f'0584 {d[4:]}'
        if d.startswith('0585'): return f'0585 {d[4:]}'
        if len(d) >= 6:  # prefisso 4 cifre generico per province toscane
            return f'{d[:4]} {d[4:]}'
    return d

def legal_form(nome):
    n = ' '+norm(nome)+' '
    if ' srls ' in n or ' s r l s ' in n: return 'SRLS'
    if ' srl ' in n or ' s r l ' in n: return 'SRL'
    if ' snc ' in n or ' s n c ' in n: return 'SNC'
    if ' sas ' in n: return 'SAS'
    return ''

import os
TITOLARI = json.load(open('titolari.json', encoding='utf-8')) if os.path.exists('titolari.json') else {}

def nome_decisore(nome):
    # 1) nome reale trovato via ricerca (registro imprese / sito / ragione sociale)
    t = TITOLARI.get(norm(nome))
    if t and t.get('nome'):
        ruolo = t.get('ruolo') or 'titolare'
        return f"{t['nome']} ({ruolo})"
    # 2) pattern esplicito nella ragione sociale "DI Nome Cognome"
    m = re.search(r'\bDI\s+([A-ZÀ-Ù][A-Za-zÀ-ù\.]+(?:\s+[A-ZÀ-Ù][A-Za-zÀ-ù\.]+)?)\b', str(nome))
    if m:
        cand = m.group(1).strip().strip('.')
        if cand.upper() not in {'TRASLOCHI','SERVIZI','TRASPORTI'} and len(cand) > 2:
            return f'{cand.title()} (titolare) - confermare in chiamata'
    # 3) fallback onesto
    return 'Da chiedere al centralino (chiedere del titolare)'

HOOK_A = ("Abbiamo creato un configuratore di preventivi su misura per le aziende di traslochi: il cliente fa il "
          "preventivo da solo sul vostro sito e voi non perdete piu mezz'ore al telefono. Le va se le mostro come "
          "funziona in 15 minuti?")
HOOK_C = ("Aiutiamo le aziende di traslochi a dare al cliente il preventivo immediato online, cosi non perde tempo "
          "al telefono e non si fa scappare i clienti. Le mostro in una demo veloce come funziona?")

# telefoni trovati via ricerca per le aziende senza numero in lista (con fonte verificata)
PHONE_OVERRIDE = {
    'viticchi traslochi': ('335 376509', 'https://www.traslochi-italia.eu/ditte/traslochi/viticchi-traslochi/4020'),
    'frassi traslochi': ('327 131 7952', 'https://www.frassitraslochi.it/contatti/'),
    'bs traslochi stagno': ('347 579 2250', 'https://www.virgilio.it/italia/vicopisano/cat/TRASLOCHI.html'),
    'riviera express traslochi': ('0584 332645', 'https://aziende.virgilio.it/traslochi/massarosa-lu/riviera-express-s-a-s-di-cappelli-massimo-and-c'),
}

rows = []
seen_name, seen_dom, seen_phone = set(), set(), set()
pilot = 0

def classify(nome, zona, dom, only_social, cent, dire, fonte, origine, nome_dec, a=None):
    """costruisce il record prospect con la classificazione standard."""
    a = a or area(zona, nome)
    lf = legal_form(nome)
    has_site = bool(dom) and 'facebook' not in dom
    inv = 'Alto' if has_site else ('Medio' if only_social else 'Basso')
    nuovo = ''
    if has_site and lf in ('SRL','SRLS'):
        nuovo = 'Societa strutturata (' + lf + ') con sito web proprio: profilo in crescita, gia investe in struttura e digitale.'
    if has_site:
        evid = f'Sito web proprio attivo ({dom}).' + (f' Forma: {lf}.' if lf else '')
    elif only_social:
        evid = 'Presenza solo social (Facebook), nessun sito proprio.' + (f' Forma: {lf}.' if lf else '')
    else:
        evid = 'Nessun sito web rilevato (presenza digitale assente).' + (f' Forma: {lf}.' if lf else '')
    if has_site:
        perche = f'Impresa di traslochi attiva ({a}) con sito proprio: gia orientata ad acquisire clienti online, fit diretto per il configuratore IsyQuote.'
    else:
        perche = f'Impresa di traslochi attiva ({a}) senza funnel online: oggi prende richieste solo al telefono, massimo margine di miglioramento col configuratore.'
    return {
        'azienda': nome, 'settore': a, 'sede': zona or 'Toscana', 'sito': dom,
        'dimensione': 'Mid-large', 'investimento': inv, 'evidenza': evid, 'nuovo_investimento': nuovo,
        'decision_maker': 'Titolare', 'nome_decisore': nome_dec,
        'perche_in_target': perche, 'cosa_dire': HOOK_A if has_site else HOOK_C,
        'telefono_centralino': cent, 'telefono_diretto': dire, 'fonte_telefono': fonte, 'origine': origine
    }

def dedup_ok(nn, dom, phone_key):
    if nn in seen_name: return False
    if dom and dom in seen_dom and 'facebook' not in dom: return False
    if phone_key and phone_key in seen_phone: return False
    seen_name.add(nn)
    if dom and 'facebook' not in dom: seen_dom.add(dom)
    if phone_key: seen_phone.add(phone_key)
    return True

# ---------- 1) LISTA CLIENTE (xlsx) ----------
for r in range(2, ws.max_row+1):
    nome = ws.cell(r,1).value
    if not nome or not str(nome).strip(): continue
    nome = str(nome).strip()
    if r in RED: continue
    nn = norm(nome)
    dom = domain((ws.cell(r,6).value or '').strip())
    if 'lctraslochieservizi' in dom or 'l c traslochi e servizi' in nn or nn.startswith('l c traslochi'):
        pilot += 1; continue
    cent, dire = split_phones(ws.cell(r,4).value)
    ovr_fonte = ''
    if not cent and nn in PHONE_OVERRIDE:
        cent, ovr_fonte = PHONE_OVERRIDE[nn]
    phone_key = re.sub(r'\D','', cent or dire)
    if nn in seen_name or (dom and dom in seen_dom and 'facebook' not in dom) or (phone_key and phone_key in seen_phone):
        continue
    if not dedup_ok(nn, dom, phone_key): continue
    gmaps = (ws.cell(r,7).value or '').strip()
    fonte = ovr_fonte or gmaps or (('https://'+dom) if dom else 'Lista cliente (ricerca Google Maps / diretta)')
    only_social = bool(dom) and 'facebook' in dom
    rows.append(classify(nome, (ws.cell(r,2).value or '').strip(), dom, only_social,
                         cent, dire, fonte, 'In lista cliente', nome_decisore(nome)))

n_cliente = len(rows)

# ---------- 2) NUOVI LEAD (scouting, new/*.json) ----------
FILE_AREA = {
    'new_firenze_citta': 'Firenze e provincia', 'new_firenze_prov': 'Firenze e provincia',
    'new_pisa': 'Pisa e area pisana', 'new_livorno': 'Livorno e costa',
    'new_prato_pistoia': 'Prato-Pistoia', 'new_arezzo_siena': 'Arezzo-Siena-Grosseto',
    'new_grosseto': 'Arezzo-Siena-Grosseto', 'new_massacarrara': 'Lucca-Versilia-Massa',
}
import glob
for f in sorted(glob.glob('new/new_*.json')):
    key = os.path.splitext(os.path.basename(f))[0]
    a = FILE_AREA.get(key, 'Toscana')
    try: data = json.load(open(f, encoding='utf-8'))
    except Exception as e: print('SKIP', f, e); continue
    for d in data:
        nome = (d.get('azienda') or '').strip()
        if not nome: continue
        nn = norm(nome)
        dom = domain(d.get('sito') or '')
        cent = fmt(re.sub(r'\D','', (d.get('telefono_centralino') or ''))) if d.get('telefono_centralino') else ''
        dire = fmt(re.sub(r'\D','', (d.get('telefono_diretto') or ''))) if d.get('telefono_diretto') else ''
        if not cent and dire: cent, dire = dire, ''   # garantisci centralino sempre pieno
        phone_key = re.sub(r'\D','', cent or dire)
        if not cent:  # cancello qualita: niente telefono -> scarta
            continue
        if nn in seen_name or (dom and dom in seen_dom and 'facebook' not in dom) or (phone_key and phone_key in seen_phone):
            continue
        if not dedup_ok(nn, dom, phone_key): continue
        nd = (d.get('nome_decisore') or '').strip()
        if nd:
            nome_dec = nd if '(' in nd else f"{nd} ({(d.get('ruolo') or 'titolare').strip()})"
        else:
            nome_dec = nome_decisore(nome)
        fonte = (d.get('fonte_telefono') or '').strip() or (('https://'+dom) if dom else 'Ricerca scouting')
        only_social = bool(dom) and 'facebook' in dom
        rows.append(classify(nome, (d.get('sede') or '').strip(), dom, only_social,
                             cent, dire, fonte, 'Nuovo lead (ricerca)', nome_dec, a=a))

n_nuovi = len(rows) - n_cliente

json.dump(rows, open('prospects.json','w',encoding='utf-8'), ensure_ascii=False, indent=1)
from collections import Counter
print(f'PROSPECT: {len(rows)} (cliente {n_cliente} + nuovi {n_nuovi}) | pilota escluso:', pilot, '| righe rosse escluse:', len(RED))
print('--- per area ---'); [print(f'{v:3d}  {k}') for k,v in Counter(x['settore'] for x in rows).most_common()]
print('--- investimento ---'); [print(f'{v:3d}  {k}') for k,v in Counter(x['investimento'] for x in rows).most_common()]
print('SRL/SRLS con sito (verdi A):', sum(1 for x in rows if x['nuovo_investimento']))
print('con centralino:', sum(1 for x in rows if x['telefono_centralino']),
      '| con diretto:', sum(1 for x in rows if x['telefono_diretto']),
      '| nome reale:', sum(1 for x in rows if 'Da chiedere' not in x['nome_decisore']))
