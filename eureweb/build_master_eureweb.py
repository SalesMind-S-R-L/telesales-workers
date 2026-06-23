#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Master Eureweb / AD Lab - shooting foto/video AI per grandi brand.
Stesso impianto del Master IMEON / Giglioli. Niente emoji.
Lista costruita per SCOUTING (no DB pregresso). 6 settori dal perimetro Giulia Rizzi.
Esclusi i 130+ clienti gia Eureweb (no doppioni / no conflitto).
"""
import urllib.parse, json, glob, re, unicodedata
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

# ---- colori (stessi del master Giglioli/IMEON) ----
C_TITLE = PatternFill('solid', fgColor='1F3864')
C_HDR   = PatternFill('solid', fgColor='1F3864')
C_A     = PatternFill('solid', fgColor='C8E6C9')  # verde
C_B     = PatternFill('solid', fgColor='FFF59D')  # giallo
C_C     = PatternFill('solid', fgColor='F2F2F2')  # grigio
C_SEC   = PatternFill('solid', fgColor='D9E1F2')  # azzurro sezione
F_TITLE = Font(color='FFFFFF', bold=True, size=14)
F_HDRW  = Font(color='FFFFFF', bold=True)
F_SUB   = Font(italic=True, size=10, color='555555')
F_B     = Font(bold=True)
WRAP    = Alignment(wrap_text=True, vertical='top')
TOP     = Alignment(vertical='top')
thin    = Side(style='thin', color='D0D0D0')
BORD    = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = openpyxl.Workbook()

def title_block(ws, title, sub):
    ws['A1'] = title; ws['A1'].font = F_TITLE; ws['A1'].fill = C_TITLE
    ws['A2'] = sub;   ws['A2'].font = F_SUB

def fill_row_title(ws, ncol):
    for c in range(1, ncol+1):
        ws.cell(row=1, column=c).fill = C_TITLE

# =====================================================================
# DATASET PROSPECT  (azienda, settore, sede, sito, prio)
# prio: A = grande advertiser molto attivo in contenuti/video/social
#       B = brand mid-large premium attivo in comunicazione
#       C = realta con budget dedicato, da lavorare dopo
# =====================================================================
AUTO='Automotive'; FOOD='Food & Beverage'; RET='Retail & GDO'
FASH='Fashion & Lifestyle'; FIN='Finance & Insurance'; TLC='Telecomunicazioni'

DATA = [
 # ---------------- AUTOMOTIVE ----------------
 ('Volkswagen Group Italia',AUTO,'Verona','volkswagen.it','A'),
 ('Stellantis Italia (Fiat)',AUTO,'Torino','stellantis.com','A'),
 ('Jeep Italia',AUTO,'Torino','jeep.it','A'),
 ('Alfa Romeo',AUTO,'Torino','alfaromeo.it','A'),
 ('Renault Italia',AUTO,'Roma','renault.it','A'),
 ('Toyota Motor Italia',AUTO,'Roma','toyota.it','A'),
 ('BMW Italia',AUTO,'San Donato Milanese (MI)','bmw.it','A'),
 ('Mercedes-Benz Italia',AUTO,'Roma','mercedes-benz.it','A'),
 ('Audi Italia',AUTO,'Verona','audi.it','A'),
 ('Ford Italia',AUTO,'Roma','ford.it','A'),
 ('Nissan Italia',AUTO,'Capena (RM)','nissan.it','A'),
 ('Hyundai Italia',AUTO,'Milano','hyundai.it','A'),
 ('Ducati Motor Holding',AUTO,'Bologna','ducati.com','A'),
 ('Dacia Italia',AUTO,'Roma','dacia.it','B'),
 ('Volvo Car Italia',AUTO,'Garbagnate Milanese (MI)','volvocars.it','B'),
 ('Cupra Italia',AUTO,'Verona','cupraofficial.it','B'),
 ('Skoda Italia',AUTO,'Verona','skoda-auto.it','B'),
 ('Opel Italia',AUTO,'Torino','opel.it','B'),
 ('Mazda Italia',AUTO,'Roma','mazda.it','B'),
 ('Honda Italia',AUTO,'Roma','honda.it','B'),
 ('MINI Italia',AUTO,'San Donato Milanese (MI)','mini.it','B'),
 ('Jaguar Land Rover Italia',AUTO,'Roma','landrover.it','B'),
 ('DS Automobiles Italia',AUTO,'Torino','dsautomobiles.it','B'),
 ('Lexus Italia',AUTO,'Roma','lexus.it','B'),
 ('Lancia',AUTO,'Torino','lancia.it','B'),
 ('Yamaha Motor Italia',AUTO,'Lesmo (MB)','yamaha-motor.it','B'),
 ('Harley-Davidson Italia',AUTO,'Milano','harley-davidson.com','B'),
 ('Bridgestone Italia',AUTO,'Roma','bridgestone.it','B'),
 ('Michelin Italia',AUTO,'Torino','michelin.it','B'),
 ('Brembo',AUTO,'Curno (BG)','brembo.com','B'),
 ('Continental Italia',AUTO,'Milano','continental.it','C'),
 ('Goodyear Italia',AUTO,'Roma','goodyear.it','C'),
 ('Marelli',AUTO,'Corbetta (MI)','marelli.com','C'),
 ('Benelli',AUTO,'Pesaro','benelli.com','C'),
 ('KTM Italia',AUTO,'Milano','ktm.com','C'),
 ('Energica Motor Company',AUTO,'Modena','energicamotor.com','C'),
 # ---------------- FOOD & BEVERAGE ----------------
 ('Ferrero',FOOD,'Alba (CN)','ferrero.it','A'),
 ('Barilla',FOOD,'Parma','barilla.com','A'),
 ('Lavazza',FOOD,'Torino','lavazza.it','A'),
 ('Gruppo Campari',FOOD,'Sesto San Giovanni (MI)','camparigroup.com','A'),
 ('Nestle Italia',FOOD,'Assago (MI)','nestle.it','A'),
 ('Heineken Italia',FOOD,'Pollein (AO)','heineken.it','A'),
 ('Sanpellegrino',FOOD,'Milano','sanpellegrino-corporate.it','A'),
 ('Granarolo',FOOD,'Bologna','granarolo.it','A'),
 ('Galbani (Lactalis Italia)',FOOD,'Milano','galbani.it','A'),
 ('Pastificio Rana',FOOD,'San Giovanni Lupatoto (VR)','giovannirana.it','A'),
 ('illycaffe',FOOD,'Trieste','illy.com','A'),
 ('De Cecco',FOOD,'Fara San Martino (CH)','dececco.com','B'),
 ('Rio Mare (Bolton Food)',FOOD,'Cermenate (CO)','riomare.it','B'),
 ('Star (GB Foods Italia)',FOOD,'Agrate Brianza (MB)','staralimentare.it','B'),
 ('Findus Italia',FOOD,'Roma','findus.it','B'),
 ('Birra Peroni',FOOD,'Roma','peroniitalia.com','B'),
 ('Segafredo Zanetti',FOOD,'Bologna','segafredo.it','B'),
 ('Kimbo',FOOD,'Melito di Napoli (NA)','kimbo.it','B'),
 ('Fratelli Branca Distillerie',FOOD,'Milano','branca.it','B'),
 ('Illva Saronno (Disaronno)',FOOD,'Saronno (VA)','illva.it','B'),
 ('Gruppo Montenegro',FOOD,'Zola Predosa (BO)','gruppomontenegro.com','B'),
 ('Ferrarelle',FOOD,'Riardo (CE)','ferrarelle.it','B'),
 ('Acqua Minerale San Benedetto',FOOD,'Scorze (VE)','sanbenedetto.it','B'),
 ('Rovagnati',FOOD,'Biassono (MB)','rovagnati.it','B'),
 ('Fratelli Beretta',FOOD,'Trezzo sull Adda (MI)','fratelliberetta.com','B'),
 ('Gruppo Veronesi (AIA)',FOOD,'San Martino Buon Albergo (VR)','aia-food.com','B'),
 ('Amadori',FOOD,'San Vittore di Cesena (FC)','amadori.it','B'),
 ('Pastificio Rummo',FOOD,'Benevento','pastarummo.it','C'),
 ('Pastificio Granoro',FOOD,'Corato (BA)','granoro.it','C'),
 ('Citterio',FOOD,'Rho (MI)','citterio.com','C'),
 ('Negroni Salumi',FOOD,'Cremona','negronisalumi.it','C'),
 ('Auricchio',FOOD,'Cremona','auricchio.it','C'),
 ('Zonin1821',FOOD,'Gambellara (VI)','zonin1821.it','C'),
 ('Marchesi Antinori',FOOD,'Firenze','antinori.it','C'),
 ('Santa Margherita Gruppo Vinicolo',FOOD,'Fossalta di Portogruaro (VE)','santamargherita.com','C'),
 ('Guido Berlucchi',FOOD,'Corte Franca (BS)','berlucchi.it','C'),
 ('Acqua Lete',FOOD,'Pratella (CE)','acqualete.it','C'),
 ('Caffe Vergnano',FOOD,'Santena (TO)','caffevergnano.com','C'),
 # ---------------- RETAIL & GDO ----------------
 ('Esselunga',RET,'Pioltello (MI)','esselunga.it','A'),
 ('Conad',RET,'Bologna','conad.it','A'),
 ('Lidl Italia',RET,'Arcole (VR)','lidl.it','A'),
 ('Eurospin',RET,'San Martino Buon Albergo (VR)','eurospin.it','A'),
 ('Carrefour Italia',RET,'Milano','carrefour.it','A'),
 ('Decathlon Italia',RET,'Milano','decathlon.it','A'),
 ('IKEA Italia',RET,'Carugate (MI)','ikea.com','A'),
 ('Leroy Merlin Italia',RET,'Rozzano (MI)','leroymerlin.it','A'),
 ('OVS',RET,'Venezia-Mestre','ovs.it','A'),
 ('Calzedonia Group',RET,'Verona','calzedonia.com','A'),
 ('Selex Gruppo Commerciale',RET,'Milano','selexgc.it','B'),
 ('Gruppo Pam (Pam Panorama)',RET,'Spinea (VE)','gruppopam.it','B'),
 ('Bennet',RET,'Montano Lucino (CO)','bennet.com','B'),
 ('MD Discount',RET,'Gricignano di Aversa (CE)','mdspa.it','B'),
 ('Penny Market Italia',RET,'Cesano Boscone (MI)','penny.it','B'),
 ('CRAI',RET,'Milano','craiweb.it','B'),
 ('Despar Italia',RET,'Casalecchio di Reno (BO)','desparitalia.it','B'),
 ('OBI Italia',RET,'Milano','obi-italia.it','B'),
 ('Bricoman Italia',RET,'Rozzano (MI)','bricoman.it','B'),
 ('Mondo Convenienza',RET,'Civitavecchia (RM)','mondoconv.it','B'),
 ('Unieuro',RET,'Forli','unieuro.it','B'),
 ('Expert Italia',RET,'Roma','expert.it','B'),
 ('Tigota (Gottardo)',RET,'Vigonza (PD)','tigota.it','B'),
 ('Acqua e Sapone',RET,'Spoltore (PE)','acquaesapone.it','B'),
 ('Douglas Italia',RET,'Milano','douglas.it','B'),
 ('Sephora Italia',RET,'Milano','sephora.it','B'),
 ('Kiko Milano',RET,'Bergamo','kikocosmetics.com','B'),
 ('La Rinascente',RET,'Milano','rinascente.it','B'),
 ('Eataly',RET,'Milano','eataly.it','B'),
 ('Deichmann Italia',RET,'Vigonza (PD)','deichmann.com','C'),
 ('Original Marines',RET,'Napoli','originalmarines.com','C'),
 ('Terranova (Teddy)',RET,'Rimini','terranovastyle.com','C'),
 ('Piazza Italia',RET,'Nola (NA)','piazzaitalia.it','C'),
 ('NaturaSi (EcorNaturaSi)',RET,'Verona','naturasi.it','C'),
 ('laFeltrinelli',RET,'Milano','lafeltrinelli.it','C'),
 ('Mondadori Retail',RET,'Milano','mondadoristore.it','C'),
 ('Kasanova',RET,'Cologno Monzese (MI)','kasanova.com','C'),
 ('Risparmio Casa',RET,'Pomezia (RM)','risparmiocasa.com','C'),
 # ---------------- FASHION & LIFESTYLE ----------------
 ('Moncler',FASH,'Milano','moncler.com','A'),
 ('Dolce & Gabbana',FASH,'Milano','dolcegabbana.com','A'),
 ('Versace',FASH,'Milano','versace.com','A'),
 ('Valentino',FASH,'Milano','valentino.com','A'),
 ('Bottega Veneta',FASH,'Milano','bottegaveneta.com','A'),
 ('Bvlgari',FASH,'Roma','bulgari.com','A'),
 ('Brunello Cucinelli',FASH,'Solomeo (PG)','brunellocucinelli.com','A'),
 ('Tod s Group',FASH,'Sant Elpidio a Mare (FM)','todsgroup.com','A'),
 ('Max Mara Fashion Group',FASH,'Reggio Emilia','maxmara.com','A'),
 ('EssilorLuxottica',FASH,'Milano','essilorluxottica.com','A'),
 ('Diesel (OTB Group)',FASH,'Breganze (VI)','diesel.com','A'),
 ('Loro Piana',FASH,'Quarona (VC)','loropiana.com','B'),
 ('Geox',FASH,'Montebelluna (TV)','geox.com','B'),
 ('Furla',FASH,'Bologna','furla.com','B'),
 ('Coccinelle',FASH,'Parma','coccinelle.com','B'),
 ('Piquadro',FASH,'Gaggio Montano (BO)','piquadro.com','B'),
 ('Pinko (Cris Conf)',FASH,'Fidenza (PR)','pinko.com','B'),
 ('Liu Jo',FASH,'Carpi (MO)','liujo.com','B'),
 ('Patrizia Pepe (Tessilform)',FASH,'Capalle (FI)','patriziapepe.com','B'),
 ('Elisabetta Franchi',FASH,'Bologna','elisabettafranchi.com','B'),
 ('Stone Island',FASH,'Ravarino (MO)','stoneisland.com','B'),
 ('Woolrich',FASH,'Bologna','woolrich.com','B'),
 ('Herno',FASH,'Lesa (NO)','herno.com','B'),
 ('K-Way (BasicNet)',FASH,'Torino','k-way.com','B'),
 ('Colmar',FASH,'Monza','colmar.it','B'),
 ('Save The Duck',FASH,'Milano','savetheduck.it','B'),
 ('Replay (Fashion Box)',FASH,'Asolo (TV)','replay.it','B'),
 ('Benetton Group',FASH,'Ponzano Veneto (TV)','benetton.com','B'),
 ('Yamamay (Inticom)',FASH,'Gallarate (VA)','yamamay.com','B'),
 ('Boggi Milano',FASH,'Cantu (CO)','boggi.com','B'),
 ('Harmont & Blaine',FASH,'Caivano (NA)','harmontblaine.com','B'),
 ('Safilo Group',FASH,'Padova','safilogroup.com','B'),
 ('Marcolin',FASH,'Longarone (BL)','marcolin.com','B'),
 ('Canali',FASH,'Sovico (MB)','canali.com','C'),
 ('Borsalino',FASH,'Spinetta Marengo (AL)','borsalino.com','C'),
 ('Morellato Group',FASH,'Santa Giustina in Colle (PD)','morellato.com','C'),
 ('Nomination',FASH,'Sesto Fiorentino (FI)','nomination.com','C'),
 ('Marco Bicego',FASH,'Trissino (VI)','marcobicego.com','C'),
 ('Collistar',FASH,'Milano','collistar.it','C'),
 ('Acqua di Parma (LVMH)',FASH,'Milano','acquadiparma.com','C'),
 ('Davines',FASH,'Parma','davines.com','C'),
 ('Bottega Verde',FASH,'Pienza (SI)','bottegaverde.it','C'),
 ('Deborah Milano',FASH,'Milano','deborahmilano.com','C'),
 ('Pupa Milano',FASH,'Trezzo sull Adda (MI)','pupamilano.com','C'),
 ('Twinset',FASH,'Carpi (MO)','twinset.com','C'),
 # ---------------- FINANCE & INSURANCE ----------------
 ('Intesa Sanpaolo',FIN,'Torino','intesasanpaolo.com','A'),
 ('UniCredit',FIN,'Milano','unicredit.it','A'),
 ('Generali Italia',FIN,'Mogliano Veneto (TV)','generali.it','A'),
 ('Allianz Italia',FIN,'Trieste','allianz.it','A'),
 ('Poste Italiane',FIN,'Roma','poste.it','A'),
 ('Banco BPM',FIN,'Milano','bancobpm.it','A'),
 ('Banca Mediolanum',FIN,'Basiglio (MI)','bancamediolanum.it','A'),
 ('FinecoBank',FIN,'Milano','finecobank.com','A'),
 ('BPER Banca',FIN,'Modena','bper.it','B'),
 ('Mediobanca',FIN,'Milano','mediobanca.com','B'),
 ('Credit Agricole Italia',FIN,'Parma','credit-agricole.it','B'),
 ('Credem',FIN,'Reggio Emilia','credem.it','B'),
 ('BNL Gruppo BNP Paribas',FIN,'Roma','bnl.it','B'),
 ('ING Italia',FIN,'Milano','ing.it','B'),
 ('Widiba',FIN,'Milano','widiba.it','B'),
 ('UnipolSai',FIN,'Bologna','unipolsai.it','B'),
 ('Reale Mutua (Reale Group)',FIN,'Torino','realemutua.it','B'),
 ('Zurich Italia',FIN,'Milano','zurich.it','B'),
 ('AXA Italia',FIN,'Milano','axa.it','B'),
 ('Groupama Assicurazioni',FIN,'Roma','groupama.it','B'),
 ('Nexi',FIN,'Milano','nexi.it','B'),
 ('American Express Italia',FIN,'Roma','americanexpress.it','B'),
 ('Vittoria Assicurazioni',FIN,'Milano','vittoriaassicurazioni.com','C'),
 ('Genertel',FIN,'Trieste','genertel.it','C'),
 ('Prima Assicurazioni',FIN,'Milano','prima.it','C'),
 ('Satispay',FIN,'Milano','satispay.com','C'),
 ('Scalapay',FIN,'Milano','scalapay.com','C'),
 ('N26 Italia',FIN,'Milano','n26.com','C'),
 ('Banca Ifis',FIN,'Venezia-Mestre','bancaifis.it','C'),
 ('Banca Sella',FIN,'Biella','sella.it','C'),
 ('illimity',FIN,'Milano','illimity.com','C'),
 ('Mastercard Italia',FIN,'Milano','mastercard.it','C'),
 # ---------------- TELECOMUNICAZIONI ----------------
 ('TIM',TLC,'Roma','tim.it','A'),
 ('Vodafone Italia',TLC,'Milano','vodafone.it','A'),
 ('WindTre',TLC,'Rho (MI)','windtre.it','A'),
 ('Fastweb',TLC,'Milano','fastweb.it','A'),
 ('Iliad Italia',TLC,'Milano','iliad.it','A'),
 ('Sky Italia',TLC,'Milano','sky.it','A'),
 ('DAZN Italia',TLC,'Milano','dazn.com','A'),
 ('Eolo',TLC,'Busto Arsizio (VA)','eolo.it','B'),
 ('Tiscali',TLC,'Cagliari','tiscali.it','B'),
 ('Open Fiber',TLC,'Milano','openfiber.it','B'),
 ('PosteMobile',TLC,'Roma','postemobile.it','B'),
 ('Very Mobile (WindTre)',TLC,'Milano','verymobile.it','B'),
 ('ho. Mobile (Vodafone)',TLC,'Milano','ho-mobile.it','C'),
 ('Kena Mobile (TIM)',TLC,'Roma','kenamobile.it','C'),
 ('CoopVoce',TLC,'Firenze','coopvoce.it','C'),
 ('RTL 102.5',TLC,'Cologno Monzese (MI)','rtl.it','C'),
 ('Mondadori Media',TLC,'Segrate (MI)','mondadori.it','C'),
]

# ---- ruoli decision maker per settore (ruoli, NON nomi: i nomi si reperiscono in fase scraping) ----
DM = {
 AUTO:['Marketing & Communication Director','Brand Manager','Head of Brand & Product Communication'],
 FOOD:['Direttore Marketing','Brand Manager','Head of Marketing & Communication'],
 RET :['Direttore Marketing & Comunicazione','Head of Brand & Communication','Marketing & CRM Director'],
 FASH:['Marketing & Communication Director','Brand Image & Communication Director','Head of Digital & Content'],
 FIN :['Chief Marketing Officer','Responsabile Comunicazione & Brand','Head of Marketing'],
 TLC :['Brand & Advertising Director','Head of Marketing Communication','Brand & Content Director'],
}

# ---- "perche in target" per settore ----
PT = {
 AUTO:'Settore tra i primi investitori adv in Italia. Campagne video/social ad alta frequenza, lanci modello e configuratori: forte fabbisogno di contenuti visivi.',
 FOOD:'Food & beverage tra i top spender pubblicitari. Cataloghi prodotto, packaging, ricette e campagne stagionali: produzione contenuti continua.',
 RET :'GDO/retail: volantini, campagne promozionali, contenuti always-on per social ed e-commerce. Budget marketing strutturati e ricorrenti.',
 FASH:'Fashion & lifestyle: shooting di collezione, lookbook, campagne stagionali. Storicamente abituati a produzioni fotografiche di alto livello.',
 FIN :'Finance/insurance: forte spesa in brand e advertising, campagne istituzionali e di prodotto su tutti i touchpoint. Esigenza di contenuti premium e scalabili.',
 TLC :'Telco/media: tra i maggiori advertiser nazionali, campagne sempre on con video e social ad altissima frequenza.',
}

# ---- ganci "cosa dire" (5 varianti, nessun prezzo, obiettivo = appuntamento) ----
HOOKS = [
 'Realizziamo shooting foto e video interamente in AI: qualita da produzione tradizionale con tempi e costi ridotti. Vorrei fissare una breve conoscitiva con chi segue contenuti e campagne.',
 'Aiutiamo brand come il vostro a produrre campagne visive (social, ADV, spot, affissioni) con shooting generati in AI. Posso fissare una call conoscitiva con il responsabile marketing?',
 'Produciamo contenuti visivi AI di alto livello per advertising e social. Vorremmo mostrarvi alcuni case in una conoscitiva: chi segue la comunicazione e i contenuti?',
 'Siamo una unit specializzata in shooting foto/video AI per grandi brand. Vorrei capire come gestite oggi la produzione contenuti e fissare un confronto conoscitivo.',
 'Creiamo set fotografici e video AI fedeli ai vostri ambienti e prodotti, pronti per ogni touchpoint. Le va un breve appuntamento conoscitivo con chi cura le campagne?',
]

# score trasparente: base prio + intensita contenuti del settore
PRIO_BASE = {'A':90,'B':75,'C':60}
SECT_W = {FASH:4, RET:3, FOOD:3, AUTO:3, TLC:2, FIN:1}

def li_search(az):
    q = urllib.parse.quote(az + ' marketing')
    return f'https://www.linkedin.com/search/results/people/?keywords={q}'

# arricchisci i record
recs = []
sect_counter = {}
for i,(az,se,sede,sito,prio) in enumerate(DATA):
    n = sect_counter.get(se,0); sect_counter[se]=n+1
    dm = DM[se][n % len(DM[se])]
    hook = HOOKS[i % len(HOOKS)]
    score = PRIO_BASE[prio] + SECT_W[se]
    canale = 'Chiamata (centralino > marketing)' if prio=='A' else ('Chiamata + LinkedIn' if prio=='B' else 'LinkedIn + email')
    recs.append(dict(az=az, se=se, sede=sede, sito=sito, prio=prio, dm=dm,
                     hook=hook, perche=PT[se], score=score, canale=canale, li=li_search(az)))

# =====================================================================
# MERGE ricerca per-azienda: evidenza investimento adv/contenuti + nuovi lead
# (file in research/*.json prodotti dagli agenti di ricerca)
# =====================================================================
def norm(s):
    s = unicodedata.normalize('NFKD', str(s)).encode('ascii','ignore').decode().lower()
    s = re.sub(r'\(.*?\)',' ', s)
    s = re.sub(r'[^a-z0-9 ]',' ', s)
    s = re.sub(r'\b(italia|italy|group|gruppo|spa|srl|s p a|holding)\b',' ', s)
    return re.sub(r'\s+',' ', s).strip()

def canon(s):
    t = str(s).lower()
    if 'auto' in t or 'motor' in t or 'moto' in t: return AUTO
    if 'food' in t or 'bever' in t: return FOOD
    if 'retail' in t or 'gdo' in t: return RET
    if 'fashion' in t or 'lifestyle' in t or 'beauty' in t or 'moda' in t or 'lusso' in t: return FASH
    if 'finance' in t or 'insur' in t or 'assicur' in t or 'bank' in t or 'fintech' in t: return FIN
    return TLC

def norm_dim(s):
    t = str(s).lower()
    if 'enter' in t or 'multinaz' in t or 'colosso' in t or 'gigante' in t: return 'Enterprise/multinazionale'
    if 'mid' in t or 'media' in t or 'nicchia' in t or 'premium' in t: return 'Mid-large'
    return 'Grande nazionale'

VALID_INV = ('Alto','Medio','Basso')
def clean_inv(v):
    v = str(v or '').strip().capitalize()
    return v if v in VALID_INV else 'Medio'
def clean_nuovo(v):
    v = str(v or '').strip()
    if not v or v.lower() in ('no','false','none'): return ''
    return v[:120]

# clienti gia Eureweb (esclusi sempre) per dedup dei nuovi lead
EXCL = ('Pirelli,Maserati,Kawasaki,MV Agusta,Peugeot,Citroen,Suzuki,Subaru,Kia,Mitsubishi,Chevrolet,Piaggio,NIU,'
 'Metzeler,Luna Rossa,Driver,MAK,Brixton,Biauto,Armani,Fondazione Prada,Damiani,Citizen,Breil,Bulova,'
 'Frederique Constant,Golden Point,Golden Lady,Motivi,Carpisa,Arena,Iceberg,Salewa,UPIM,Acqua dell Elba,'
 'Locman,Pompea,Camomilla,Clinians,Natura Verde,Geomar,Revita Care,Sodalis,ACBC,Vic Matie,Cafe Noir,'
 'Simonetta,Liabel,Sport Specialist,Cisalfa,Scarpe e Scarpe,Nicla,Coop,Mutti,Garofalo,Pernigotti,Carapelli,'
 'Caffarel,Amica Chips,Tennent s,Fileni,Frescobaldi,Pellini,Bofrost,Recla,Witor s,L Angelica,Fresco Pesce,'
 'Il Pescatore,Cassina,Poltrona Frau,Snaidero,Piscine Castiglione,Grandform,Edilkamin,Samsung,Acer,Lenovo,'
 'Euronics,Media World,MediaWorld,Energizer,Fujitsu,Sharp,Sennheiser,Olympia Splendid,Konica Minolta,iRobot,'
 'Olympus,Polti,Meliconi,Cerved,Henkel,GLS,Sisal,WEBUILD,Sant Anna,Hangar,Tupperware,Einhell,Motul,'
 'Brico.io,Upower,Weber,MaxMeyer,Plan Hotel,GioStyle,Urban Fitness,Marka,Martini,North Sails').split(',')
excl_norm = {norm(x) for x in EXCL}

# carica ricerca
verif = {}           # norm(nome) -> dict
new_leads_raw = []
for f in sorted(glob.glob('/Users/simocors/Desktop/telesales/eureweb/research/*.json')):
    raw = open(f).read().strip()
    if raw.startswith('```'):
        raw = raw.split('```')[1]
        if raw.startswith('json'): raw = raw[4:]
    d = json.loads(raw)
    for v in d.get('verificate', []):
        verif[norm(v.get('azienda',''))] = v
    new_leads_raw.extend(d.get('nuovi_lead', []))

# arricchisci i 206 di base + segna i fuori-target
base_norms = {norm(x['az']) for x in recs}
dropped = []
enriched = []
unmatched = []
for x in recs:
    v = verif.get(norm(x['az']))
    if v is None:
        unmatched.append(x['az'])
        x['inv']='Medio'; x['evid']='Da verificare'; x['nuovo']=''; x['origine']='In lista iniziale'
        enriched.append(x); continue
    if v.get('in_target', True) is False:
        dropped.append(x['az']); continue
    x['inv'] = clean_inv(v.get('investimento'))
    x['evid'] = str(v.get('evidenza','') or '').strip()[:300] or 'Da verificare'
    x['nuovo'] = clean_nuovo(v.get('nuovo_investimento'))
    p = str(v.get('prio','') or '').strip().upper()
    if p in ('A','B','C'):
        x['prio'] = p
        x['score'] = PRIO_BASE[p] + SECT_W[canon(x['se'])]
    x['origine'] = 'In lista iniziale'
    enriched.append(x)
recs = enriched

# aggiungi i nuovi lead (dedup vs base, vs esclusi, vs gia aggiunti)
def build_lead(L, origine, dim_from_agent=False):
    name = str(L.get('azienda','')).strip()
    se_disp = str(L.get('settore','')).strip() or 'Altro'
    if 'energ' in se_disp.lower() or 'utility' in se_disp.lower():
        se_disp = 'Telco / Energia-Utility'
    else:
        se_disp = canon(se_disp)
    cse = canon(se_disp)
    p = str(L.get('prio','') or 'B').strip().upper()
    if p not in ('A','B','C'): p='B'
    n = sect_counter.get(cse,0); sect_counter[cse]=n+1
    rec = dict(az=name, se=se_disp, sede=str(L.get('sede','') or ''), sito=str(L.get('sito','') or ''),
               prio=p, dm=DM[cse][n % len(DM[cse])], hook=HOOKS[(n) % len(HOOKS)], perche=PT[cse],
               score=PRIO_BASE[p]+SECT_W[cse],
               canale=('Chiamata (centralino > marketing)' if p=='A' else ('Chiamata + LinkedIn' if p=='B' else 'LinkedIn + email')),
               li=li_search(name), inv=clean_inv(L.get('investimento')),
               evid=str(L.get('evidenza','') or '').strip()[:300] or 'Da verificare',
               nuovo=clean_nuovo(L.get('nuovo_investimento')), origine=origine)
    if dim_from_agent:
        rec['dim_hint'] = norm_dim(L.get('dimensione','Mid-large'))
    return rec

seen = set(base_norms) | excl_norm
n_new1 = 0
for L in new_leads_raw:
    nn = norm(L.get('azienda',''))
    if not nn or nn in seen: continue
    seen.add(nn)
    recs.append(build_lead(L,'Nuovo lead (ricerca)')); n_new1+=1

# ricerca 2: nuovi lead MID-LARGE lavorabili e verificati (con dimensione dichiarata dall'agente)
new2_raw = []
for f in sorted(glob.glob('/Users/simocors/Desktop/telesales/eureweb/research2/*.json')):
    raw = open(f).read().strip()
    if raw.startswith('```'):
        raw = raw.split('```')[1]
        if raw.startswith('json'): raw = raw[4:]
    try: new2_raw.extend(json.loads(raw).get('nuovi_lead', []))
    except Exception: pass
n_new2 = 0
for L in new2_raw:
    nn = norm(L.get('azienda',''))
    if not nn or nn in seen: continue
    seen.add(nn)
    recs.append(build_lead(L,'Nuovo lead mid-large (ricerca 2)', dim_from_agent=True)); n_new2+=1

n_new = n_new1 + n_new2
n_drop = len(dropped)

# ---- telefoni verificati (centralino/sede da fonte ufficiale, file tel/out_*.json) ----
telmap = {}
for f in glob.glob('/Users/simocors/Desktop/telesales/eureweb/tel/out_*.json'):
    try: arr = json.load(open(f))
    except Exception: continue
    for e in (arr if isinstance(arr, list) else []):
        t = str(e.get('telefono','') or '').strip()
        if t: telmap[norm(e.get('azienda',''))] = {'t': t, 'f': str(e.get('fonte','') or '').strip()}
# cellulari/diretti pubblicati su fonte ufficiale (file tel/mob_*.json)
mobmap = {}
for f in glob.glob('/Users/simocors/Desktop/telesales/eureweb/tel/mob_*.json'):
    try: arr = json.load(open(f))
    except Exception: continue
    for e in (arr if isinstance(arr, list) else []):
        c = str(e.get('cellulare','') or e.get('telefono','') or '').strip()
        if c: mobmap[norm(e.get('azienda',''))] = c
# titolare / responsabile acquisti / cellulare personale (file tel/owner_*.json)
ownmap = {}
for f in glob.glob('/Users/simocors/Desktop/telesales/eureweb/tel/owner_*.json'):
    try: arr = json.load(open(f))
    except Exception: continue
    for e in (arr if isinstance(arr, list) else []):
        ownmap[norm(e.get('azienda',''))] = {
            'tit': str(e.get('titolare','') or '').strip(),
            'acq': str(e.get('acquisti','') or '').strip(),
            'cp': str(e.get('cellulare','') or '').strip(),
        }
# decisore marketing (file tel/dm_*.json): nome, ruolo, linkedin, email, cellulare (solo verificati/pubblici)
dmmap = {}
for f in glob.glob('/Users/simocors/Desktop/telesales/eureweb/tel/dm_*.json'):
    try: arr = json.load(open(f))
    except Exception: continue
    for e in (arr if isinstance(arr, list) else []):
        dmmap[norm(e.get('azienda',''))] = {
            'n': str(e.get('dm_nome','') or '').strip(),
            'r': str(e.get('dm_ruolo','') or '').strip(),
            'li': str(e.get('dm_linkedin','') or '').strip(),
            'em': str(e.get('dm_email','') or '').strip(),
            'ce': str(e.get('dm_cell','') or '').strip(),
        }
for x in recs:
    m = telmap.get(norm(x['az']))
    x['tel'] = m['t'] if m else ''
    x['telf'] = m['f'] if m else ''
    x['cell'] = mobmap.get(norm(x['az']), '')
    d = dmmap.get(norm(x['az']), {})
    x['dm_nome'] = d.get('n', ''); x['dm_ruolo'] = d.get('r', ''); x['dm_li'] = d.get('li', '')
    x['dm_email'] = d.get('em', ''); x['dm_cell'] = d.get('ce', '')
    o = ownmap.get(norm(x['az']), {})
    x['titolare'] = o.get('tit', '').split(' (')[0].strip()  # nome pulito, senza note tra parentesi
    x['acquisti'] = o.get('acq', '').split(' (')[0].strip()
    x['cellpers'] = o.get('cp', '')
n_tel = sum(1 for x in recs if x.get('tel'))
n_cell = sum(1 for x in recs if x.get('cell'))
n_tit = sum(1 for x in recs if x.get('titolare'))
n_acq = sum(1 for x in recs if x.get('acquisti'))
n_dm = sum(1 for x in recs if x.get('dm_nome'))
n_dmli = sum(1 for x in recs if x.get('dm_li'))
n_dmem = sum(1 for x in recs if x.get('dm_email'))

# ---- DIMENSIONE e ACCESSIBILITA del decision maker ----
# Enterprise/multinazionale = giganti tipo Haribo/Maserati: centralino + procurement + gare, ciclo lungo => Difficile
# Grande nazionale = grande ma piu raggiungibile => Media
# Mid-large = brand premium/di nicchia con budget, DM raggiungibile => Buona (i piu lavorabili)
GIANTS_RAW = [
 'Volkswagen','Stellantis','Fiat','Jeep','Alfa Romeo','Lancia','Renault','Toyota','BMW','Mercedes-Benz','Audi','Ford',
 'Nissan','Hyundai','Volvo','Cupra','Skoda','Opel','Mazda','Honda','MINI','Jaguar Land Rover','DS Automobiles','Lexus',
 'Dacia','Ducati','Yamaha','Harley-Davidson','Bridgestone','Michelin','Continental','Goodyear','Brembo','KTM','BYD',
 'MG Motor','Leapmotor','Iveco',
 'Ferrero','Barilla','Lavazza','Campari','Nestle','Sanpellegrino','Galbani','Granarolo','Rana','illycaffe','De Cecco',
 'Rio Mare','Star','Findus','Birra Peroni','Heineken','Coca-Cola HBC','Muller','Conserve','AIA','Veronesi','Amadori',
 'Segafredo','San Benedetto','Beretta','Bauli',
 'Esselunga','Conad','Carrefour','Lidl','Eurospin','Decathlon','IKEA','Leroy Merlin','Selex','Despar','Penny','OVS',
 'Calzedonia','Intimissimi','Tezenis','Unieuro','La Rinascente','Eataly','Douglas','Sephora','Kiko','OBI','Bricoman',
 'MD','Pam','Aldi','MediaWorld','Bennet',
 'Moncler','Dolce','Versace','Valentino','Bottega Veneta','Bvlgari','Brunello Cucinelli','Tod','Max Mara',
 'EssilorLuxottica','Luxottica','Diesel','Loro Piana','Geox','Benetton','Safilo','Marcolin','Furla','Stone Island',
 'Acqua di Parma','Pomellato','Aeffe',
 'Intesa Sanpaolo','UniCredit','Generali','Allianz','Poste','Banco BPM','Banca Mediolanum','FinecoBank','BPER',
 'Mediobanca','Credit Agricole','Credem','BNL','ING','UnipolSai','Reale Mutua','Zurich','AXA','Groupama','Nexi',
 'American Express','Mastercard','Banca Generali','Findomestic','Agos','Compass','Cofidis','Telepass','Anima','Amundi',
 'TIM','Vodafone','WindTre','Fastweb','Iliad','Sky','DAZN','Open Fiber','Enel','Eni','A2A','Acea','Edison','Iren',
 'Hera','Engie','Mediaset','Warner Bros','Spotify',
]
MID_RAW = [
 'Benelli','Sabelt','Sparco','OMP','Alpinestars','Givi','Fantic','Beta','Energica',
 'Rummo','Granoro','Citterio','Negroni','Auricchio','Zonin','Antinori','Santa Margherita','Berlucchi','Acqua Lete',
 'Caffe Vergnano','Caffe Borbone','Ferrari Trento','Kimbo','Branca','Montenegro','Disaronno','Illva','Ferrarelle','Rovagnati',
 'Davines','Collistar','Deborah','Pupa','Bottega Verde','Borsalino','Canali','Nomination','Morellato','Marco Bicego',
 'Twinset','Piquadro','Coccinelle','Save The Duck','Herno','K-Way','Colmar','Boggi','Harmont','Replay','Yamamay',
 'Patrizia Pepe','Pinko','Liu Jo','Elisabetta Franchi','Woolrich','Dainese','Diadora','Lotto','Sergio Tacchini',
 'Buccellati','Stroili','Brosway','L Erbolario','Diego Dalla Palma',
 'Prima Assicurazioni','Satispay','Scalapay','N26','Banca Ifis','Banca Sella','Vittoria','Genertel','Verti','ConTe',
 'Eolo','Tiscali','RTL','Aruba','Arcaplanet','Original Marines','Terranova','Piazza Italia','NaturaSi','Kasanova',
 'Risparmio Casa','Tigota','Acqua e Sapone','Deichmann','PittaRosso','Trony','Sorgenia',
]
giants = {norm(x) for x in GIANTS_RAW}
mids = {norm(x) for x in MID_RAW}

# ---- ri-prioritizzazione uniforme su segnali reali (investimento + nuovo investimento) ----
# A = gia investe Alto E si sta muovendo ora (nuovo investimento) -> piu caldo
# B = investe Alto (senza segnale nuovo) OPPURE ha un segnale di nuovo investimento
# C = il resto (investimento medio/basso senza segnali)
INV_PTS = {'Alto':60,'Medio':40,'Basso':20}
ACC_PTS = {'Buona':8,'Media':4,'Difficile':0}
ACC_OF = {'Enterprise/multinazionale':'Difficile','Grande nazionale':'Media','Mid-large':'Buona'}
for x in recs:
    nn = norm(x['az'])
    if nn in giants: dim='Enterprise/multinazionale'
    elif x.get('dim_hint'): dim=x['dim_hint']
    elif nn in mids: dim='Mid-large'
    else: dim='Grande nazionale'
    x['dim']=dim; x['acc']=ACC_OF[dim]
    has_nuovo = bool(x.get('nuovo'))
    inv = x.get('inv','Medio')
    # A = investe Alto e si muove ora; B = investe Alto, o ha segnale, o e un mid-large accessibile che investe; C = resto
    if inv=='Alto' and has_nuovo: p='A'
    elif inv=='Alto' or has_nuovo or (inv=='Medio' and x['acc']=='Buona'): p='B'
    else: p='C'
    x['prio'] = p
    # score: segnale d'acquisto + accessibilita (a parita di priorita salgono i piu lavorabili)
    x['score'] = INV_PTS.get(inv,40) + (25 if has_nuovo else 0) + ACC_PTS[x['acc']] + SECT_W[canon(x['se'])]
    # quick win = facile da lavorare (accessibilita buona) E investe davvero (Alto/Medio)
    x['quickwin'] = 'Quick win' if (x['acc']=='Buona' and inv in ('Alto','Medio')) else ''
    x['canale'] = 'Chiamata (centralino > marketing)' if p=='A' else ('Chiamata + LinkedIn' if p=='B' else 'LinkedIn + email')

n_giant = sum(1 for x in recs if x['dim'].startswith('Enterprise'))
n_grande = sum(1 for x in recs if x['dim']=='Grande nazionale')
n_mid = sum(1 for x in recs if x['dim']=='Mid-large')
n_qw = sum(1 for x in recs if x['quickwin'])

order = {'A':0,'B':1,'C':2}
recs.sort(key=lambda x:(order[x['prio']], -x['score'], x['se'], x['az']))

nA = sum(1 for x in recs if x['prio']=='A')
nB = sum(1 for x in recs if x['prio']=='B')
nC = sum(1 for x in recs if x['prio']=='C')
TOT = len(recs)
sect_count = {}
for x in recs: sect_count[x['se']] = sect_count.get(x['se'],0)+1

# clienti gia Eureweb esclusi dallo scouting (dal sito /clienti) - per nota
ESCLUSI = ('Pirelli, Maserati, Peugeot, Citroen, Suzuki, Kia, Mitsubishi, Piaggio, Metzeler, Luna Rossa, '
 'Armani, Fond. Prada, Damiani, Golden Point, Golden Lady, Motivi, Carpisa, Iceberg, Salewa, UPIM, '
 'Pompea, Camomilla, Liabel, Vic Matie, Cafe Noir, North Sails, Arena, Cisalfa, Scarpe e Scarpe, '
 'Coop, Mutti, Garofalo, Pernigotti, Carapelli, Caffarel, Amica Chips, Fileni, Frescobaldi, Pellini, Bofrost, '
 'Samsung, Acer, Lenovo, Euronics, Media World, Cassina, Poltrona Frau, Snaidero, Henkel, Sisal, Martini')

# =====================================================================
# HOME
# =====================================================================
ws = wb.active; ws.title = 'HOME'
title_block(ws,'MASTER EUREWEB / AD LAB',
    'Strumento di lavoro Telesales per generare appuntamenti - shooting foto/video AI per grandi brand - aggiornato 15 giugno 2026')
rows = [
 ('',),
 ('OBIETTIVO: fissare appuntamenti conoscitivi tra AD Lab (unit Eureweb) e i decision maker marketing di grandi brand italiani che investono in pubblicita e contenuti.',),
 ('CLIENTE: Eureweb / AD Lab - referente Giulia Rizzi (g.rizzi@eureweb.com). Servizio: shooting foto/video interamente generati in AI per social, ADV, spot, affissioni.',),
 ('GARANZIA CONTRATTO: minimo 6-10 appuntamenti in target/mese (giugno-luglio, stop agosto). Qualita > quantita: pochi appuntamenti ma buoni.',),
 ('REGOLA FERREA: in chiamata NON si fanno MAI prezzi (li gestisce Giulia in appuntamento). L\'obiettivo della chiamata e SOLO l\'appuntamento.',),
 ('',),
 ('SE SEI UN SETTER: 1) Leggi la tab "GUIDA_SETTER"   2) Lavora dalla tab "LISTA_PROSPECT" partendo dai VERDI (A) in alto.',),
 ('',),
 ('COSA TROVI IN OGNI TAB',),
 ('Tab','A chi serve','Cosa contiene'),
 ('HOME','Tutti','Indice, obiettivo, riepilogo numerico, regole d\'oro in pillole, ultimo sync'),
 ('GUIDA_SETTER','Setter','Guida completa: chi chiami, workflow, script (apertura + varianti), scenari, obiezioni, regole, esiti, tono'),
 ('LISTA_PROSPECT','Setter','I prospect in target ordinati per priorita (A verde, B giallo, C grigio). Colonna "Cosa dire" + Esito/Note/Appuntamento da compilare'),
 ('ICP_e_INSIGHT','Telesales / Giulia','Profilo cliente ideale + insight di mercato (dato verificato -> uso in chiamata)'),
 ('KPI_TRACKING','Tutti','Target appuntamenti, review settimanale giu-lug, prossimi passi'),
 ('DB_COMPLETO','Riferimento','Tutti i prospect + legenda esiti/colori/score'),
 ('',),
 ('RIEPILOGO NUMERICO',),
 (f'Prospect totali in lista ({n_new} nuovi da ricerca, {n_drop} rimossi fuori target)', TOT),
 (f'  VERDE  A - investe Alto E segnale di nuovo investimento ora (di cui {n_qw} "quick win" raggiungibili)', nA),
 ('  GIALLO B - investe Alto, oppure ha un segnale di nuovo investimento', nB),
 ('  GRIGIO C - investe in modo medio/basso senza segnali immediati, da lavorare dopo', nC),
 ('',),
 ('PROSPECT PER SETTORE (perimetro Giulia Rizzi, email 15/6)',),
 ('  Automotive', sect_count.get(AUTO,0)),
 ('  Food & Beverage', sect_count.get(FOOD,0)),
 ('  Retail & GDO', sect_count.get(RET,0)),
 ('  Fashion & Lifestyle', sect_count.get(FASH,0)),
 ('  Finance & Insurance', sect_count.get(FIN,0)),
 ('  Telecomunicazioni', sect_count.get(TLC,0)),
 ('',),
 ('CRITERIO DI SELEZIONE (dal perimetro Giulia): aziende attive in Italia con forte presenza in comunicazione, advertising, contenuti digitali, campagne video, social, branded content.',),
 (f'Ogni azienda e stata VERIFICATA azienda-per-azienda (ricerca web): colonne Investimento / Evidenza / Nuovo investimento. Lavorabilita: {n_giant} enterprise/multinazionali (difficili, tipo Haribo/Maserati), {n_grande} grandi nazionali, {n_mid} mid-large piu facili. Tra i verdi salgono in alto i piu raggiungibili.',),
 ('NOTA: esclusi dallo scouting i 130+ brand gia clienti Eureweb (no doppioni / no conflitto). Es: '+ESCLUSI[:130]+'... (elenco completo in DB_COMPLETO).',),
 ('I clienti gia Eureweb restano un canale "caldo" separato che Giulia puo attivare internamente: non vanno chiamati dai setter.',),
 ('',),
 ('TARGET APPUNTAMENTI',),
 ('  Appuntamenti in target / mese (garanzia contratto)', '6-10'),
 ('',),
 ('REGOLE D\'ORO IN PILLOLE (dettaglio in GUIDA_SETTER)',),
 ('  1. In chiamata NON si fanno MAI prezzi. L\'obiettivo e l\'APPUNTAMENTO conoscitivo.',),
 ('  2. Chiami a nome di "AD Lab, la unit di produzione contenuti AI di Eureweb".',),
 ('  3. Superare il filtro (centralino/segreteria) per arrivare al referente marketing: vedi script sez. 5.',),
 ('  4. Lista da far approvare a Giulia PRIMA di partire con le chiamate.',),
 ('  5. Meglio pochi appuntamenti ottimi e in target che tanti mediocri.',),
 ('',),
 ('ULTIMO SYNC - 15 GIUGNO 2026 (call onboarding 11/6 + email perimetro settori Giulia Rizzi 15/6)',),
]
r = 3
for tup in rows:
    for j,v in enumerate(tup):
        ws.cell(row=r, column=1+j, value=v)
    r += 1
fill_row_title(ws, 3)
# formattazioni
for label_row in (5,11,20,26,34,40,45):
    ws.cell(row=label_row,column=1).font = F_B
# header tabella tab
ws.cell(row=12,column=1).fill=C_HDR; ws.cell(row=12,column=1).font=F_HDRW
for c in range(1,4):
    ws.cell(row=12,column=c).fill=C_HDR; ws.cell(row=12,column=c).font=F_HDRW
widths={'A':74,'B':22,'C':70}
for k,v in widths.items(): ws.column_dimensions[k].width=v
for rr in range(3, r):
    ws.cell(row=rr,column=1).alignment = WRAP

# =====================================================================
# GUIDA_SETTER
# =====================================================================
ws = wb.create_sheet('GUIDA_SETTER')
title_block(ws,'GUIDA SETTER - AD LAB (Eureweb)',
    'Tutto quello che serve per chiamare grandi brand e fissare appuntamenti conoscitivi per gli shooting foto/video AI')
G = [
 ('',),
 ('REGOLA NUMERO 1: l\'obiettivo della chiamata e SOLO fissare un appuntamento conoscitivo con il referente marketing. In chiamata NON si fanno mai prezzi: li gestisce Giulia in appuntamento.',True),
 ('',),
 ('1. CHI STAI CHIAMANDO',True),
 ('Grandi aziende italiane (o filiali italiane di gruppi internazionali) che investono molto in pubblicita e produzione contenuti, nei 6 settori: Automotive, Food & Beverage, Retail & GDO, Fashion & Lifestyle, Finance & Insurance, Telecomunicazioni.',),
 ('Cosa vendiamo: AD Lab e la unit di Eureweb che realizza shooting foto e video INTERAMENTE generati con intelligenza artificiale. Stessa qualita (o superiore) di una produzione tradizionale, con tempi e costi ridotti. I contenuti si usano su social, campagne ADV, spot TV, affissioni: tutti i touchpoint.',),
 ('Perche interessa: questi brand producono contenuti di continuo (lanci, stagioni, promo, social always-on). Noi riduciamo tempi e costi mantenendo un livello alto. Esempi di lavori: hotellerie di lusso, fashion, automotive (mostrati da Giulia in appuntamento).',),
 ('Chi e il referente: responsabile/direttore marketing o comunicazione; in alternativa brand manager, head of content/digital. Nelle aziende molto grandi puntiamo all\'entita ITALIANA del gruppo.',),
 ('',),
 ('2. WORKFLOW GIORNALIERO',True),
 ('Quando','Cosa fare'),
 ('Mattina inizio','Apri LISTA_PROSPECT. Parti dal primo VERDE (A) in alto. Verifica che la riga non abbia gia un Esito.'),
 ('Prima della chiamata','Leggi "Perche in target" e "Cosa dire": sai gia il gancio. Dai un occhio veloce al sito/social del brand (vedi come comunicano).'),
 ('Durante','Usa lo script (sez. 4-5). Non leggerlo a memoria. Primo obiettivo: superare il filtro e arrivare al marketing.'),
 ('Subito dopo','Aggiorna ESITO, NOTE, e se fissi: APPUNTAMENTO (data/ora). Non aspettare fine giornata.'),
 ('Fissato un appuntamento','Inseriscilo SUBITO sul calendario condiviso e avvisa Giulia. Niente prezzi promessi al cliente.'),
 ('Fine giornata','Mini-report (5 righe): chiamate, aziende raggiunte, referenti agganciati, appuntamenti fissati.'),
 ('Settimanale','Review: porta i 2-3 prospect piu interessanti e gli eventuali "non in target" da segnalare a Giulia.'),
 ('',),
 ('3. COSA SIGNIFICA OGNI COLORE',True),
 ('Colore','Significato'),
 ('VERDE - priorita A','Chiama QUESTI PRIMA. Gia investono molto in adv/contenuti E hanno un segnale di nuovo investimento ora (rebrand, nuova linea, nuovo testimonial/CMO, espansione): sono i piu caldi. Vedi colonne Evidenza e Nuovo investimento.'),
 ('GIALLO - priorita B','Dopo i verdi. Investono molto in comunicazione, oppure hanno un segnale di nuovo investimento. Buon potenziale.'),
 ('GRIGIO - priorita C','Solo dopo verdi e gialli. Investono in modo medio/basso senza segnali immediati. Lavorali con efficienza, spesso meglio via LinkedIn/email.'),
 ('',),
 ('4. SCRIPT - struttura della chiamata',True),
 ('4.A - Apertura con il centralino/filtro',True),
 ('"Buongiorno, sono [NOME] di AD Lab, la unit di produzione contenuti di Eureweb. Vorrei parlare con chi si occupa di marketing e comunicazione: produzione contenuti e campagne. Me lo passa, per favore?"',),
 ('-> Se chiede di cosa si tratta: "Realizziamo shooting foto e video in AI per i brand. Volevo solo capire chi segue questa parte e fissare un confronto conoscitivo."',),
 ('4.B - Aggancio con il referente marketing',True),
 ('"Buongiorno [NOME], la chiamo da AD Lab, unit di Eureweb. Realizziamo shooting foto e video interamente in AI per brand come il vostro: stessa qualita di una produzione tradizionale, con tempi e costi piu bassi. Come gestite oggi la produzione dei vostri contenuti per campagne e social?"',),
 ('-> Lascia parlare. Capisci se producono internamente, con agenzie, con quale frequenza. Annota.',),
 ('4.C - Chiusura appuntamento (obiettivo della chiamata)',True),
 ('"Le propongo un confronto conoscitivo di 20-30 minuti: la nostra referente vi mostra alcuni case reali (hotellerie di lusso, fashion, automotive) e capiamo insieme se ha senso per voi. Le va meglio [giorno X] o [giorno Y]?"',),
 ('-> Proponi 2 opzioni concrete. Fissa data e ora. Niente prezzi, niente preventivi al telefono.',),
 ('-> Inserisci subito in calendario e avvisa Giulia.',),
 ('',),
 ('5. SUPERARE IL FILTRO (centralino/segreteria) - tecniche',True),
 ('Tono sicuro','Parla come un pari, non come chi "vende". "Mi passa il marketing, grazie" funziona meglio di "potrei eventualmente parlare con...".'),
 ('Chiedi il nome','"Chi segue la comunicazione/contenuti da voi?" Annotalo: alla prossima chiamata chiedi direttamente di lui/lei.'),
 ('Email mirata','Se non ti passano: "Mi puo dare la mail diretta del responsabile marketing? Mando due righe e un case." Poi richiama.'),
 ('LinkedIn','Per B e C spesso si arriva prima via LinkedIn al referente (vedi colonna LinkedIn). Connetti + messaggio breve, poi chiamata.'),
 ('Orari','Centralini piu sgombri 9:00-9:30 e 14:00-15:00. Evita lunedi mattina e venerdi pomeriggio.'),
 ('',),
 ('6. GESTIONE OBIEZIONI',True),
 ('Obiezione','Cosa rispondere'),
 ('"Di cosa si tratta?"','"Realizziamo shooting foto/video interamente in AI per i brand. Volevo capire chi segue i contenuti e proporre un breve confronto conoscitivo."'),
 ('"Mandate una mail"','"Volentieri, a chi la indirizzo nello specifico? La mando e la richiamo io tra qualche giorno per un riscontro." Poi richiama davvero.'),
 ('"Lavoriamo gia con un\'agenzia"','"Perfetto, non sostituiamo nessuno: siamo un\'opzione in piu sulla produzione visiva, piu veloce e flessibile. Le va di vedere alcuni case in 20 minuti?"'),
 ('"Quanto costa?"','"Dipende dal progetto, e proprio per questo serve un confronto: la referente le fa vedere i case e poi costruisce una proposta su misura. Le fisso l\'appuntamento?"'),
 ('"Non abbiamo budget ora"','"Capisco. L\'incontro e conoscitivo, cosi quando partira un progetto ci avete gia visti. Le va comunque un breve confronto?"'),
 ('"Non ci interessa"','"Capito, grazie. Le lascio il riferimento, se in futuro vi servisse produzione contenuti ci siamo." Esito = Non interessato.'),
 ('',),
 ('7. REGOLE D\'ORO (sempre)',True),
 ('1. MAI prezzi al telefono.','Li gestisce Giulia in appuntamento, sempre su misura.'),
 ('2. L\'obiettivo e l\'APPUNTAMENTO conoscitivo.','Non vendere il servizio in chiamata: vendi l\'incontro.'),
 ('3. Punta al referente marketing.','Supera il filtro, prendi il nome, arriva alla persona giusta.'),
 ('4. Entita italiana per i gruppi grandi.','Per multinazionali, cerca il marketing della filiale italiana.'),
 ('5. Non bruciare il brand.','Tono curato e professionale: sono aziende importanti, l\'immagine conta.'),
 ('6. Segnala i non in target.','Se un\'azienda non e adatta, segnalala: non si conteggia ed evitiamo di perdere tempo.'),
 ('',),
 ('8. ESITI - significato (usa il menu a tendina, non testo libero)',True),
 ('Da contattare','Stato iniziale.'),
 ('Non risposto','Hai chiamato e NESSUNO ha risposto. Riprova entro 24-48h. (Solo se nessuno risponde.)'),
 ('Da richiamare','Hai parlato col centralino/segreteria o non era il momento. Metti data prossimo contatto.'),
 ('Referente individuato','Hai il nome/contatto del referente marketing ma non ancora agganciato. Annota nome.'),
 ('Interessato','Il referente ha mostrato interesse, in attesa di fissare. Da chiudere.'),
 ('Appuntamento fissato','Hai fissato il confronto conoscitivo e l\'hai messo in calendario. (Questo e il "win".)'),
 ('Da ricontattare piu avanti','Interesse ma non ora (es. budget gia speso/campagna fatta). Annota quando.'),
 ('Non in target','Non adatto (troppo piccolo, no budget contenuti). Segnala a Giulia, non si conteggia.'),
 ('Non interessato','Ha detto chiaro di no. Non insistere.'),
 ('Numero/contatto errato','Recapito sbagliato. Annota in note.'),
 ('',),
 ('9. NOTE - come scriverle',True),
 ('Brevi, professionali, come le scriverebbe un commerciale. Riassumi: con chi hai parlato (ruolo), come producono oggi i contenuti, prossimo passo. Mai dettagli tecnici interni, mai gergo. Es: "Parlato con assistente marketing, referente e la Brand Manager (Sara). Producono con agenzia esterna. Mando mail + case, richiamo lun 22."',),
 ('',),
 ('10. TONO DI VOCE',True),
 ('Sicuro','Parli con grandi brand: pari grado, mai supplicante.'),
 ('Curioso','Sei li per capire come producono i contenuti, non per "vendere" subito.'),
 ('Concreto','Frasi corte. "Le va meglio giovedi o lunedi?" meglio di lunghi giri di parole.'),
 ('Professionale','Aziende importanti: cura linguaggio e immagine. Niente fretta, niente pressione.'),
 ('',),
 ('11. FINE GIORNATA - mini-report (5 righe)',True),
 ('Esempio: "AD Lab - lun 16 giu | 32 chiamate, 14 aziende raggiunte | 5 referenti individuati | 2 appuntamenti (mer 18 h15 Brand Mng X, gio 19 h11 Mktg Dir Y) | 1 da ricontattare a settembre. Domani continuo dai verdi food."',),
]
r = 3
for tup in G:
    text = tup[0]; bold = len(tup) > 1 and tup[1] is True
    second = tup[1] if (len(tup) > 1 and tup[1] is not True) else None
    ws.cell(row=r, column=1, value=text)
    if bold: ws.cell(row=r, column=1).font = F_B
    if second is not None:
        ws.cell(row=r, column=2, value=second)
    ws.cell(row=r, column=1).alignment = WRAP
    ws.cell(row=r, column=2).alignment = WRAP
    r += 1
fill_row_title(ws, 2)
ws.column_dimensions['A'].width = 62
ws.column_dimensions['B'].width = 80

# =====================================================================
# LISTA_PROSPECT
# =====================================================================
ws = wb.create_sheet('LISTA_PROSPECT')
cols = ['#','Azienda','Prio','Quick win','Settore','Dimensione','Accessibilita DM','Sede','Sito','Telefono centralino','Cellulare (diretto)',
        'Decision Maker (ruolo)','LinkedIn (ricerca)','Investimento','Evidenza investimento adv/contenuti','Nuovo investimento',
        'Perche in target','Cosa dire (gancio)','Canale approccio','Esito','Note','Appuntamento (data/ora)',
        'Titolare / Legale rappresentante','Responsabile acquisti','Cellulare titolare/acquisti',
        'DM marketing (nome)','DM marketing (ruolo)','DM marketing LinkedIn','DM email (se pubblica)','DM cellulare (se pubblico)']
title_block(ws,'LISTA_PROSPECT - la tua lista di lavoro',
    f'{nA} verdi (A) - {nB} gialli (B) - {nC} grigi (C) su {TOT} prospect ({n_new} nuovi). {n_tel} centralini verificati. Cellulare diretto: solo dove pubblico/da enrichment. In chiamata NON si fanno prezzi.')
ws['A3']='Parti dai VERDI in alto. Telefono centralino = numero ufficiale verificato (si chiama e si chiede il marketing). Cellulare (diretto) = vuoto dove non disponibile in modo affidabile (mai inventato).'
ws['A3'].font=F_SUB
hdr_row=4
for j,c in enumerate(cols,1):
    cell=ws.cell(row=hdr_row,column=j,value=c); cell.fill=C_HDR; cell.font=F_HDRW; cell.alignment=WRAP
fillmap={'A':C_A,'B':C_B,'C':C_C}
r=hdr_row+1
for i,x in enumerate(recs,1):
    vals=[i, x['az'], x['prio'], x['quickwin'], x['se'], x['dim'], x['acc'], x['sede'], x['sito'], x.get('tel',''), x.get('cell',''),
          x['dm'], x['li'], x['inv'], x['evid'], x['nuovo'],
          x['perche'], x['hook'], x['canale'], 'Da contattare', '', '',
          x.get('titolare',''), x.get('acquisti',''), x.get('cellpers',''),
          x.get('dm_nome',''), x.get('dm_ruolo',''), x.get('dm_li',''), x.get('dm_email',''), x.get('dm_cell','')]
    for j,v in enumerate(vals,1):
        cell=ws.cell(row=r,column=j,value=v)
        cell.fill=fillmap[x['prio']]
        cell.alignment = WRAP if j in (13,15,17,18) else TOP
    ws.cell(row=r,column=10).number_format='@'  # telefono centralino come testo (mantiene lo zero)
    ws.cell(row=r,column=11).number_format='@'  # cellulare come testo
    ws.cell(row=r,column=25).number_format='@'  # cellulare titolare/acquisti come testo
    ws.cell(row=r,column=30).number_format='@'  # cellulare DM marketing come testo
    r+=1
last=r-1
esiti='"Da contattare,Non risposto,Da richiamare,Referente individuato,Interessato,Appuntamento fissato,Da ricontattare piu avanti,Non in target,Non interessato,Numero/contatto errato"'
dv=DataValidation(type='list',formula1=esiti,allow_blank=True)
ws.add_data_validation(dv); dv.add(f'T{hdr_row+1}:T{last}')
prio_dv=DataValidation(type='list',formula1='"A,B,C"',allow_blank=True)
ws.add_data_validation(prio_dv); prio_dv.add(f'C{hdr_row+1}:C{last}')
inv_dv=DataValidation(type='list',formula1='"Alto,Medio,Basso"',allow_blank=True)
ws.add_data_validation(inv_dv); inv_dv.add(f'N{hdr_row+1}:N{last}')
acc_dv=DataValidation(type='list',formula1='"Difficile,Media,Buona"',allow_blank=True)
ws.add_data_validation(acc_dv); acc_dv.add(f'G{hdr_row+1}:G{last}')
widths={'A':5,'B':28,'C':5,'D':10,'E':19,'F':22,'G':14,'H':22,'I':22,'J':18,'K':18,'L':24,'M':36,'N':12,'O':58,'P':30,'Q':36,'R':44,'S':24,'T':22,'U':28,'V':18,'W':30,'X':26,'Y':20,'Z':26,'AA':30,'AB':40,'AC':30,'AD':22}
for k,v in widths.items(): ws.column_dimensions[k].width=v
ws.freeze_panes='A5'
ws.auto_filter.ref=f'A{hdr_row}:AD{last}'

# =====================================================================
# ICP_e_INSIGHT   (insight di mercato: aggiornati dopo ricerca, qui base verificabile)
# =====================================================================
ws=wb.create_sheet('ICP_e_INSIGHT')
title_block(ws,'ICP e INSIGHT - mercato contenuti AI per grandi brand',
    'Profilo del cliente ideale + sintesi ricerca di mercato (dettaglio in Dossier_Mercato_Eureweb.pdf)')
r=4
ws.cell(row=r,column=1,value='PROFILO CLIENTE IDEALE (azienda da prenotare in appuntamento)').font=F_B; r+=1
for c,t in enumerate(['Parametro','Valore'],1):
    cell=ws.cell(row=r,column=c,value=t); cell.fill=C_HDR; cell.font=F_HDRW
r+=1
icp=[
 ('Tipo azienda','Grande brand o filiale italiana di gruppo, forte investitore in pubblicita e produzione contenuti'),
 ('Settori (perimetro Giulia)','Automotive, Food & Beverage, Retail & GDO, Fashion & Lifestyle, Finance & Insurance, Telecomunicazioni'),
 ('Segnale piu forte','Gia attiva con campagne video/social ad alta frequenza, shooting ricorrenti, budget contenuti allocato'),
 ('Referente da agganciare','Direttore/Responsabile Marketing o Comunicazione; in alternativa Brand Manager, Head of Content/Digital'),
 ('Sblocca-budget','C-level (CEO/CFO/CMO): spesso non fa la call ma autorizza il budget'),
 ('Per i gruppi internazionali','Puntare alla entita ITALIANA (es. la filiale italiana del gruppo) col proprio marketing e budget'),
 ('Da escludere / depriorizzare','PMI con budget contenuti basso; aziende senza attivita di comunicazione strutturata; clienti gia Eureweb'),
 ('Dato utile pre-call','Quanto investe in ADV (Meta Ad Library, campagne attive): segnale di budget e di attivita'),
 ('Valore di riferimento (interno, NON in call)','Servizio AD Lab da 15.000 EUR a salire; benchmark produzione video tradizionale citato ~50.000 EUR'),
]
for p,v in icp:
    ws.cell(row=r,column=1,value=p); c=ws.cell(row=r,column=2,value=v); c.alignment=WRAP; r+=1
r+=2
ws.cell(row=r,column=1,value='INSIGHT DI MERCATO -> uso in chiamata (in aggiornamento dalla ricerca, vedi Dossier)').font=F_B; r+=1
for c,t in enumerate(['Area','Insight chiave','Come usarlo in chiamata'],1):
    cell=ws.cell(row=r,column=c,value=t); cell.fill=C_HDR; cell.font=F_HDRW; cell.alignment=WRAP
r+=1
ins=[
 ('Video in crescita','Il video advertising in Italia vale ~2,9 mld EUR nel 2025 (+16%, 41% del digitale): e il formato che cresce di piu. (Osservatorio PoliMi)','"Il video e il formato che cresce di piu nel mercato pubblicitario: volevo capire come gestite oggi la produzione dei vostri contenuti video."'),
 ('AI gia in casa','L\'84% delle grandi imprese italiane ha gia almeno una licenza di AI generativa. (Osservatorio AI PoliMi)','"So che realta come la vostra usano gia l\'AI internamente: noi la applichiamo alla produzione creativa per i brand. Confrontiamoci."'),
 ('AI nel video adv','Nel 2025 il 30% degli spot video nel mondo e creato o migliorato con AI generativa (era 22% nel 2024). (IAB)','"Quasi un terzo del video advertising globale passa ormai per l\'AI: vorrei mostrarvi cosa significa per un brand del vostro livello."'),
 ('I big lo fanno','Coca-Cola ha generato 70.000 videoclip con AI per un solo spot. (Mediatrends)','"I grandi brand producono migliaia di varianti con l\'AI: il punto non e \'se\' ma \'come farlo bene\'. Ve lo mostriamo in incontro?"'),
 ('Settori in crescita','Food & beverage e finance/insurance aumentano gli investimenti adv (finance +11,7% nel 2025). (Nielsen)','"Il vostro settore e tra i pochi che stanno aumentando gli investimenti in comunicazione: e il momento giusto per fare di piu col budget."'),
 ('Settori sotto pressione','Automotive (2025 -5%), retail/GDO (-10%) e telco (-10%) tagliano i budget adv. (Nielsen)','"So che c\'e massima attenzione ai costi: proprio per questo volevo presentarvi un modo per produrre contenuti di alto livello in meno tempo."'),
 ('Risparmio tempi/costi','La produzione AI riduce in modo sostanziale tempi e costi rispetto allo shooting tradizionale (stime di settore). Produzione spot di alta fascia: decine di migliaia di EUR.','"Senza numeri al telefono perche dipendono dal progetto: la differenza in tempi e flessibilita e enorme. Preferisco mostrarvelo dal vivo."'),
 ('Chi decide','Decide il marketing (CMO/Direttore/Head of Content); il procurement entra dopo sulle condizioni. (analisi B2B)','"Per non farvi perdere tempo: di solito la persona giusta per una panoramica e chi segue marketing e contenuti. Parlo con la persona corretta?"'),
]
for a,b,c in ins:
    ws.cell(row=r,column=1,value=a)
    x=ws.cell(row=r,column=2,value=b); x.alignment=WRAP
    y=ws.cell(row=r,column=3,value=c); y.alignment=WRAP
    r+=1
fill_row_title(ws,3)
ws.column_dimensions['A'].width=22; ws.column_dimensions['B'].width=58; ws.column_dimensions['C'].width=58

# =====================================================================
# KPI_TRACKING
# =====================================================================
ws=wb.create_sheet('KPI_TRACKING')
title_block(ws,'KPI_TRACKING - target + review settimanale','Campagna giugno-luglio 2026 (stop agosto). Primo elenco prospect da far approvare a Giulia.')
r=4
ws.cell(row=r,column=1,value='TARGET').font=F_B; r+=1
for c,t in enumerate(['Obiettivo','Target','Attuale','Note'],1):
    cell=ws.cell(row=r,column=c,value=t); cell.fill=C_HDR; cell.font=F_HDRW
r+=1
for o,t,a,n in [('Appuntamenti in target fissati / mese','6-10',0,'Garanzia contratto; Giulia se ne aspetta di piu'),
                ('Referenti marketing individuati','-',0,'Nome + contatto del decision maker'),
                ('Prospect approvati da Giulia','-',0,'Lista da validare PRIMA delle chiamate')]:
    ws.cell(row=r,column=1,value=o);ws.cell(row=r,column=2,value=t);ws.cell(row=r,column=3,value=a);ws.cell(row=r,column=4,value=n);r+=1
r+=1
ws.cell(row=r,column=1,value='REVIEW SETTIMANALE (giu-lug 2026)').font=F_B; r+=1
for c,t in enumerate(['Settimana','Date','N. chiamate','Aziende raggiunte','Referenti individuati','Appuntamenti fissati'],1):
    cell=ws.cell(row=r,column=c,value=t); cell.fill=C_HDR; cell.font=F_HDRW
r+=1
for w,d in [('W1','16-20 giu'),('W2','23-27 giu'),('W3','30 giu-4 lug'),('W4','7-11 lug'),
            ('W5','14-18 lug'),('W6','21-25 lug'),('W7','28 lug-1 ago')]:
    ws.cell(row=r,column=1,value=w);ws.cell(row=r,column=2,value=d);r+=1
r+=1
ws.cell(row=r,column=1,value='PROSSIMI PASSI').font=F_B; r+=1
for s in ['1. Inviare a Giulia il primo elenco prospect (questa lista) per approvazione PRIMA di chiamare.',
          '2. Recuperare in fase scraping: numero diretto e nome del referente marketing per i prospect A (verdi).',
          '3. Preparare/condividere lo script v1 (questa guida) e raccogliere da Giulia i suoi script per A/B test.',
          '4. Arricchire la colonna "Telefono" e "Decision Maker (nome)" partendo dai verdi.',
          '5. Verificare per i prospect A la presenza di campagne attive (Meta Ad Library) come segnale di budget.',
          '6. Report fine mese del lavoro svolto (aziende contattate + esiti) a Giulia.']:
    ws.cell(row=r,column=1,value=s); r+=1
fill_row_title(ws,6)
for k,v in {'A':46,'B':16,'C':14,'D':18,'E':22,'F':20}.items(): ws.column_dimensions[k].width=v

# =====================================================================
# DB_COMPLETO
# =====================================================================
ws=wb.create_sheet('DB_COMPLETO')
title_block(ws,'DB COMPLETO - tutti i prospect (backup) + legenda',
    f'{TOT} prospect in target ({n_new} nuovi da ricerca; {n_drop} rimossi fuori target). Perimetro 6 settori (email Giulia 15/6). Tutti verificati per investimento adv/contenuti.')
cols=['#','Prio','Azienda','Settore','Dimensione','Accessibilita DM','Sede','Sito','Telefono centralino','Cellulare (diretto)','Investimento','Nuovo investimento','Evidenza investimento adv/contenuti','Fonte telefono','Origine','Score','Titolare / Legale rappresentante','Responsabile acquisti','Cellulare titolare/acquisti','DM marketing (nome)','DM marketing (ruolo)','DM marketing LinkedIn','DM email (se pubblica)','DM cellulare (se pubblico)']
hdr=4
for j,c in enumerate(cols,1):
    cell=ws.cell(row=hdr,column=j,value=c); cell.fill=C_HDR; cell.font=F_HDRW; cell.alignment=WRAP
r=hdr+1
for i,x in enumerate(recs,1):
    vals=[i,x['prio'],x['az'],x['se'],x['dim'],x['acc'],x['sede'],x['sito'],x.get('tel',''),x.get('cell',''),x['inv'],x['nuovo'],x['evid'],x.get('telf',''),x['origine'],x['score'],x.get('titolare',''),x.get('acquisti',''),x.get('cellpers',''),x.get('dm_nome',''),x.get('dm_ruolo',''),x.get('dm_li',''),x.get('dm_email',''),x.get('dm_cell','')]
    for j,v in enumerate(vals,1):
        cell=ws.cell(row=r,column=j,value=v)
        cell.fill=fillmap[x['prio']]
        cell.alignment = WRAP if j in (12,13) else TOP
    ws.cell(row=r,column=9).number_format='@'   # telefono centralino come testo
    ws.cell(row=r,column=10).number_format='@'  # cellulare come testo
    ws.cell(row=r,column=19).number_format='@'  # cellulare titolare/acquisti come testo
    ws.cell(row=r,column=24).number_format='@'  # cellulare DM come testo
    r+=1
dblast=r-1
ws.freeze_panes='A5'; ws.auto_filter.ref=f'A{hdr}:X{dblast}'
for k,v in {'A':5,'B':6,'C':28,'D':19,'E':24,'F':14,'G':22,'H':22,'I':18,'J':18,'K':12,'L':30,'M':62,'N':40,'O':20,'P':7,'Q':30,'R':26,'S':20,'T':26,'U':30,'V':40,'W':30,'X':22}.items():
    ws.column_dimensions[k].width=v
# legenda
r=dblast+3
ws.cell(row=r,column=1,value='LEGENDA').font=F_B; r+=1
for k,v in [('Prio A (verde)','Grande advertiser molto attivo in contenuti/video/social - chiama subito'),
            ('Prio B (giallo)','Brand mid-large premium attivo in comunicazione'),
            ('Prio C (grigio)','Realta piu piccola ma con budget contenuti - dopo, spesso via LinkedIn/email'),
            ('Quick win','Azienda lavorabile: accessibilita Buona (DM raggiungibile) E investe davvero (Alto/Medio) in adv/contenuti. I primi da chiamare per fare appuntamenti in fretta.'),
            ('Dimensione','Enterprise/multinazionale (giganti tipo Haribo/Maserati) | Grande nazionale | Mid-large (premium/nicchia). Stima per fascia, non fatturato puntuale.'),
            ('Accessibilita DM','Quanto e facile arrivare al decisore e fissare: Difficile (procurement/gare, ciclo lungo) | Media | Buona (DM raggiungibile). Inversamente legata alla dimensione.'),
            ('Investimento','Livello stimato di spesa in adv/contenuti (Alto/Medio/Basso) da ricerca web per azienda'),
            ('Evidenza investimento adv/contenuti','Prova concreta che gia investe (spot, campagne, social/video, shooting) con anno/segnale - dato del criterio Giulia'),
            ('Nuovo investimento','Segnale che sta per investire/aumentare ora (rebrand, nuova linea, nuovo testimonial/CMO, espansione): aggancio caldo'),
            ('Origine','In lista iniziale = scouting primo giro; Nuovo lead (ricerca) = aggiunto dopo verifica per-azienda'),
            ('Score','Punteggio priorita = base priorita (A 90 / B 75 / C 60) + intensita contenuti del settore (Fashion +4, Retail/Food/Auto +3, Telco +2, Finance +1)'),
            ('Decision Maker (ruolo)','RUOLO da agganciare; il NOME si reperisce in fase scraping (anti-errore: non si inventano nomi)'),
            ('Telefono centralino','Numero UFFICIALE della sede italiana VERIFICATO da fonte ufficiale (pagina Contatti del sito, registro imprese, elenco verificato). Vuoto = non verificabile, da non inventare. Formato nazionale senza +39. Si chiama e si chiede il marketing.'),
            ('Cellulare (diretto)','Cellulare/linea diretta del decisore: NON e un dato pubblico per le grandi aziende. Si popola solo da provider B2B a pagamento o dove pubblicato su fonte ufficiale. Vuoto = non disponibile in modo affidabile (mai inventato).'),
            ('Titolare / Legale rappresentante','Nome e cognome del titolare/legale rappresentante/AD verificato da fonte pubblica (registro imprese, sito, LinkedIn). Vuoto = non verificato.'),
            ('Responsabile acquisti','Nome e cognome del responsabile acquisti SOLO se pubblicamente verificabile. Per AD Lab il vero compratore e il marketing: per le grandi aziende questo campo e spesso vuoto.'),
            ('Cellulare titolare/acquisti','Cellulare di titolare/acquisti SOLO se pubblicato su fonte ufficiale. Quasi sempre vuoto (non e un dato pubblico).'),
            ('Fonte telefono','URL della fonte ufficiale da cui e stato letto il numero (tracciabilita anti-errore).'),
            ('Esclusi','Clienti gia Eureweb (dal sito /clienti): non in lista per evitare doppioni e conflitti. Restano canale caldo interno di Giulia.')]:
    ws.cell(row=r,column=1,value=k).font=F_B; c=ws.cell(row=r,column=3,value=v); c.alignment=WRAP; r+=1
fill_row_title(ws,len(cols))

out='/Users/simocors/Desktop/telesales/eureweb/Master_Eureweb_ADLab.xlsx'
wb.save(out)
print('SALVATO:',out)
print('Tab:',wb.sheetnames)
print('Prospect totali:',TOT,'| A:',nA,'B:',nB,'C:',nC)
print('Nuovi lead: ricerca1=',n_new1,'ricerca2(mid-large)=',n_new2,'tot=',n_new,'| Rimossi fuori target:',n_drop,dropped)
print('Dimensione:',n_giant,'enterprise /',n_grande,'grandi /',n_mid,'mid-large | Quick win:',n_qw)
print('Telefoni verificati:',n_tel,'/',TOT,'| cellulari diretti:',n_cell)
print('Titolari:',n_tit,'| Resp. acquisti:',n_acq)
print('DM marketing:',n_dm,'| con LinkedIn:',n_dmli,'| con email pubblica:',n_dmem)
print('Verificate non matchate (rimaste generiche):',len(unmatched),unmatched)
print('Per settore:',sect_count)
