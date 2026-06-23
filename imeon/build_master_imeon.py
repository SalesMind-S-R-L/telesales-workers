#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Master IMEON - L'agenzia alle torri. Ispirato al master Giglioli."""
import pickle, re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

ded = pickle.load(open('/tmp/imeon_final.pkl', 'rb'))

# ---- colori (stessi del master Giglioli) ----
C_TITLE = PatternFill('solid', fgColor='1F3864')
C_HDR   = PatternFill('solid', fgColor='1F3864')
C_A     = PatternFill('solid', fgColor='C8E6C9')  # verde
C_B     = PatternFill('solid', fgColor='FFF59D')  # giallo
C_C     = PatternFill('solid', fgColor='F2F2F2')  # grigio
C_SEC   = PatternFill('solid', fgColor='D9E1F2')  # azzurro sezione
F_TITLE = Font(color='FFFFFF', bold=True, size=14)
F_HDRW  = Font(color='FFFFFF', bold=True)
F_SUB   = Font(italic=True, size=10, color='555555')
F_B     = Font(bold=True)
WRAP    = Alignment(wrap_text=True, vertical='top')
TOP     = Alignment(vertical='top')
thin    = Side(style='thin', color='D0D0D0')
BORD    = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = openpyxl.Workbook()

def title_block(ws, title, sub):
    ws['A1'] = title; ws['A1'].font = F_TITLE; ws['A1'].fill = C_TITLE
    ws['A2'] = sub;   ws['A2'].font = F_SUB

def fill_row_title(ws, row, ncol):
    for c in range(1, ncol+1):
        ws.cell(row=1, column=c).fill = C_TITLE

# ============================================================ HOME
ws = wb.active; ws.title = 'HOME'
title_block(ws, 'MASTER IMEON - L\'AGENZIA ALLE TORRI', 'Strumento di lavoro Telesales per l\'acquisizione immobili - aggiornato 12 giugno 2026')
rows = [
 ('',),
 ('OBIETTIVO: acquisire immobili da vendere. Il cliente (Diego) vende tutto cio che acquisisce; il collo di bottiglia e l\'ACQUISIZIONE.',),
 ('SE SEI UN SETTER: 1) Leggi la tab "GUIDA_SETTER"   2) Lavora dalla tab "SETTER_CHIAMATE" partendo dai VERDI in alto.',),
 ('',),
 ('COSA TROVI IN OGNI TAB',),
 ('Tab','A chi serve','Cosa contiene','Quando usarla'),
 ('HOME','Tutti','Indice, riepilogo numerico, regole d\'oro in pillole','Inizio sessione'),
 ('GUIDA_SETTER','Setter','Guida completa: chi chiami, workflow, script (approccio indiretto), 4 scenari, obiezioni, regole, esiti, tono','Leggi PRIMA volta, tieni a portata'),
 ('SETTER_CHIAMATE','Setter','2.490 contatti chiamabili ordinati per priorita (A verde, B giallo, C grigio). Colonna "Cosa dire" + Esito/Note/Appuntamento da compilare','Operativo quotidiano'),
 ('ICP_e_INSIGHT','Telesales','Profilo target acquisizione + 10 insight dalla ricerca di mercato (prezzi, domanda, leve)','Strategia e taratura script'),
 ('KPI_TRACKING','Tutti','Target appuntamenti, review settimanale giu-ago, azioni immediate','Ogni venerdi'),
 ('DB_COMPLETO','Riferimento','Tutti i 2.603 contatti (incl. 113 senza numero) + legenda esiti','Backup e riferimento'),
 ('',),
 ('RIEPILOGO NUMERICO',),
 ('Contatti totali (uniti i 2 file, dedup per cellulare)', 2603),
 ('  - con cellulare (chiamabili subito)', 2482),
 ('  - solo telefono fisso', 8),
 ('  - senza alcun numero (solo DB, per ora si saltano)', 113),
 ('',),
 ('PRIORITA (solo chiamabili)',),
 ('  VERDE  A - caldi: Acquisto recente (2024-25), in zona, segnale di vendita (permuta/prima casa)', 700),
 ('  GIALLO B - buoni: altri Acquisto recenti / in zona', 920),
 ('  GRIGIO C - dopo: locazione, acquisto vecchi, comuni lontani, solo-fisso', 870),
 ('',),
 ('SORGENTE DATI',),
 ('  File "Richieste con residenza" 2025-06-27 (ufficio attivo)', 1514),
 ('  File "Richieste con residenza" 2025-01-11 (ufficio precedente)', 1677),
 ('  Duplicati rimossi (stesso cellulare, tenuto il piu recente)', 588),
 ('',),
 ('TARGET (da confermare con Diego)',),
 ('  Appuntamenti di valutazione fissati / settimana', '> da definire'),
 ('',),
 ('REGOLE D\'ORO IN PILLOLE (dettaglio in GUIDA_SETTER)',),
 ('  1. NON chiedere diretto "vuole vendere casa?" - parti dalla casa cercata, poi chiedi "mutuo o appartamento da vendere?". Al cliente: sempre "valutazione", mai "acquisizione".',),
 ('  2. L\'obiettivo della chiamata e l\'APPUNTAMENTO DI VALUTAZIONE. Si fissa "a prescindere" sul Google Calendar di Diego.',),
 ('  3. NON inviare messaggi di conferma appuntamento (rischio cancellazione). Solo reminder automatico calendario, se attivo.',),
 ('  4. Chiamare a nome de "L\'agenzia alle torri".',),
 ('  5. NON bruciare i contatti: anche un no oggi e un si tra qualche anno (ciclo casa 5-8 anni). Chiamata sempre fatta bene.',),
 ('',),
 ('ULTIMO SYNC - 12 GIUGNO 2026 (call onboarding con Diego)',),
]
r = 3
for tup in rows:
    for j, v in enumerate(tup):
        ws.cell(row=r, column=1+j, value=v)
    r += 1
fill_row_title(ws, 1, 4)
# format
ws['A8'].font = F_B
for rr in [9]:  # header of table
    for c in range(1,5):
        ws.cell(row=rr,column=c).fill=C_HDR; ws.cell(row=rr,column=c).font=F_HDRW
for label_row in (5,8,17,23,28,33,36):
    ws.cell(row=label_row,column=1).font = F_B
widths={'A':62,'B':24,'C':70,'D':30}
for k,v in widths.items(): ws.column_dimensions[k].width=v
ws['A6'].font=F_B

# ============================================================ GUIDA_SETTER
ws = wb.create_sheet('GUIDA_SETTER')
title_block(ws,'GUIDA SETTER','Tutto quello che serve per chiamare le richieste pregresse de L\'agenzia alle torri e fissare appuntamenti di acquisizione')
G = [
 ('',),
 ('REGOLA NUMERO 1: il successo della chiamata e l\'APPUNTAMENTO DI VALUTAZIONE per Diego (o un suo collaboratore). Tutto il resto e dettaglio.',True),
 ('',),
 ('1. CHI STAI CHIAMANDO',True),
 ('Le persone in lista in passato hanno fatto una RICHIESTA di ricerca casa (acquisto o affitto) tramite gli uffici Tecnocasa di Diego. Oggi Diego lavora come "L\'agenzia alle torri" (zona Legnaia, Scandicci, Ponte a Greve / Isolotto).',),
 ('Logica: chi cerca/cercava casa nell\'80% dei casi deve anche VENDERE quella attuale. Quindi ogni richiesta-acquirente e potenzialmente un immobile da acquisire. Questo e l\'oro della lista.',),
 ('Molti avranno gia trovato casa: non e un problema, vanno comunque "mantenuti" per il futuro (ciclo casa 5-8 anni).',),
 ('',),
 ('2. WORKFLOW GIORNALIERO',True),
 ('Quando','Cosa fare'),
 ('Mattina inizio','Apri SETTER_CHIAMATE. Parti dal primo VERDE in alto. Verifica che la riga non abbia gia un Esito.'),
 ('Prima della chiamata','Leggi la colonna "Cosa dire": ti dice cosa cercava e qual e il gancio (permuta/prima casa = ha quasi certamente da vendere).'),
 ('Durante','Usa lo script (sez. 4). Non leggerlo a memoria, adatta al tono. Tieni davanti i 4 scenari (sez. 5).'),
 ('Subito dopo','Aggiorna ESITO, NOTE, e se fissi: APPUNTAMENTO (data/ora). Non aspettare fine giornata.'),
 ('Fissato un appuntamento','Inseriscilo SUBITO sul Google Calendar condiviso di Diego. NON mandare messaggio di conferma al cliente.'),
 ('Fine giornata','Mini-report (5 righe) al referente Telesales su WhatsApp: chiamate, contattati, appuntamenti fissati, note strane.'),
 ('Settimanale','Venerdi review con il referente Telesales: porta i 2-3 nominativi piu interessanti.'),
 ('',),
 ('3. COSA SIGNIFICA OGNI COLORE',True),
 ('Colore','Significato'),
 ('VERDE - priorita A','Chiama QUESTI PRIMA. Acquisto recente (2024-2025), in zona, con segnale di vendita (permuta o prima casa). Alta probabilita di avere un immobile da acquisire.'),
 ('GIALLO - priorita B','Dopo i verdi. Altri acquirenti recenti o in zona. Buona probabilita.'),
 ('GRIGIO - priorita C','Solo dopo verdi e gialli. Locazione (spesso inquilini, bassa probabilita di possedere), acquisti vecchi, comuni lontani, solo-fisso. Chiamali con efficienza.'),
 ('',),
 ('4. SCRIPT - metodo di Diego (come approccia lui)',True),
 ('REGOLA CHIAVE: la domanda qualificante NON e "sta vendendo casa?" (controproducente sul bene primario). Si parte dalla casa che cercava, poi si chiede "per l\'acquisto pensava a un MUTUO o ha un APPARTAMENTO DA VENDERE?". Cosi emerge l\'immobile da acquisire in modo naturale. E ai clienti si parla SEMPRE di "valutazione dell\'immobile", MAI di "acquisizione".',),
 ('4.A - Presentazione + info sulla casa cercata',True),
 ('"Buongiorno, sono [NOME] de L\'agenzia alle torri. Parlo con [NOME CLIENTE]? In passato ci aveva contattato perche stava cercando casa in zona. La chiamo per capire se sta ancora cercando: cos\'e che le serviva di preciso?"',),
 ('-> Lascia parlare. Fatti dire tipologia, zona, budget, tempi.',),
 ('4.B - Domanda qualificante (il cuore - parole di Diego)',True),
 ('"Perfetto. E per l\'acquisto, pensava di prendere un mutuo oppure ha un appartamento da vendere?"',),
 ('-> Se "ho un appartamento da vendere" (o se era gia indicato PERMUTA): e il gancio. Vai all\'appuntamento di valutazione (4.C).',),
 ('-> Se "mutuo / nessun immobile": e comunque un potenziale acquirente per Diego. Registra cosa cerca, Esito = "Sta cercando casa".',),
 ('4.C - Chiusura appuntamento di valutazione (fissare "a prescindere")',True),
 ('"Allora le fisso un passaggio per la valutazione del suo immobile: viene un mio collaboratore, le porta un\'analisi sui prezzi reali della zona. Le va meglio [giorno X] o [giorno Y]?"',),
 ('-> Proponi 2 opzioni concrete. Fissa data e ora precise. NON dire "poi la richiamo per confermare".',),
 ('-> Inserisci subito sul Google Calendar di Diego. Niente messaggio di conferma. (Internamente e un appuntamento di ACQUISIZIONE, ma al cliente si dice solo VALUTAZIONE.)',),
 ('',),
 ('5. I 4 SCENARI - cosa fare quando il cliente risponde',True),
 ('A. HA GIA TROVATO/COMPRATO casa',''),
 ('Cosa fai:','Non chiudere secco. "Perfetto, complimenti. E quella di prima l\'ha venduta o la tiene?" Se la tiene/affitta -> possibile acquisizione futura. Comunque mantieni il contatto: "La richiamo tra qualche tempo, queste cose tornano." Esito = "Ha gia trovato/non cerca piu".'),
 ('B. STA ANCORA CERCANDO + ha un immobile da vendere',''),
 ('Cosa fai:','GANCIO. Fissa appuntamento di valutazione (sez. 4.C). Esito = "Appuntamento fissato".'),
 ('C. STA CERCANDO ma NON ha da vendere (es. primo acquisto, in affitto)',''),
 ('Cosa fai:','Resta utile: registra cosa cerca (potenziale acquirente per Diego). "La tengo presente, se trovo qualcosa di adatto la chiamo." Esito = "Sta cercando casa". Nota cosa cerca.'),
 ('D. NON cerca e NON vende',''),
 ('Cosa fai:','Chiudi gentile, mantieni: "Grazie, se in futuro le servisse una valutazione ci siamo." Esito = "Non possiede da vendere" o "Non interessato". NON insistere, NON bruciare.'),
 ('',),
 ('6. GESTIONE OBIEZIONI',True),
 ('Obiezione','Cosa rispondere'),
 ('"Chi siete? Chi vi ha dato il numero?"','"Ci aveva contattato in passato per cercare casa, tramite l\'ufficio in zona. Oggi siamo L\'agenzia alle torri. La richiamo proprio per capire se le serve ancora una mano."'),
 ('"Ho gia un\'agenzia"','"Capisco. Le lascio comunque un riferimento per una valutazione gratuita e aggiornata: i prezzi in zona sono saliti del 15-18% in un anno, magari la sua casa vale piu di quanto pensa."'),
 ('"Quanto vale la mia casa?"','Non dare cifre al telefono. "Glielo dice con precisione il collega in valutazione, sui prezzi reali degli ultimi atti nella sua via. Le fisso un passaggio?"'),
 ('"Non ho tempo adesso"','Proponi orario preciso. "La richiamo domani alle 10 o alle 16, cosa preferisce?" Annota data/ora.'),
 ('"Mandatemi una mail"','"Volentieri. Le richiamo io tra 5 giorni per un riscontro." Poi richiama davvero.'),
 ('"Non mi interessa, non chiamatemi piu"','Rispetta. "Capito, grazie, buona giornata." Esito = "Non interessato". Non richiamare.'),
 ('',),
 ('7. REGOLE D\'ORO (sempre)',True),
 ('1. NON chiedere diretto se vende.','Domanda di Diego: "per l\'acquisto pensava a un mutuo o ha un appartamento da vendere?".'),
 ('2. Al cliente di\' sempre VALUTAZIONE.','Mai dire "acquisizione". Internamente e acquisizione, al telefono e "valutazione dell\'immobile".'),
 ('3. L\'obiettivo e l\'APPUNTAMENTO.','La valutazione e la vendita le fa Diego/collaboratore di persona.'),
 ('4. Fissa "a prescindere".','Una volta agganciato, fissa data e ora. Non sfissare, non rimandare a "poi confermo".'),
 ('5. MAI cifre/valutazioni al telefono.','Le da il collega in sopralluogo.'),
 ('6. NON bruciare il contatto.','Anche un no oggi vale tra 5-8 anni. Tono sempre curato.'),
 ('7. Chiama a nome de "L\'agenzia alle torri".',''),
 ('',),
 ('8. ESITI - significato (usa il menu a tendina, non testo libero)',True),
 ('Da chiamare','Stato iniziale.'),
 ('Non risposto','Hai squillato e NESSUNO ha risposto. Riprova entro 24-48h. (Solo se nessuno risponde.)'),
 ('Da richiamare','Hai parlato ma non era il momento/persona, oppure segreteria/centralino. Metti data prossimo contatto.'),
 ('Sta cercando casa','Potenziale acquirente per Diego, ma non ha (ancora) da vendere. Annota cosa cerca.'),
 ('Ha un immobile da vendere','Confermato che possiede e valuterebbe. Fissa subito appuntamento.'),
 ('Appuntamento fissato','Hai fissato la valutazione e l\'hai messa sul Calendar. (Questo e il "win".)'),
 ('Ha gia trovato/non cerca piu','Ha risolto. Mantieni per il futuro.'),
 ('Non possiede da vendere','Non ha immobili da vendere (es. in affitto). Bassa priorita.'),
 ('Non interessato','Ha detto chiaro di non voler essere ricontattato. NON richiamare.'),
 ('Numero errato','Numero sbagliato. Annota in note.'),
 ('',),
 ('9. NOTE - come scriverle',True),
 ('Brevi, professionali, come le scriverebbe un commerciale. Riassumi: cosa cerca / cosa ha da vendere / quando richiamare. Mai dettagli tecnici interni, mai gergo. Es: "Cerca trilocale Isolotto, ha bilo da vendere in zona, valutazione fissata gio 19/6 h18."',),
 ('',),
 ('10. TONO DI VOCE',True),
 ('Calmo','Mai affannato. Anche se la chiamata va storta.'),
 ('Curioso','Sei li per capire cosa cerca, non per vendere. Domande aperte.'),
 ('Concreto','Frasi corte. "Le va meglio giovedi o lunedi?" meglio di lunghi giri di parole.'),
 ('Pari grado','Professionale, mai supplicante, mai aggressivo.'),
 ('',),
 ('11. FINE GIORNATA - mini-report al referente Telesales (WhatsApp, 5 righe)',True),
 ('Esempio: "Torri - gio 12 giu | 28 chiamate, 13 contattati | 3 appuntamenti valutazione (mer 18 h17, gio 19 h18, ven 20 h10) | 2 stanno ancora cercando (li ho segnati) | 1 numero errato. Domani continuo dai verdi."',),
]
r = 3
for tup in G:
    text = tup[0]; bold = len(tup) > 1 and tup[1] is True
    second = tup[1] if (len(tup) > 1 and tup[1] is not True) else None
    ws.cell(row=r, column=1, value=text)
    if bold: ws.cell(row=r, column=1).font = F_B
    if second is not None:
        ws.cell(row=r, column=2, value=second)
    ws.cell(row=r, column=1).alignment = WRAP
    ws.cell(row=r, column=2).alignment = WRAP
    r += 1
fill_row_title(ws, 1, 2)
ws.column_dimensions['A'].width = 60
ws.column_dimensions['B'].width = 78

# ============================================================ SETTER_CHIAMATE
ws = wb.create_sheet('SETTER_CHIAMATE')
callable_ = [x for x in ded if x['fband'] in ('A','B','C','C-fisso')]
order = {'A':0,'B':1,'C':2,'C-fisso':3}
callable_.sort(key=lambda x: (order[x['fband']], -x['score'], -x['year']))

def euro(p):
    if isinstance(p,(int,float)) and p>0: return f"{int(p):,}".replace(',','.')+" EUR"
    return ''
def cosa_dire(x):
    tipo = x['tipo'] or 'immobile'
    zona = x['comune'].title() if x['comune'] else ''
    anno = x['year'] or ''
    bud = euro(x['prezzo'])
    if x['motiv']=='Locazione':
        return f"Cercava in AFFITTO {tipo} a {zona} ({anno}). Probabile inquilino. Sonda comunque se possiede un immobile da vendere."
    if x['dest']=='Permuta':
        return f"PERMUTA - cercava {tipo} a {zona} ({anno}) cambiando casa: ha quasi certamente un immobile da vendere. Gancio diretto su valutazione."
    if x['dest']=='Prima Casa':
        return f"Cercava PRIMA CASA: {tipo} a {zona} {('~'+bud) if bud else ''} ({anno}). Probabile debba vendere l'attuale. Chiedi: mutuo o appartamento da vendere?"
    if x['dest']=='Investimento':
        return f"Investitore: cercava {tipo} a {zona} {('~'+bud) if bud else ''} ({anno}). Spesso possiede altri immobili. Chiedi: mutuo o immobile da vendere/valutare?"
    return f"Cercava {tipo} a {zona} {('~'+bud) if bud else ''} ({anno}). Apri sulla casa cercata, poi chiedi: mutuo o appartamento da vendere? Se ha da vendere, fissa la valutazione."

cols = ['#','Nome','Prio','Zona/Comune','Tipo cercato','Budget','Motivo','Quando',
        'Cosa dire (gancio acquisizione)','Dettaglio richiesta (storico)','Cellulare','Tel fisso',
        'Esito','Note chiamata','Appuntamento (data/ora)']
title_block(ws,'SETTER_CHIAMATE - la tua lista di lavoro',
    f'{sum(1 for x in callable_ if x["fband"]=="A")} verdi (A) - {sum(1 for x in callable_ if x["fband"]=="B")} gialli (B) - {sum(1 for x in callable_ if x["fband"] in ("C","C-fisso"))} grigi (C). Inizia dai verdi in alto. Ordina/filtra a piacere ma lavora top-down.')
ws['A3']='VERDE = chiama PRIMA   |   GIALLO = dopo i verdi   |   GRIGIO = solo dopo i gialli'
ws['A3'].font=F_SUB
hdr_row=4
for j,c in enumerate(cols,1):
    cell=ws.cell(row=hdr_row,column=j,value=c); cell.fill=C_HDR; cell.font=F_HDRW; cell.alignment=WRAP
r=hdr_row+1
fillmap={'A':C_A,'B':C_B,'C':C_C,'C-fisso':C_C}
for i,x in enumerate(callable_,1):
    prio = 'A' if x['fband']=='A' else ('B' if x['fband']=='B' else 'C')
    vals=[i, x['nome'], prio, x['comune'].title(), x['tipo'], euro(x['prezzo']),
          x['motiv'], (x['data'] or ''), cosa_dire(x), (x['esig'] or '')[:200],
          ("'"+x['cell']) if x['cell'] else '', ("'"+x['telfisso']) if x['telfisso'] else '',
          'Da chiamare', '', '']
    for j,v in enumerate(vals,1):
        cell=ws.cell(row=r,column=j,value=v)
        cell.fill=fillmap[x['fband']]
        cell.alignment = WRAP if j in (9,10) else TOP
    r+=1
last=r-1
# data validation esiti
esiti='"Da chiamare,Non risposto,Da richiamare,Sta cercando casa,Ha un immobile da vendere,Appuntamento fissato,Ha gia trovato/non cerca piu,Non possiede da vendere,Non interessato,Numero errato"'
dv=DataValidation(type='list',formula1=esiti,allow_blank=True)
ws.add_data_validation(dv); dv.add(f'M{hdr_row+1}:M{last}')
prio_dv=DataValidation(type='list',formula1='"A,B,C"',allow_blank=True)
ws.add_data_validation(prio_dv); prio_dv.add(f'C{hdr_row+1}:C{last}')
widths={'A':5,'B':22,'C':5,'D':16,'E':16,'F':11,'G':13,'H':12,'I':52,'J':46,'K':13,'L':12,'M':22,'N':34,'O':18}
for k,v in widths.items(): ws.column_dimensions[k].width=v
ws.freeze_panes='A5'
ws.auto_filter.ref=f'A{hdr_row}:O{last}'

# ============================================================ ICP_e_INSIGHT
ws=wb.create_sheet('ICP_e_INSIGHT')
title_block(ws,'ICP e INSIGHT - mercato acquisizione Firenze sud-ovest','Profilo del lead di acquisizione + sintesi ricerca di mercato (dettaglio in Dossier_Mercato_IMEON.md)')
r=4
ws.cell(row=r,column=1,value='PROFILO LEAD DI ACQUISIZIONE (chi cerca casa e probabilmente ha casa da vendere)').font=F_B; r+=1
for c,t in enumerate(['Parametro','Valore'],1):
    cell=ws.cell(row=r,column=c,value=t); cell.fill=C_HDR; cell.font=F_HDRW
r+=1
icp=[
 ('Profilo dominante','Coppia/famiglia in upgrade (da bilo/trilocale a trilo/quadrilocale per i figli) - 74% degli acquirenti'),
 ('Segnale piu forte','Destinazione PERMUTA o PRIMA CASA + acquisto recente (2024-2025)'),
 ('Zona di valore','Firenze sud-ovest (Isolotto, Legnaia, Soffiano, Ponte a Greve) + Scandicci'),
 ('Taglio piu richiesto','Trilocale (2 camere) 32,7% - "i trilocali vanno a ruba"'),
 ('Tempi di permanenza','Quota crescente di famiglie cambia casa ogni 5-8 anni (mobilita 6,4%/anno)'),
 ('Chi escludere/depriorizzare','Locazione (probabili inquilini, non proprietari), comuni lontani, richieste pre-2022'),
 ('Acquirente da investimento','22,5% a Firenze (% piu alta d\'Italia), 77% in contanti -> spesso possiede immobili a reddito'),
]
for p,v in icp:
    ws.cell(row=r,column=1,value=p); c=ws.cell(row=r,column=2,value=v); c.alignment=WRAP; r+=1
r+=1
ws.cell(row=r,column=1,value='10 INSIGHT DI MERCATO -> uso nello script').font=F_B; r+=1
for c,t in enumerate(['Area','Insight chiave (dati verificati)','Come usarlo in chiamata'],1):
    cell=ws.cell(row=r,column=c,value=t); cell.fill=C_HDR; cell.font=F_HDRW; cell.alignment=WRAP
r+=1
ins=[
 ('Mercato del venditore','Firenze: vendita in ~97 giorni, sconto medio sul richiesto solo 7% (vs 11% nazionale)','"Le case qui si vendono in meno di 3 mesi con sconti minimi. Se ha un immobile e il momento giusto."'),
 ('Prezzi in corsa','Sud-ovest +15-18% in 12 mesi (Isolotto 3.732, Legnaia-Soffiano 3.946 EUR/mq). Firenze verso 5.000','"I prezzi qui sono saliti del 15-18% in un anno: forse la sua casa vale piu di quanto pensa."'),
 ('Offerta scarsa','Offerta in calo, domanda solida -> chi vende ha potere','"Pochissime case sul mercato, tanti acquirenti in attesa. Se vende ora ha la fila."'),
 ('Deve prima vendere','74% acquirenti sono famiglie in trade-up: chi compra deve vendere l\'attuale','"Quando cercava casa, doveva anche venderne una? Quella che ha oggi e ancora sua?"'),
 ('Trilocale tirato','Trilocale = 32,7% della domanda, taglio piu richiesto','"I trilocali in zona vanno a ruba, ho gia acquirenti che li cercano."'),
 ('Isolotto rotazione','Zona piu economica della citta (3.732 EUR/mq, affitto 15,5) -> alta rotazione','Targetizzare chi e in zona da 5-8 anni: tipico momento di upgrade.'),
 ('Valutazione PRO','La valutazione "gratis e generica" attira perditempo; funziona quella professionale','"Le porto un\'analisi sui prezzi reali degli ultimi atti nella sua via, non una stima a caso."'),
 ('Case Green','Direttiva UE: classi energetiche basse perdono valore','"Con le nuove regole UE le case in classe bassa rischiano di valere meno: meglio capire dove si colloca."'),
 ('190 competitor','~190 agenzie in Q4 Isolotto-Legnaia, 32 a Scandicci -> guerra di velocita','Chiudere SEMPRE su appuntamento fisico ravvicinato; puntare all\'esclusiva (ROI 5-8x).'),
 ('Investitori cash','77% degli investimenti a Firenze in contanti','Per chi ha immobile a reddito: "Ho acquirenti liquidi pronti, lo valutiamo?"'),
]
for a,b,c in ins:
    ws.cell(row=r,column=1,value=a);
    x=ws.cell(row=r,column=2,value=b); x.alignment=WRAP
    y=ws.cell(row=r,column=3,value=c); y.alignment=WRAP
    r+=1
fill_row_title(ws,1,3)
ws.column_dimensions['A'].width=22; ws.column_dimensions['B'].width=58; ws.column_dimensions['C'].width=58

# ============================================================ KPI_TRACKING
ws=wb.create_sheet('KPI_TRACKING')
title_block(ws,'KPI_TRACKING - target + review settimanale','Primo aggiornamento a Diego entro lunedi 15 giugno 2026')
r=4
ws.cell(row=r,column=1,value='TARGET').font=F_B; r+=1
for c,t in enumerate(['Obiettivo','Target','Attuale','Note'],1):
    cell=ws.cell(row=r,column=c,value=t); cell.fill=C_HDR; cell.font=F_HDRW
r+=1
for o,t,a,n in [('Appuntamenti valutazione fissati','da definire',0,'Setter su lista pregresso'),
                ('Immobili acquisiti (incarichi)','da definire',0,'Esito appuntamenti'),
                ('Contatti "mantenuti" per futuro','-',0,'Chi ha gia trovato / non vende ora')]:
    ws.cell(row=r,column=1,value=o);ws.cell(row=r,column=2,value=t);ws.cell(row=r,column=3,value=a);ws.cell(row=r,column=4,value=n);r+=1
r+=1
ws.cell(row=r,column=1,value='REVIEW SETTIMANALE').font=F_B; r+=1
for c,t in enumerate(['Settimana','Date','N. chiamate','Contattati','Appuntamenti fissati','Acquisizioni'],1):
    cell=ws.cell(row=r,column=c,value=t); cell.fill=C_HDR; cell.font=F_HDRW
r+=1
for w,d in [('W1','12-13 giu'),('W2','16-20 giu'),('W3','23-27 giu'),('W4','30 giu-4 lug'),
            ('W5','7-11 lug'),('W6','14-18 lug'),('W7','21-25 lug'),('W8','28 lug-1 ago')]:
    ws.cell(row=r,column=1,value=w);ws.cell(row=r,column=2,value=d);r+=1
r+=1
ws.cell(row=r,column=1,value='PROSSIMI PASSI (dalla call onboarding)').font=F_B; r+=1
for s in ['1. Diego invia via mail lo script ideale di acquisizione -> integrarlo nella GUIDA_SETTER (sez. 4).',
          '2. Diego condivide il Google Calendar con i setter (impostazioni > condivisione > indirizzo fornito).',
          '3. Valutare/attivare reminder automatico calendario per gli appuntamenti (no messaggi di conferma).',
          '4. Report a Diego entro lunedi 15 giugno 2026.',
          '5. (Fase 2) Scouting annunci privati in zona da matchare con le richieste in lista -> bacino acquisizione piu ampio.',
          '6. (Fase 2) Recupero numeri per i 113 contatti senza telefono.']:
    ws.cell(row=r,column=1,value=s); r+=1
fill_row_title(ws,1,6)
for k,v in {'A':40,'B':16,'C':14,'D':18,'E':20,'F':16}.items(): ws.column_dimensions[k].width=v

# ============================================================ DB_COMPLETO
ws=wb.create_sheet('DB_COMPLETO')
title_block(ws,'DB COMPLETO - tutti i contatti (backup) + legenda',
    f'{len(ded)} contatti unici (incl. {sum(1 for x in ded if x["fband"]=="X")} senza numero). Sorgente: 2 file "Richieste con residenza" uniti e deduplicati.')
cols=['#','Prio','Nome','Cellulare','Tel fisso','Comune','Indirizzo','Civico','Motivazione','Budget','Tipologia','Data richiesta','Destinazione','Dettaglio richiesta (storico)','Score','Fonte']
hdr=4
for j,c in enumerate(cols,1):
    cell=ws.cell(row=hdr,column=j,value=c); cell.fill=C_HDR; cell.font=F_HDRW; cell.alignment=WRAP
allrec=sorted(ded,key=lambda x:(order.get(x['fband'],9),-x['score']))
r=hdr+1
for i,x in enumerate(allrec,1):
    prio={'A':'A','B':'B','C':'C','C-fisso':'C','X':'NO TEL'}[x['fband']]
    esig=(x['esig'] or '')[:160]
    vals=[i,prio,x['nome'],("'"+x['cell']) if x['cell'] else '',("'"+x['telfisso']) if x['telfisso'] else '',
          x['comune'].title(),x['indirizzo'],x['civico'],x['motiv'],euro(x['prezzo']),x['tipo'],
          x['data'],x['dest'],esig,x['score'],x['src']]
    for j,v in enumerate(vals,1):
        cell=ws.cell(row=r,column=j,value=v)
        if x['fband']=='A':cell.fill=C_A
        elif x['fband']=='B':cell.fill=C_B
        elif x['fband']=='X':cell.fill=PatternFill('solid',fgColor='F8CBAD')
        else:cell.fill=C_C
        cell.alignment=TOP
    r+=1
dblast=r-1
ws.freeze_panes='A5'; ws.auto_filter.ref=f'A{hdr}:P{dblast}'
for k,v in {'A':5,'B':7,'C':22,'D':13,'E':12,'F':14,'G':20,'H':7,'I':12,'J':12,'K':12,'L':13,'N':45,'O':6,'P':24}.items():
    ws.column_dimensions[k].width=v
# legenda
r=dblast+3
ws.cell(row=r,column=1,value='LEGENDA').font=F_B; r+=1
for k,v in [('Prio A (verde)','Acquisto recente, in zona, segnale di vendita - chiama subito'),
            ('Prio B (giallo)','Altri acquirenti recenti/in zona'),
            ('Prio C (grigio)','Locazione, acquisti vecchi, comuni lontani, solo-fisso'),
            ('NO TEL (arancio)','Nessun numero: non chiamabili ora, da arricchire in fase 2'),
            ('Score','Punteggio di priorita acquisizione (motivazione+recency+zona+segnale vendita)'),
            ('Fonte','File di origine: 2025-06 ufficio attivo / 2025-01 ufficio precedente')]:
    ws.cell(row=r,column=1,value=k).font=F_B; c=ws.cell(row=r,column=3,value=v); c.alignment=WRAP; r+=1
fill_row_title(ws,1,len(cols))

out='/Users/simocors/Desktop/telesales/imeon/Master_IMEON_Torri.xlsx'
wb.save(out)
print('SALVATO:',out)
print('Tab:',wb.sheetnames)
print('SETTER_CHIAMATE righe dati:',last-hdr_row)
print('DB righe:',dblast-hdr)
