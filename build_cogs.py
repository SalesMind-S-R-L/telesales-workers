from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.utils import get_column_letter

SKILLS_SCRIPTS = "/Users/simocors/Library/Application Support/Claude/local-agent-mode-sessions/skills-plugin/7d64f5c6-e116-44b4-ba5e-9e9ef66a3ccb/0128c0c0-3597-4ad6-8f1c-9f636b750686/skills/xlsx/scripts"

# ── Colours ────────────────────────────────────────────────────────────────
BLUE_DARK   = "1F3864"   # header bg
WHITE       = "FFFFFF"
GREY_LIGHT  = "F2F2F2"
BLUE_INPUT  = "0000FF"   # hardcoded input text (industry std)
BLACK       = "000000"   # formula text
GREEN_LINK  = "008000"   # cross-sheet links
YELLOW_BG   = "FFFF00"

# ── Borders ────────────────────────────────────────────────────────────────
thin = Side(style="thin", color="CCCCCC")
med  = Side(style="medium", color="888888")
THIN_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
MED_BORDER  = Border(left=med,  right=med,  top=med,  bottom=med)

def hdr_font(size=10):
    return Font(name="Arial", bold=True, color=WHITE, size=size)

def inp_font(bold=False):
    return Font(name="Arial", color=BLUE_INPUT, bold=bold)

def calc_font():
    return Font(name="Arial", color=BLACK)

def link_font():
    return Font(name="Arial", color=GREEN_LINK)

def hdr_fill():
    return PatternFill("solid", start_color=BLUE_DARK)

def alt_fill():
    return PatternFill("solid", start_color=GREY_LIGHT)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center")

def right_align():
    return Alignment(horizontal="right", vertical="center")

FMT_EUR  = u'€#,##0.00;(€#,##0.00);"-"'
FMT_USD  = u'$#,##0.00;($#,##0.00);"-"'
FMT_PCT  = '0.0%;(0.0%);"-"'
FMT_NUM  = '#,##0;(#,##0);"-"'

def apply_border(cell):
    cell.border = THIN_BORDER

def style_header(cell, size=10):
    cell.font = hdr_font(size)
    cell.fill = hdr_fill()
    cell.alignment = center()
    apply_border(cell)

def style_input(cell, fmt=FMT_USD, bold=False):
    cell.font = inp_font(bold)
    cell.number_format = fmt
    cell.alignment = right_align()
    apply_border(cell)

def style_calc(cell, fmt=FMT_USD):
    cell.font = calc_font()
    cell.number_format = fmt
    cell.alignment = right_align()
    apply_border(cell)

def style_link(cell, fmt=FMT_USD):
    cell.font = link_font()
    cell.number_format = fmt
    cell.alignment = right_align()
    apply_border(cell)

# ══════════════════════════════════════════════════════════════════════════════
wb = Workbook()

# ── Sheet 0: Assumptions ──────────────────────────────────────────────────
ws_a = wb.active
ws_a.title = "Assumptions"

ws_a.row_dimensions[1].height = 30
ws_a.merge_cells("A1:D1")
t = ws_a["A1"]
t.value = "COGS Tech Analysis – Assumptions"
t.font = Font(name="Arial", bold=True, size=14, color=WHITE)
t.fill = hdr_fill()
t.alignment = center()

headers_a = ["Parameter", "Value", "Unit", "Source / Note"]
for c, h in enumerate(headers_a, 1):
    cell = ws_a.cell(row=2, column=c, value=h)
    style_header(cell)

assumptions = [
    ("GHL Agency Unlimited",  297,    "$/mese",   "GoHighLevel pricing, May 2026"),
    ("ElevenLabs rate",       0.08,   "$/min",     "ElevenLabs outbound voice, May 2026"),
    ("Telnyx SIP rate",       0.0075, "$/min",     "Average of $0.005–$0.01"),
    ("Instantly (shared)",    42,     "$/mese",    "Instantly.ai plan, May 2026"),
    ("Vercel / Cloudflare",   0,      "$/mese",    "Free tier"),
    ("EUR/USD",               1.0,    "ratio",     "1:1 assumption for simplicity"),
    ("",                      None,   "",          ""),
    ("SalesMind Start – prezzo",        650,  "€/mese", "Pack pricing"),
    ("SalesMind Pro – prezzo",         1100,  "€/mese", "Pack pricing"),
    ("SalesMind Domination – prezzo",  1800,  "€/mese", "Pack pricing"),
    ("AIVoice Max Enterprise – prezzo",3500,  "€/mese", "Pack pricing – stima"),
    ("",                      None,   "",          ""),
    ("SalesMind Start – min AI",         0,   "min/mese", "No voice AI"),
    ("SalesMind Pro – min AI",         100,   "min/mese", "Centro range 0-200"),
    ("SalesMind Domination – min AI",  350,   "min/mese", "Centro range 200-500"),
    ("AIVoice Max Enterprise – min AI",1250,  "min/mese", "Centro range 500-2000"),
]

for r, (param, val, unit, note) in enumerate(assumptions, 3):
    ws_a.cell(row=r, column=1, value=param).font = Font(name="Arial", bold=(val is not None))
    c_val = ws_a.cell(row=r, column=2, value=val)
    if val is not None:
        c_val.font = inp_font(bold=True)
        if isinstance(val, float) and val < 1:
            c_val.number_format = '$#,##0.0000'
        elif isinstance(val, int) and val >= 100:
            c_val.number_format = FMT_EUR if "prezzo" in param else FMT_USD
        c_val.alignment = right_align()
    ws_a.cell(row=r, column=3, value=unit).font = Font(name="Arial", color="555555")
    ws_a.cell(row=r, column=4, value=note).font  = Font(name="Arial", color="555555", italic=True)
    if r % 2 == 0 and val is not None:
        for c in range(1, 5):
            ws_a.cell(row=r, column=c).fill = alt_fill()

ws_a.column_dimensions["A"].width = 38
ws_a.column_dimensions["B"].width = 14
ws_a.column_dimensions["C"].width = 14
ws_a.column_dimensions["D"].width = 42
ws_a.freeze_panes = "A3"

# Named ranges for cross-sheet reference (row numbers in Assumptions sheet)
# Row 3=GHL, 4=11labs, 5=telnyx, 6=instantly
# Row 8=start€, 9=pro€, 10=dom€, 11=ent€
# Row 13=start_min, 14=pro_min, 15=dom_min, 16=ent_min
ASS = "Assumptions"
GHL_CELL        = f"{ASS}!$B$3"
ELEVEN_CELL     = f"{ASS}!$B$4"
TELNYX_CELL     = f"{ASS}!$B$5"
INST_CELL       = f"{ASS}!$B$6"
PRICE_ROW       = {0: f"{ASS}!$B$8", 1: f"{ASS}!$B$9", 2: f"{ASS}!$B$10", 3: f"{ASS}!$B$11"}
MIN_ROW         = {0: f"{ASS}!$B$13", 1: f"{ASS}!$B$14", 2: f"{ASS}!$B$15", 3: f"{ASS}!$B$16"}

# ── Sheet 1: COGS Analysis (main) ─────────────────────────────────────────
ws1 = wb.create_sheet("COGS Analysis")

ws1.row_dimensions[1].height = 36
ws1.merge_cells("A1:O1")
t = ws1["A1"]
t.value = "COGS Tech Analysis – Pack Telesales AI"
t.font = Font(name="Arial", bold=True, size=14, color=WHITE)
t.fill = hdr_fill()
t.alignment = center()

# Legend row
ws1.merge_cells("A2:O2")
leg = ws1["A2"]
leg.value = (
    "Testo BLU = input modificabile     |     "
    "Testo NERO = formula calcolata     |     "
    "Testo VERDE = link da altro foglio"
)
leg.font = Font(name="Arial", size=9, italic=True, color="444444")
leg.alignment = center()
leg.fill = PatternFill("solid", start_color="E8ECF0")

# Sub-headers row 3
ws1.merge_cells("A3:C3")
ws1.merge_cells("D3:H3")
ws1.merge_cells("I3:O3")
for cell_addr, label in [("A3",""), ("D3","SCENARIO: 10 CLIENTI ATTIVI"), ("I3","SCENARIO: 20 CLIENTI ATTIVI")]:
    c = ws1[cell_addr]
    c.value = label
    c.font = Font(name="Arial", bold=True, size=10, color=WHITE)
    c.fill = PatternFill("solid", start_color="2E5597")
    c.alignment = center()
    c.border = THIN_BORDER
for col in range(1, 16):
    ws1.cell(row=3, column=col).border = THIN_BORDER

# Column headers row 4
col_headers = [
    "Pack",             # A
    "Prezzo (€/mese)",  # B
    "Min AI/mese",      # C
    # 10 cli
    "GHL 10 cli ($)",   # D
    "ElevenLabs ($)",   # E
    "Telnyx ($)",       # F
    "Instantly 10 cli ($)", # G
    "COGS Totale 10 cli ($)", # H
    "Margine Lordo 10 cli (€)", # I
    "Margine % 10 cli", # J  -- wait, restructure: I=margine, J=marg%
    # 20 cli
    "GHL 20 cli ($)",   # K
    "Instantly 20 cli ($)", # L
    "COGS Totale 20 cli ($)", # M
    "Margine Lordo 20 cli (€)", # N
    "Margine % 20 cli", # O
]
# recount: A=1..O=15 → 15 cols
for c, h in enumerate(col_headers, 1):
    cell = ws1.cell(row=4, column=c, value=h)
    style_header(cell, size=9)

ws1.row_dimensions[4].height = 42

PACKS = ["SalesMind Start", "SalesMind Pro", "SalesMind Domination", "AIVoice Max Enterprise"]

for i, pack in enumerate(PACKS):
    r = 5 + i
    fill = alt_fill() if i % 2 == 0 else PatternFill("solid", start_color=WHITE)

    # A: Pack name
    c_a = ws1.cell(row=r, column=1, value=pack)
    c_a.font = Font(name="Arial", bold=True, size=10)
    c_a.alignment = left()
    c_a.border = THIN_BORDER
    c_a.fill = fill

    # B: Prezzo – link from Assumptions
    c_b = ws1.cell(row=r, column=2, value=f"={PRICE_ROW[i]}")
    style_link(c_b, FMT_EUR); c_b.fill = fill

    # C: Min AI – link from Assumptions
    c_c = ws1.cell(row=r, column=3, value=f"={MIN_ROW[i]}")
    style_link(c_c, FMT_NUM); c_c.fill = fill

    # D: GHL 10 = GHL / 10
    c_d = ws1.cell(row=r, column=4, value=f"={GHL_CELL}/10")
    style_calc(c_d); c_d.fill = fill

    # E: ElevenLabs = min * rate
    c_e = ws1.cell(row=r, column=5, value=f"=C{r}*{ELEVEN_CELL}")
    style_calc(c_e); c_e.fill = fill

    # F: Telnyx = min * rate
    c_f = ws1.cell(row=r, column=6, value=f"=C{r}*{TELNYX_CELL}")
    style_calc(c_f); c_f.fill = fill

    # G: Instantly 10
    c_g = ws1.cell(row=r, column=7, value=f"={INST_CELL}/10")
    style_calc(c_g); c_g.fill = fill

    # H: COGS total 10
    c_h = ws1.cell(row=r, column=8, value=f"=D{r}+E{r}+F{r}+G{r}")
    c_h.font = Font(name="Arial", bold=True, color=BLACK)
    c_h.number_format = FMT_USD; c_h.alignment = right_align()
    c_h.border = THIN_BORDER; c_h.fill = fill

    # I: Margine lordo 10 = Prezzo - COGS
    c_i = ws1.cell(row=r, column=9, value=f"=B{r}-H{r}")
    c_i.font = Font(name="Arial", bold=True, color=BLACK)
    c_i.number_format = FMT_EUR; c_i.alignment = right_align()
    c_i.border = THIN_BORDER; c_i.fill = fill

    # J: Margine % 10
    c_j = ws1.cell(row=r, column=10, value=f"=IF(B{r}>0,(B{r}-H{r})/B{r},0)")
    c_j.font = Font(name="Arial", bold=True, color=BLACK)
    c_j.number_format = FMT_PCT; c_j.alignment = right_align()
    c_j.border = THIN_BORDER; c_j.fill = fill

    # K: GHL 20
    c_k = ws1.cell(row=r, column=11, value=f"={GHL_CELL}/20")
    style_calc(c_k); c_k.fill = fill

    # L: Instantly 20
    c_l = ws1.cell(row=r, column=12, value=f"={INST_CELL}/20")
    style_calc(c_l); c_l.fill = fill

    # M: COGS total 20 = GHL20 + ElevenLabs + Telnyx + Instantly20
    c_m = ws1.cell(row=r, column=13, value=f"=K{r}+E{r}+F{r}+L{r}")
    c_m.font = Font(name="Arial", bold=True, color=BLACK)
    c_m.number_format = FMT_USD; c_m.alignment = right_align()
    c_m.border = THIN_BORDER; c_m.fill = fill

    # N: Margine lordo 20
    c_n = ws1.cell(row=r, column=14, value=f"=B{r}-M{r}")
    c_n.font = Font(name="Arial", bold=True, color=BLACK)
    c_n.number_format = FMT_EUR; c_n.alignment = right_align()
    c_n.border = THIN_BORDER; c_n.fill = fill

    # O: Margine % 20
    c_o = ws1.cell(row=r, column=15, value=f"=IF(B{r}>0,(B{r}-M{r})/B{r},0)")
    c_o.font = Font(name="Arial", bold=True, color=BLACK)
    c_o.number_format = FMT_PCT; c_o.alignment = right_align()
    c_o.border = THIN_BORDER; c_o.fill = fill

# Conditional formatting on margin % cols (J and O, rows 5-8)
from openpyxl.formatting.rule import ColorScaleRule
# Red if <70%, Yellow if ~77.5%, Green if >85%
for col_letter in ["J", "O"]:
    rng = f"{col_letter}5:{col_letter}8"
    ws1.conditional_formatting.add(rng, ColorScaleRule(
        start_type="num", start_value=0.0, start_color="FF4C4C",
        mid_type="num",   mid_value=0.775,  mid_color="FFFF00",
        end_type="num",   end_value=1.0,   end_color="00B050"
    ))

# Column widths
col_widths = [26, 16, 14, 16, 14, 12, 18, 20, 22, 16, 16, 18, 20, 22, 16]
for i, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

ws1.freeze_panes = "D5"

# ── Sheet 2: Dettaglio per Scenario ─────────────────────────────────────
ws2 = wb.create_sheet("Dettaglio per Scenario")

ws2.row_dimensions[1].height = 36
ws2.merge_cells("A1:L1")
t2 = ws2["A1"]
t2.value = "Dettaglio COGS per Scenario – Breakdown Voci di Costo"
t2.font = Font(name="Arial", bold=True, size=14, color=WHITE)
t2.fill = hdr_fill()
t2.alignment = center()

# Two tables: 10 cli (A-F) and 20 cli (H-M) with gap col G
def build_scenario_table(ws, start_col, n_clients, scenario_label):
    # Scenario header spanning 6 cols
    sc = get_column_letter(start_col)
    ec = get_column_letter(start_col + 5)
    ws.merge_cells(f"{sc}2:{ec}2")
    cell = ws[f"{sc}2"]
    cell.value = f"SCENARIO: {n_clients} CLIENTI ATTIVI"
    cell.font = Font(name="Arial", bold=True, size=11, color=WHITE)
    cell.fill = PatternFill("solid", start_color="2E5597")
    cell.alignment = center()
    cell.border = THIN_BORDER
    for col in range(start_col, start_col+6):
        ws.cell(row=2, column=col).border = THIN_BORDER

    sub_headers = ["Pack", f"GHL (${297}/{n_clients})", "ElevenLabs ($)", "Telnyx ($)", f"Instantly (${42}/{n_clients})", "COGS Totale ($)"]
    for offset, h in enumerate(sub_headers):
        cell = ws.cell(row=3, column=start_col+offset, value=h)
        style_header(cell, size=9)

    ws.row_dimensions[3].height = 36

    for i, pack in enumerate(PACKS):
        r = 4 + i
        fill = alt_fill() if i % 2 == 0 else PatternFill("solid", start_color=WHITE)

        # Pack name
        c_p = ws.cell(row=r, column=start_col, value=f"='COGS Analysis'!A{5+i}")
        c_p.font = link_font(); c_p.font = Font(name="Arial", bold=True, color=GREEN_LINK)
        c_p.alignment = left(); c_p.border = THIN_BORDER; c_p.fill = fill

        # GHL
        ghl_col = get_column_letter(start_col+1)
        c_ghl = ws.cell(row=r, column=start_col+1, value=f"={GHL_CELL}/{n_clients}")
        style_calc(c_ghl); c_ghl.fill = fill

        # ElevenLabs
        min_ref = MIN_ROW[i]
        c_el = ws.cell(row=r, column=start_col+2, value=f"={min_ref}*{ELEVEN_CELL}")
        style_calc(c_el); c_el.fill = fill

        # Telnyx
        c_tel = ws.cell(row=r, column=start_col+3, value=f"={min_ref}*{TELNYX_CELL}")
        style_calc(c_tel); c_tel.fill = fill

        # Instantly
        c_inst = ws.cell(row=r, column=start_col+4, value=f"={INST_CELL}/{n_clients}")
        style_calc(c_inst); c_inst.fill = fill

        # COGS total
        s = get_column_letter(start_col+1)
        e = get_column_letter(start_col+4)
        c_tot = ws.cell(row=r, column=start_col+5, value=f"=SUM({s}{r}:{e}{r})")
        c_tot.font = Font(name="Arial", bold=True, color=BLACK)
        c_tot.number_format = FMT_USD; c_tot.alignment = right_align()
        c_tot.border = THIN_BORDER; c_tot.fill = fill

    # Summary row
    r_sum = 8
    ws.cell(row=r_sum, column=start_col, value="TOTALE / MEDIA").font = Font(name="Arial", bold=True)
    ws.cell(row=r_sum, column=start_col).border = THIN_BORDER
    for offset in range(1, 6):
        col_l = get_column_letter(start_col + offset)
        c_s = ws.cell(row=r_sum, column=start_col+offset, value=f"=AVERAGE({col_l}4:{col_l}7)")
        c_s.font = Font(name="Arial", bold=True, color=BLACK)
        c_s.number_format = FMT_USD; c_s.alignment = right_align()
        c_s.border = THIN_BORDER
        c_s.fill = PatternFill("solid", start_color="D9E1F2")

build_scenario_table(ws2, start_col=1, n_clients=10, scenario_label="10 CLI")
build_scenario_table(ws2, start_col=8, n_clients=20, scenario_label="20 CLI")

# Gap column
ws2.column_dimensions["G"].width = 2

# Margine table (rows 10-16)
ws2.merge_cells("A10:L10")
hdr10 = ws2["A10"]
hdr10.value = "CONFRONTO MARGINE LORDO (Prezzo – COGS)"
hdr10.font = Font(name="Arial", bold=True, size=11, color=WHITE)
hdr10.fill = hdr_fill(); hdr10.alignment = center()
for col in range(1, 13):
    ws2.cell(row=10, column=col).border = THIN_BORDER

marg_headers = ["Pack", "Prezzo (€)", "COGS 10 cli ($)", "Marg. 10 cli (€)", "Marg. % 10 cli",
                "", "Pack", "Prezzo (€)", "COGS 20 cli ($)", "Marg. 20 cli (€)", "Marg. % 20 cli", ""]
for c, h in enumerate(marg_headers, 1):
    cell = ws2.cell(row=11, column=c, value=h)
    if h:
        style_header(cell, size=9)

for i, pack in enumerate(PACKS):
    r = 12 + i
    fill = alt_fill() if i % 2 == 0 else PatternFill("solid", start_color=WHITE)
    # 10-cli side
    ws2.cell(row=r, column=1, value=pack).font = Font(name="Arial", bold=True); ws2.cell(row=r, column=1).border=THIN_BORDER; ws2.cell(row=r,column=1).fill=fill
    c_pr  = ws2.cell(row=r, column=2, value=f"={PRICE_ROW[i]}"); style_link(c_pr, FMT_EUR); c_pr.fill=fill
    c_cg  = ws2.cell(row=r, column=3, value=f"='COGS Analysis'!H{5+i}"); style_link(c_cg); c_cg.fill=fill
    c_mg  = ws2.cell(row=r, column=4, value=f"=B{r}-C{r}"); style_calc(c_mg, FMT_EUR); c_mg.fill=fill; c_mg.font=Font(name="Arial", bold=True, color=BLACK)
    c_mp  = ws2.cell(row=r, column=5, value=f"=IF(B{r}>0,(B{r}-C{r})/B{r},0)"); style_calc(c_mp, FMT_PCT); c_mp.fill=fill; c_mp.font=Font(name="Arial", bold=True, color=BLACK)
    # gap
    # 20-cli side
    ws2.cell(row=r, column=7, value=pack).font = Font(name="Arial", bold=True); ws2.cell(row=r, column=7).border=THIN_BORDER; ws2.cell(row=r,column=7).fill=fill
    c_pr2 = ws2.cell(row=r, column=8, value=f"={PRICE_ROW[i]}"); style_link(c_pr2, FMT_EUR); c_pr2.fill=fill
    c_cg2 = ws2.cell(row=r, column=9, value=f"='COGS Analysis'!M{5+i}"); style_link(c_cg2); c_cg2.fill=fill
    c_mg2 = ws2.cell(row=r, column=10, value=f"=H{r}-I{r}"); style_calc(c_mg2, FMT_EUR); c_mg2.fill=fill; c_mg2.font=Font(name="Arial", bold=True, color=BLACK)
    c_mp2 = ws2.cell(row=r, column=11, value=f"=IF(H{r}>0,(H{r}-I{r})/H{r},0)"); style_calc(c_mp2, FMT_PCT); c_mp2.fill=fill; c_mp2.font=Font(name="Arial", bold=True, color=BLACK)

# Cond formatting margini % in sheet2
for col_letter in ["E", "K"]:
    rng = f"{col_letter}12:{col_letter}15"
    ws2.conditional_formatting.add(rng, ColorScaleRule(
        start_type="num", start_value=0.0, start_color="FF4C4C",
        mid_type="num",   mid_value=0.775,  mid_color="FFFF00",
        end_type="num",   end_value=1.0,   end_color="00B050"
    ))

# Column widths ws2
cw2 = [24, 14, 16, 16, 14, 2, 24, 14, 16, 16, 14, 2]
for i, w in enumerate(cw2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

ws2.freeze_panes = "A4"

# ── Save ──────────────────────────────────────────────────────────────────
OUT = "/Users/simocors/Desktop/telesales/COGS_Tech_Pack.xlsx"
wb.save(OUT)
print(f"Saved: {OUT}")
