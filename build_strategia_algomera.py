#!/usr/bin/env python3
"""Genera Strategia_Algomera_Telesales.xlsx - deliverable strategia omnicanale."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Palette sobria
NAVY = "1F3864"      # titoli sezione
BLUE = "2E5496"      # header tabelle
LIGHT = "D9E1F2"     # righe alternate / box
ACCENT = "C9A227"    # oro tenue accenti
GREY = "595959"
WHITE = "FFFFFF"
GREEN = "E2EFDA"
RED = "FCE4E4"

f_title = Font(name="Calibri", size=20, bold=True, color=NAVY)
f_sub = Font(name="Calibri", size=11, italic=True, color=GREY)
f_h1 = Font(name="Calibri", size=13, bold=True, color=NAVY)
f_th = Font(name="Calibri", size=11, bold=True, color=WHITE)
f_b = Font(name="Calibri", size=11, color="000000")
f_bb = Font(name="Calibri", size=11, bold=True, color="000000")
f_small = Font(name="Calibri", size=10, color=GREY)

fill_navy = PatternFill("solid", fgColor=NAVY)
fill_blue = PatternFill("solid", fgColor=BLUE)
fill_light = PatternFill("solid", fgColor=LIGHT)
fill_green = PatternFill("solid", fgColor=GREEN)
fill_red = PatternFill("solid", fgColor=RED)
fill_white = PatternFill("solid", fgColor=WHITE)

thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

wrap = Alignment(wrap_text=True, vertical="top")
wrap_c = Alignment(wrap_text=True, vertical="center", horizontal="center")
center = Alignment(horizontal="center", vertical="center")
left_top = Alignment(horizontal="left", vertical="top", wrap_text=True)

wb = openpyxl.Workbook()

def style_header_row(ws, row, ncols, fill=fill_blue):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = f_th
        cell.alignment = wrap_c
        cell.border = border

def title_block(ws, title, subtitle, span=6):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    ws.cell(row=1, column=1, value=title).font = f_title
    ws.cell(row=1, column=1).alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 30
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span)
    ws.cell(row=2, column=1, value=subtitle).font = f_sub

# ============================================================
# TAB 1 - PANORAMICA
# ============================================================
ws = wb.active
ws.title = "Panoramica"
ws.sheet_view.showGridLines = False
title_block(ws, "Strategia commerciale Algòmera x Telesales", "Piano omnicanale coordinato — Chiamate · LinkedIn · Email   |   v1 giugno 2026", span=2)

ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 95

rows = [
    ("", ""),
    ("OBIETTIVO", "Generare appuntamenti qualificati in videocall con decisori realmente in target. Meno appuntamenti ma giusti: ogni prospect deve aver compreso che Algòmera sviluppa software 100% su misura e di sua proprietà, non un prodotto pronto."),
    ("I TRE CANALI", "LinkedIn + Email lavorano come UNA sequenza coordinata sullo stesso contatto. Le chiamate dei setter sono una traccia parallela e indipendente sulla lista lead già condivisa. Tutti gli esiti confluiscono sullo stesso foglio Google in tempo reale."),
    ("PERIMETRO (ICP)", "Aziende 11-50 dipendenti · fatturato ≥ 2 mln · tutta Italia con priorità Lombardia (poi Campania e Lazio). Settori: manifatturiero/produzione, logistica/trasporti, edilizia/costruzioni."),
    ("ESCLUSIONI", "Fuori target: IT/software, web agency, marketing/comunicazione (competitor o budget insufficiente). Solo decisori C-level (titolare, AD, DG, responsabile acquisti/produzione), mai figure intermedie senza potere di firma."),
    ("BLACKLIST", "La lista è già stata incrociata con la vostra blacklist (3.270 domini): 10 aziende coincidenti sono evidenziate in rosso sul foglio ed escluse da TUTTI i canali (chiamate, LinkedIn, email)."),
    ("LEVE DI VALORE", "Software su misura e di vostra proprietà · codice sorgente del cliente · nessuna licenza né canone · integrazione AI · caso Promove (preventivi da 3 giorni a 2 ore) · 560+ progetti, Le Fonti Award, Google Partner, AssoSoftware."),
    ("COSA SERVE PER PARTIRE", "1) Accessi Sales Navigator di Francesco  ·  2) Approvazione dei copy LinkedIn ed email  ·  3) Conferma volumi LinkedIn  ·  4) Caselle email in riscaldamento (10-14 giorni)."),
]
r = 3
for k, v in rows:
    if k == "":
        r += 1
        continue
    ws.cell(row=r, column=1, value=k).font = f_h1
    ws.cell(row=r, column=1).alignment = left_top
    ws.cell(row=r, column=1).fill = fill_light
    ws.cell(row=r, column=1).border = border
    ws.cell(row=r, column=2, value=v).font = f_b
    ws.cell(row=r, column=2).alignment = left_top
    ws.cell(row=r, column=2).border = border
    ws.row_dimensions[r].height = 52
    r += 1

# ============================================================
# TAB 2 - SEQUENZA LINKEDIN + EMAIL
# ============================================================
ws = wb.create_sheet("Sequenza LinkedIn+Email")
ws.sheet_view.showGridLines = False
title_block(ws, "Sequenza coordinata LinkedIn + Email", "Una sola cadenza sullo stesso contatto. Qualsiasi risposta interrompe la sequenza e passa al contatto manuale.", span=6)

widths = [10, 14, 13, 40, 30, 26]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

hr = 4
headers = ["Step", "Canale", "Quando", "Azione", "Se RISPONDE", "Se NON risponde"]
for c, h in enumerate(headers, 1):
    ws.cell(row=hr, column=c, value=h)
style_header_row(ws, hr, 6)
ws.row_dimensions[hr].height = 22

seq = [
    ("1", "LinkedIn", "Giorno 0", "Richiesta di connessione SENZA nota (tasso di accettazione più alto).", "—", "Se non accettata entro 7 giorni: salta direttamente all'Email 1 (Step 4)."),
    ("2", "LinkedIn", "Giorno 0 (+ qualche ora dall'accettazione)", "Messaggio 1: presentazione breve + leva di valore + caso Promove + proposta di 15 min.", "Annotata su foglio (col. LI_Msg1_Risposta) → passa a contatto manuale/setter.", "Attendi 2 giorni, poi Step 3."),
    ("3", "LinkedIn", "Giorno +2", "Messaggio 2: follow-up con leva 'software di vostra proprietà, nessun lock-in'.", "Annotata su foglio (col. LI_Msg2_Risposta) → passa a contatto manuale/setter.", "Attendi 3 giorni, poi Step 4."),
    ("4", "Email", "Giorno +5", "Email 1 (da casella algòmera.pro, firma Elia Sorrentino): hook pain point di settore + Promove + proposta call con Francesco.", "Annotata su foglio (col. Email1_Risposta) → passa a contatto manuale/setter.", "Attendi 4 giorni, poi Step 5."),
    ("5", "Email", "Giorno +9", "Email 2 (break-up): referenze (Sofidel, Dompé, Openjobmetis) + porta lasciata aperta.", "Annotata su foglio (col. Email2_Risposta) → passa a contatto manuale/setter.", "Sequenza chiusa. Lead marcato 'nurturing' per ripresa futura."),
]
r = hr + 1
for i, row in enumerate(seq):
    fillc = fill_white if i % 2 == 0 else fill_light
    for c, val in enumerate(row, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = f_b if c != 1 else f_bb
        cell.alignment = wrap if c != 1 else center
        cell.fill = fillc
        cell.border = border
    ws.row_dimensions[r].height = 70
    r += 1

r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
note = ("REGOLA CHIAVE: la sequenza è automatica solo finché c'è silenzio. Alla prima risposta su qualsiasi canale, "
        "il contatto esce dalla sequenza, l'esito viene scritto sul foglio condiviso e la conversazione passa a una persona. "
        "Durata totale ciclo: 9 giorni. Le chiamate dei setter (vedi tab dedicato) corrono in parallelo e non dipendono da questa cadenza.")
ws.cell(row=r, column=1, value=note).font = Font(name="Calibri", size=10, bold=True, color=NAVY)
ws.cell(row=r, column=1).alignment = left_top
ws.cell(row=r, column=1).fill = fill_green
ws.row_dimensions[r].height = 56
for c in range(1, 7):
    ws.cell(row=r, column=c).border = border

# ============================================================
# TAB 3 - CHIAMATE (traccia separata)
# ============================================================
ws = wb.create_sheet("Chiamate (traccia separata)")
ws.sheet_view.showGridLines = False
title_block(ws, "Chiamate setter — binario parallelo", "Indipendente dalla sequenza digitale. Stessa lista, stesso foglio condiviso, esiti in tempo reale.", span=2)
ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 95

crows = [
    ("LISTA", "630 aziende in target già consegnate sul foglio Google condiviso. Chiamabili effettive: 599 (609 con telefono verificato, meno 10 in blacklist)."),
    ("PRIORITÀ", "Si parte dalla Lombardia, poi si estende a Campania, Lazio e resto Italia."),
    ("APERTURA", "Aggancio per ruolo: si chiede del decisore per nome dove presente (251 aziende), altrimenti 'mi passa il titolare/responsabile?'. Script telefonico già condiviso, adattato al tono Algòmera."),
    ("MESSAGGIO", "Software su misura e di vostra proprietà, senza licenze, sviluppato sul processo reale. Si qualifica subito: dimensioni, fatturato, ruolo del contatto, comprensione del servizio. Obiettivo: fissare videocall con Francesco."),
    ("QUALIFICA (kill criteria)", "11-50 dipendenti · fatturato ≥ 2 mln · interlocutore decisore · ha capito che è sviluppo custom. Se manca anche uno solo: non si fissa, si segna come fuori target. Meglio zero che un appuntamento sprecato."),
    ("ESITI", "Tracciati sul foglio con dropdown (es. Appuntamento fissato, Da richiamare, Non interessato, Fuori target, Non risponde). Note brevi e professionali."),
    ("ANTI NO-SHOW", "Appuntamenti fissati max ~2 settimane avanti. Reminder pre-call (chi siamo / cosa facciamo + caso Promove) a T-2 giorni e T-2 ore per ridurre i mancati appuntamenti."),
    ("INDIPENDENZA", "Le chiamate NON aspettano LinkedIn/email e viceversa. Un'azienda può essere in sequenza digitale e contemporaneamente in coda chiamate; alla prima risposta su un canale si coordina tutto sul foglio per non sovrapporsi."),
]
r = 3
for k, v in crows:
    ws.cell(row=r, column=1, value=k).font = f_h1
    ws.cell(row=r, column=1).alignment = left_top
    ws.cell(row=r, column=1).fill = fill_light
    ws.cell(row=r, column=1).border = border
    ws.cell(row=r, column=2, value=v).font = f_b
    ws.cell(row=r, column=2).alignment = left_top
    ws.cell(row=r, column=2).border = border
    ws.row_dimensions[r].height = 52
    r += 1

# ============================================================
# TAB 4 - SCRIPT
# ============================================================
ws = wb.create_sheet("Script")
ws.sheet_view.showGridLines = False
title_block(ws, "Script — copy LinkedIn ed Email", "Da approvare prima del lancio. Tono consulenziale, niente vendite aggressive. [Nome] e [settore] sono personalizzati per contatto.", span=2)
ws.column_dimensions["A"].width = 26
ws.column_dimensions["B"].width = 99

def script_block(ws, r, label, body):
    ws.cell(row=r, column=1, value=label).font = f_bb
    ws.cell(row=r, column=1).fill = fill_blue
    ws.cell(row=r, column=1).font = f_th
    ws.cell(row=r, column=1).alignment = left_top
    ws.cell(row=r, column=1).border = border
    ws.cell(row=r, column=2, value=body).font = f_b
    ws.cell(row=r, column=2).alignment = left_top
    ws.cell(row=r, column=2).border = border
    # altezza ~ righe
    lines = body.count("\n") + max(1, len(body) // 95)
    ws.row_dimensions[r].height = max(60, 16 * (lines + 1))
    return r + 1

li1 = ("Buongiorno [Nome], grazie per il collegamento.\n\n"
       "Lavoro in Algòmera, software house con sedi a Milano, Rovigo e Brindisi. Aiutiamo aziende come la vostra a sostituire i fogli Excel e i passaggi manuali tra produzione, magazzino e ufficio con un software proprietario, sviluppato sul vostro processo reale e senza licenze ricorrenti.\n\n"
       "Per un cliente nel [settore] abbiamo ridotto il processo preventivi da 3 giorni a 2 ore.\n\n"
       "Le farebbe senso uno scambio di 15 minuti per capire se può valere anche per voi? Senza impegno.")

li2 = ("[Nome], le rilancio brevemente.\n\n"
       "Capisco i tempi. Le lascio un punto che spesso interessa chi è nella sua posizione: i progetti che sviluppiamo restano interamente di vostra proprietà, codice sorgente compreso. Nessun abbonamento, nessun vincolo con il fornitore.\n\n"
       "Se le va un confronto rapido, anche solo per vedere come stiamo lavorando con realtà simili alla vostra, mi indichi un giorno questa o la prossima settimana. Altrimenti nessun problema.")

em1_subj = ("OGGETTO (3 varianti da testare):\n"
            "  •  Domanda veloce su [Nome Azienda]\n"
            "  •  Preventivi in [settore] in 2 ore?\n"
            "  •  Software su misura per [Nome Azienda]")

em1 = ("Buongiorno [Nome],\n\n"
       "sono Elia di Algòmera. Le scrivo perché nelle aziende del vostro settore vediamo spesso lo stesso schema: produzione, ufficio tecnico e magazzino lavorano su strumenti diversi (Excel, gestionali standard, email) e ogni passaggio porta via tempo, genera errori e rallenta la pianificazione.\n\n"
       "Noi sviluppiamo software 100% su misura, di vostra proprietà, costruito sui vostri processi reali. Niente licenze, niente canoni. Per Promove abbiamo ridotto il processo di preventivazione da 3 giorni a 2 ore.\n\n"
       "Se è un tema che vi tocca, le propongo 15 minuti di videocall con Francesco Buongiorno del nostro team, solo per capire se ha senso approfondire.\n\n"
       "Cosa ne pensa?\n\n"
       "Un saluto,\nElia Sorrentino\nAlgòmera — Milano · Rovigo · Brindisi\n[link Calendly]")

em2_subj = "OGGETTO:  Ultimo messaggio, [Nome]"

em2 = ("[Nome],\n\n"
       "non avendo ricevuto riscontro al messaggio precedente, immagino che il momento non sia quello giusto.\n\n"
       "Le lascio comunque un riferimento: tra i progetti che abbiamo consegnato ci sono realtà come Sofidel, Dompé e Openjobmetis, con il codice sorgente sempre di proprietà del cliente. Se in futuro vi servirà un software che parli davvero con il vostro processo, ci trovate qui.\n\n"
       "Le auguro buon lavoro.\nElia")

r = 3
r = script_block(ws, r, "LinkedIn — Messaggio 1 (dopo accettazione)", li1)
r = script_block(ws, r, "LinkedIn — Messaggio 2 (+2 giorni, se silenzio)", li2)
r = script_block(ws, r, "Email 1 — Oggetto", em1_subj)
r = script_block(ws, r, "Email 1 — Corpo (+5 giorni)", em1)
r = script_block(ws, r, "Email 2 — Oggetto", em2_subj)
r = script_block(ws, r, "Email 2 — Corpo, break-up (+9 giorni)", em2)

r += 1
ws.cell(row=r, column=1, value="PERSONALIZZAZIONE [settore]").font = f_h1
ws.cell(row=r, column=1).fill = fill_light
ws.cell(row=r, column=1).alignment = left_top
ws.cell(row=r, column=1).border = border
hook = ("Il segnaposto [settore] e l'hook iniziale si adattano in base ai pain point del contatto:\n"
        "  •  EDILIZIA/COSTRUZIONI — informazioni che viaggiano lente tra ufficio tecnico, amministrazione e cantieri; nessun controllo in tempo reale sugli avanzamenti.\n"
        "  •  MECCANICA/PRODUZIONE — reparti e sistemi non integrati; errori e rilavorazioni da dati non aggiornati; scarsa visibilità sull'avanzamento produttivo.\n"
        "  •  LOGISTICA/TRASPORTI — nessuna visibilità in tempo reale su spedizioni e flotte; processi frammentati tra ufficio, magazzino e autisti.\n"
        "  •  PLASTICA/GOMMA — gestione commesse e magazzino su strumenti scollegati; passaggi manuali tra produzione e amministrazione.\n"
        "  •  FOOD/DISTRIBUZIONE — tracciabilità e gestione ordini dispersa tra gestionali ed Excel; carico amministrativo manuale elevato.")
ws.cell(row=r, column=2, value=hook).font = f_b
ws.cell(row=r, column=2).alignment = left_top
ws.cell(row=r, column=2).border = border
ws.row_dimensions[r].height = 150

# ============================================================
# TAB 5 - TARGET / ICP
# ============================================================
ws = wb.create_sheet("Target ICP Sales Navigator")
ws.sheet_view.showGridLines = False
title_block(ws, "Target & filtri Sales Navigator", "Profili da contattare su LinkedIn. Stesso ICP della lista chiamate.", span=2)
ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 95

icp = [
    ("GEOGRAFIA", "Italia. Priorità Lombardia, poi Campania e Lazio, quindi resto del territorio."),
    ("DIMENSIONE (Headcount)", "11-50 dipendenti. Sotto gli 11: budget insufficiente. Sopra i 50: trattative troppo lunghe."),
    ("FATTURATO", "≥ 2 milioni di euro."),
    ("SETTORI INCLUSI", "Manifatturiero / Industrial Machinery, Costruzioni, Logistica & Supply Chain, Trasporti, Produzione (food, plastica, gomma, tessile), Wholesale."),
    ("SETTORI ESCLUSI", "Information Technology, Computer Software, Internet, Marketing & Advertising, Design. Sono competitor o senza budget adeguato."),
    ("RUOLI DECISORE", "CEO, Founder, Owner, Titolare, Amministratore Delegato, Amministratore Unico, Direttore Generale, Direttore Operations, Direttore Acquisti, Responsabile Produzione/Logistica."),
    ("SENIORITY", "Owner, Partner, CXO, VP, Director. Esclusi Manager e Specialist senza potere decisionale."),
    ("VOLUMI SOSTENIBILI", "80-100 richieste di connessione a settimana per profilo (~400/mese). Oltre 120/settimana LinkedIn inizia a limitare l'account. 50 messaggi/settimana."),
    ("FONTE CONTATTI", "Si parte dai decisori già presenti nella lista lead (priorità Lombardia), si amplia via Sales Navigator con gli stessi filtri per alimentare il volume mensile. Aziende in blacklist sempre escluse."),
]
r = 3
for k, v in icp:
    fillk = fill_red if "ESCLUSI" in k else fill_light
    ws.cell(row=r, column=1, value=k).font = f_h1
    ws.cell(row=r, column=1).alignment = left_top
    ws.cell(row=r, column=1).fill = fillk
    ws.cell(row=r, column=1).border = border
    ws.cell(row=r, column=2, value=v).font = f_b
    ws.cell(row=r, column=2).alignment = left_top
    ws.cell(row=r, column=2).border = border
    ws.row_dimensions[r].height = 46
    r += 1

# ============================================================
# TAB 6 - COORDINAMENTO & TIMELINE
# ============================================================
ws = wb.create_sheet("Coordinamento e Timeline")
ws.sheet_view.showGridLines = False
title_block(ws, "Coordinamento operativo & avvio", "Chi fa cosa, accessi, tracciamento, tempistiche.", span=2)
ws.column_dimensions["A"].width = 32
ws.column_dimensions["B"].width = 93

co = [
    ("RUOLI — TELESALES", "Setter per le chiamate + gestione outreach LinkedIn ed email. Aggiornamento del foglio condiviso in tempo reale."),
    ("RUOLI — ALGÒMERA", "Francesco Buongiorno: conduce gli appuntamenti di vendita (videocall). Valentina Spagnulo: referente unica del progetto."),
    ("FOGLIO CONDIVISO", "Si aggiungono 6 colonne stato per il tracciamento omnicanale: LI_Connessione, LI_Msg1_Risposta, LI_Msg2_Risposta, Email1_Risposta, Email2_Risposta, Esito_Chiamata. Tutto visibile in tempo reale a entrambe le parti."),
    ("ALLINEAMENTO", "Call ricorrente di 30 minuti a settimana (Google Chat) con Valentina e Francesco per rivedere insieme il foglio e tarare messaggi e target. Niente gruppo WhatsApp (policy Algòmera)."),
]
r = 3
for k, v in co:
    ws.cell(row=r, column=1, value=k).font = f_h1
    ws.cell(row=r, column=1).alignment = left_top
    ws.cell(row=r, column=1).fill = fill_light
    ws.cell(row=r, column=1).border = border
    ws.cell(row=r, column=2, value=v).font = f_b
    ws.cell(row=r, column=2).alignment = left_top
    ws.cell(row=r, column=2).border = border
    ws.row_dimensions[r].height = 48
    r += 1

# Sales Navigator access options
r += 1
ws.cell(row=r, column=1, value="ACCESSI SALES NAVIGATOR — opzioni").font = Font(name="Calibri", size=13, bold=True, color=NAVY)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
r += 1
sn = [
    ("Opzione A — profilo di Francesco (raccomandata)", "Francesco ci fornisce accesso al suo profilo LinkedIn (con Sales Navigator attivo) e autorizza il nostro dispositivo al primo accesso. Operiamo da un unico IP italiano fisso. Vantaggio: il prospect ritrova in call la stessa persona vista su LinkedIn."),
    ("Opzione B — seat dedicato", "Algòmera aggiunge un seat Sales Navigator per noi (utente Telesales). Più separato, ma il prospect non vede Francesco e si perde la coerenza LinkedIn-call."),
    ("Opzione C — tool di delega", "Strumento terzo che opera senza condividere password. Setup tecnico ~1 settimana e costo aggiuntivo."),
]
for k, v in sn:
    ws.cell(row=r, column=1, value=k).font = f_bb
    ws.cell(row=r, column=1).alignment = left_top
    ws.cell(row=r, column=1).fill = fill_green if "raccomandata" in k else fill_white
    ws.cell(row=r, column=1).border = border
    ws.cell(row=r, column=2, value=v).font = f_b
    ws.cell(row=r, column=2).alignment = left_top
    ws.cell(row=r, column=2).border = border
    ws.row_dimensions[r].height = 56
    r += 1

# Timeline
r += 1
ws.cell(row=r, column=1, value="TEMPISTICHE DI AVVIO").font = Font(name="Calibri", size=13, bold=True, color=NAVY)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
r += 1
tl = [
    ("Oggi", "Call strategica: accessi Sales Navigator, approvazione copy, conferma volumi."),
    ("Da domani", "Riscaldamento delle 3 caselle algòmera.pro (10-14 giorni) per proteggere la deliverability."),
    ("Lunedì 15/06", "Partono le chiamate sulla Lombardia + le richieste di connessione LinkedIn (appena ricevuti gli accessi)."),
    ("Da ~29/06", "Partono le prime email a caselle riscaldate, in volumi mirati e crescenti."),
]
for k, v in tl:
    ws.cell(row=r, column=1, value=k).font = f_bb
    ws.cell(row=r, column=1).alignment = left_top
    ws.cell(row=r, column=1).fill = fill_light
    ws.cell(row=r, column=1).border = border
    ws.cell(row=r, column=2, value=v).font = f_b
    ws.cell(row=r, column=2).alignment = left_top
    ws.cell(row=r, column=2).border = border
    ws.row_dimensions[r].height = 38
    r += 1

# Punti da confermare
r += 1
ws.cell(row=r, column=1, value="DA CONFERMARE IN CALL").font = Font(name="Calibri", size=13, bold=True, color="C00000")
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
r += 1
conf = ("1) Il caso Promove è citabile per nome nei messaggi? Sofidel / Dompé / Openjobmetis sono citabili (già pubblici sul sito)?\n"
        "2) Volume LinkedIn 80-100 connessioni/settimana confermato?\n"
        "3) Modalità accessi Sales Navigator (opzione A / B / C)?\n"
        "4) Conferma che il pagamento è partito (necessario per l'avvio operativo).")
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
ws.cell(row=r, column=1, value=conf).font = f_b
ws.cell(row=r, column=1).alignment = left_top
ws.cell(row=r, column=1).fill = fill_red
ws.cell(row=r, column=1).border = border
ws.row_dimensions[r].height = 90

# Page setup: fit-to-width così la stampa/PDF non spezza le colonne testo
for sh in wb.worksheets:
    sh.page_setup.orientation = "landscape"
    sh.page_setup.fitToWidth = 1
    sh.page_setup.fitToHeight = 0
    sh.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    sh.page_margins = openpyxl.worksheet.page.PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)

out = "/Users/simocors/Desktop/telesales/Strategia_Algomera_Telesales.xlsx"
wb.save(out)
print("salvato:", out)
print("tab:", wb.sheetnames)
