#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Finalizza la lista Look: applica gli update telefoni (R2) e nomi (titolare/acquisti)
a prospects.json, rigenera il Master col builder standard, poi aggiunge 2 colonne
("Titolare (nome)", "Resp. acquisti (nome)") a LISTA_PROSPECT e DB_COMPLETO.

USO: python3 look/finalize_look_list.py
Legge gli output dei workflow dai path passati (o di default) e aggiorna i file.
"""
import os, re, json, sys, subprocess, unicodedata
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

ROOT = "/Users/simocors/Desktop/telesales"
TASKS = "/private/tmp/claude-501/-Users-simocors-Desktop-telesales/9d3cec9f-a672-420f-83a4-f63d51ec6d5d/tasks"
COMBO_OUT = os.path.join(TASKS, "wpjrvgbil.output")   # enrichment combinato (telefoni + nomi)
PROSPECTS = os.path.join(ROOT, "look/prospects.json")

def digits(s): return re.sub(r"[^0-9]", "", str(s or ""))
def norm(s):
    s = unicodedata.normalize("NFKD", str(s or "")).encode("ascii", "ignore").decode().lower()
    return re.sub(r"[^a-z0-9]+", " ", s).strip()
def valid_fixed(s):
    d = digits(s); return d.startswith("0") and 6 <= len(d) <= 11
def valid_any(s):
    d = digits(s)
    if d.startswith(("800","199","892","899","840","848")): return False
    return (d.startswith("0") and 6 <= len(d) <= 11) or (d.startswith("3") and 9 <= len(d) <= 11)
def load_result(path):
    d = json.loads(open(path, encoding="utf-8").read())
    res = d.get("result")
    if isinstance(res, str): res = json.loads(res)
    return res
def clean_name(s):
    s = re.sub(r"\s+", " ", str(s or "")).strip()
    # scarta valori non-nome (ruoli generici, vuoti, troppo corti/lunghi)
    if len(s) < 4 or len(s) > 60: return ""
    low = s.lower()
    if low in ("n/d","nd","na","n.a.","sconosciuto","non disponibile","titolare","amministratore","-"): return ""
    if not re.search(r"[a-zA-Zàèéìòù]{2,}\s+[a-zA-Zàèéìòù]{2,}", s): return ""  # serve nome+cognome
    return s

m = json.load(open(PROSPECTS, encoding="utf-8"))

# enrichment combinato: telefoni + nomi
addc = addd = addt = adda = 0
if os.path.exists(COMBO_OUT) and os.path.getsize(COMBO_OUT) > 0:
    try:
        for u in load_result(COMBO_OUT).get("results", []):
            i = u.get("idx")
            if not isinstance(i, int) or not (0 <= i < len(m)): continue
            r = m[i]
            if len(digits(r.get("telefono_centralino"))) < 6 and valid_fixed(u.get("telefono_centralino")):
                r["telefono_centralino"] = u["telefono_centralino"]; addc += 1
            if len(digits(r.get("telefono_diretto"))) < 6:
                nd = u.get("telefono_diretto", "")
                if valid_any(nd) and digits(nd) != digits(r.get("telefono_centralino")):
                    r["telefono_diretto"] = nd; addd += 1
            t = clean_name(u.get("titolare_nome")); a = clean_name(u.get("acquisti_nome"))
            if t and not r.get("titolare_nome"): r["titolare_nome"] = t; addt += 1
            if a and not r.get("acquisti_nome"): r["acquisti_nome"] = a; adda += 1
        print(f"combo applicato: +{addc} centralino, +{addd} diretti, +{addt} titolare, +{adda} resp. acquisti")
    except Exception as e:
        print("combo: output non pronto:", e); raise SystemExit(1)
else:
    print("combo: output assente o vuoto, NON finalizzo (attendere fine workflow)."); raise SystemExit(1)

for r in m:
    r.setdefault("titolare_nome", ""); r.setdefault("acquisti_nome", "")
json.dump(m, open(PROSPECTS, "w", encoding="utf-8"), ensure_ascii=False, indent=0)

# 3) rigenera Master col builder standard
subprocess.run(["python3", os.path.join(ROOT, "_KIT_CLIENTE/builders/build_master.py"), os.path.join(ROOT, "look")], check=True)

# 4) aggiungi colonne nomi (append in coda) a LISTA_PROSPECT e DB_COMPLETO
name_by_az = {norm(r["azienda"]): (r.get("titolare_nome",""), r.get("acquisti_nome","")) for r in m}
fpath = os.path.join(ROOT, "look/Master_Look_SRLS.xlsx")
wb = openpyxl.load_workbook(fpath)
HDR = PatternFill("solid", fgColor="1F3864"); HF = Font(color="FFFFFF", bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")
fillmap = {"A": PatternFill("solid", fgColor="C8E6C9"), "B": PatternFill("solid", fgColor="FFF59D"), "C": PatternFill("solid", fgColor="F2F2F2")}

def add_cols(ws, hdr_row, az_col, prio_col):
    c0 = ws.max_column
    ws.cell(hdr_row, c0+1, "Titolare (nome)").fill = HDR; ws.cell(hdr_row, c0+1).font = HF
    ws.cell(hdr_row, c0+2, "Resp. acquisti (nome)").fill = HDR; ws.cell(hdr_row, c0+2).font = HF
    for r in range(hdr_row+1, ws.max_row+1):
        az = norm(ws.cell(r, az_col).value)
        t, a = name_by_az.get(az, ("", ""))
        prio = ws.cell(r, prio_col).value if prio_col else "B"
        fill = fillmap.get(prio, fillmap["B"])
        for off, val in ((1, t), (2, a)):
            cell = ws.cell(r, c0+off, val); cell.fill = fill; cell.alignment = WRAP
    ws.column_dimensions[openpyxl.utils.get_column_letter(c0+1)].width = 24
    ws.column_dimensions[openpyxl.utils.get_column_letter(c0+2)].width = 24

# LISTA_PROSPECT: header riga 4, Azienda col 2, Prio col 3
add_cols(wb["LISTA_PROSPECT"], 4, 2, 3)
# DB_COMPLETO: header riga 4, Azienda col 3, Prio col 2
add_cols(wb["DB_COMPLETO"], 4, 3, 2)
wb.save(fpath)

# recap
withc = sum(1 for r in m if len(digits(r.get("telefono_centralino"))) >= 6)
withd = sum(1 for r in m if len(digits(r.get("telefono_diretto"))) >= 6)
witht = sum(1 for r in m if r.get("titolare_nome"))
witha = sum(1 for r in m if r.get("acquisti_nome"))
print(f"\nTOTALE {len(m)} | centralino {withc} | diretto {withd} | titolare {witht} | resp.acquisti {witha}")
print("Master aggiornato:", fpath)
