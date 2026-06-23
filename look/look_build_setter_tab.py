#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Costruisce la scheda "SETTER - DA CHIUDERE" nel Master Look: lista corta dei lead
caldi/tiepidi da far richiamare al setter per chiudere l'appuntamento. Colora anche
gli esiti in LISTA_PROSPECT. Poi ricarica il Master su Drive.

Usabile standalone:  python3 look/look_build_setter_tab.py
Oppure importato dall'analyzer (build_setter_tab(path) + push).
"""
import os, re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
MASTER = os.path.join(ROOT, "look", "Master_Look_SRLS.xlsx")
DRIVE_FILE_ID = "1-zrThfta_mZSL40iGYDLIjwOc5cF3oQv"
SETTER_TAB = "SETTER - DA CHIUDERE"

# colonne LISTA_PROSPECT (1-based)
L_AZ, L_CENT, L_DIR, L_RUOLO, L_ESITO, L_NOTE, L_APP, L_TIT = 2, 10, 11, 12, 21, 22, 23, 24

CALDO = {"Interessato", "Referente individuato", "Appuntamento fissato", "Da ricontattare piu avanti"}
PROSSIMO = {
    "Interessato": "Era interessato: richiama e fissa la consulenza col consulente.",
    "Referente individuato": "Chiedi del referente/titolare per nome e fissa la consulenza.",
    "Appuntamento fissato": "Conferma l'appuntamento e passa il lead a Riccardo.",
    "Da ricontattare piu avanti": "Richiama come da nota; valuta tempi e manda materiale se serve.",
    "Da richiamare": "Richiama il decisore (vedi nota per orario) e proponi la consulenza.",
}
F_HDR = Font(color="FFFFFF", bold=True)
C_HDR = PatternFill("solid", fgColor="1F3864")
C_CALDO = PatternFill("solid", fgColor="C8E6C9")   # verde
C_TIEPIDO = PatternFill("solid", fgColor="FFF59D")  # giallo
C_CHIUSO = PatternFill("solid", fgColor="F2F2F2")   # grigio
WRAP = Alignment(wrap_text=True, vertical="top")
TOP = Alignment(vertical="top")

def digits(s): return re.sub(r"[^0-9]", "", str(s or ""))

def build_setter_tab(path=MASTER):
    wb = openpyxl.load_workbook(path)
    ws = wb["LISTA_PROSPECT"]

    # 1) raccogli caldi/tiepidi e traccia le righe warm
    rows = []
    warm_rows = []  # (riga, prio)
    for r in range(5, ws.max_row + 1):
        az = ws.cell(r, L_AZ).value
        if not az:
            continue
        esito = (ws.cell(r, L_ESITO).value or "").strip()
        nota = (ws.cell(r, L_NOTE).value or "").strip()
        # SOLO lead CALDI nel foglio setter. I tiepidi ("Da richiamare") NON ci vanno mai.
        if esito in CALDO:
            prio = "CALDO"
            warm_rows.append((r, prio))
            tel = ws.cell(r, L_DIR).value or "" if digits(ws.cell(r, L_DIR).value) else (ws.cell(r, L_CENT).value or "")
            ref = (ws.cell(r, L_TIT).value or "").strip() or (ws.cell(r, L_RUOLO).value or "").strip()
            rows.append({
                "prio": prio, "az": az, "tel": tel, "ref": ref, "esito": esito,
                "nota": nota, "prossimo": PROSSIMO.get(esito, "Richiama e proponi la consulenza."),
                "app": ws.cell(r, L_APP).value or "",
            })
    rows.sort(key=lambda x: (0 if x["prio"] == "CALDO" else 1, x["az"]))

    # DECOLORA tutte le righe dati, poi colora SOLO i warm (caldo verde, tiepido giallo)
    NOFILL = PatternFill(fill_type=None)
    maxc = ws.max_column
    for rr in range(5, ws.max_row + 1):
        for cc in range(1, maxc + 1):
            ws.cell(rr, cc).fill = NOFILL
    for rr, prio in warm_rows:
        fill = C_CALDO if prio == "CALDO" else C_TIEPIDO
        for cc in range(1, maxc + 1):
            ws.cell(rr, cc).fill = fill

    # 2) (ri)crea la scheda SETTER
    if SETTER_TAB in wb.sheetnames:
        del wb[SETTER_TAB]
    sh = wb.create_sheet(SETTER_TAB, 0)  # in cima
    sh["A1"] = "SETTER - DA CHIUDERE"
    sh["A1"].font = Font(bold=True, size=14, color="1F3864")
    sh["A2"] = (f"{len(rows)} lead caldi (interessati, referenti individuati, da ricontattare). "
                "Richiamali per chiudere l'appuntamento.")
    sh["A2"].font = Font(italic=True, size=10, color="555555")
    cols = ["Priorita", "Azienda", "Telefono da chiamare", "Referente / Titolare", "Esito AI",
            "Cosa e successo", "Cosa fare adesso", "Setter", "Esito setter", "Appuntamento (data/ora)"]
    hdr = 4
    for j, c in enumerate(cols, 1):
        cell = sh.cell(hdr, j, c); cell.fill = C_HDR; cell.font = F_HDR; cell.alignment = WRAP
    r = hdr + 1
    for x in rows:
        vals = [x["prio"], x["az"], x["tel"], x["ref"], x["esito"], x["nota"], x["prossimo"], "", "", x["app"]]
        fill = C_CALDO if x["prio"] == "CALDO" else C_TIEPIDO
        for j, v in enumerate(vals, 1):
            cell = sh.cell(r, j, v); cell.fill = fill
            cell.alignment = WRAP if j in (6, 7) else TOP
        sh.cell(r, 3).number_format = "@"
        r += 1
    last = r - 1
    # dropdown esito setter
    if last >= hdr + 1:
        dv = DataValidation(type="list", formula1='"Appuntamento fissato,Richiamato - interessato,Richiamato - da risentire,Non interessato,Non risposto"', allow_blank=True)
        sh.add_data_validation(dv); dv.add(f"I{hdr+1}:I{last}")
    W = {"A": 10, "B": 32, "C": 18, "D": 24, "E": 20, "F": 60, "G": 42, "H": 16, "I": 26, "J": 20}
    for k, v in W.items():
        sh.column_dimensions[k].width = v
    sh.freeze_panes = "A5"
    if last >= hdr:
        sh.auto_filter.ref = f"A{hdr}:J{last}"
    wb.save(path)
    return len(rows), sum(1 for x in rows if x["prio"] == "CALDO")

def push_drive(path=MASTER):
    from google.oauth2.service_account import Credentials
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaFileUpload
    creds = Credentials.from_service_account_file(os.path.join(ROOT, "service-account.json"),
            scopes=["https://www.googleapis.com/auth/drive"])
    build("drive", "v3", credentials=creds).files().update(
        fileId=DRIVE_FILE_ID,
        media_body=MediaFileUpload(path, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        supportsAllDrives=True).execute()

if __name__ == "__main__":
    n, caldi = build_setter_tab()
    print(f"Scheda '{SETTER_TAB}' creata: {n} lead ({caldi} caldi). Colorati gli esiti in LISTA_PROSPECT.")
    push_drive()
    print("Master ricaricato su Drive.")
