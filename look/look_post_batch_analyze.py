#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Look — Analyzer post-chiamata. Legge le conversazioni dell'agente Marco-Look su
ElevenLabs, classifica l'esito e scrive una NOTA in linguaggio umano nel Master
(colonna Note), poi ricarica il Master aggiornato sul file Drive.

Regole NOTA (come richiesto):
- Non risposto / fallita / segreteria: nota BREVE ("Non risposto." / "Risponde la
  segreteria, richiamare.").
- Conversazione reale: nota RICCA e umana — con CHI si e parlato, gestore attuale,
  modalita, interesse, se manda la fattura, eventuale appuntamento, prossimo passo.
  Mai termini tecnici/AI.

USO:
    python3 look/look_post_batch_analyze.py --dry-run        # mostra, non scrive
    python3 look/look_post_batch_analyze.py                  # scrive Master + Drive
    python3 look/look_post_batch_analyze.py --hours 12       # solo ultime 12h
"""
import os, re, json, sys, argparse, datetime, unicodedata, urllib.request, urllib.parse
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
MASTER = os.path.join(ROOT, "look", "Master_Look_SRLS.xlsx")
DRIVE_FILE_ID = "1-zrThfta_mZSL40iGYDLIjwOc5cF3oQv"
LOOK_AGENT_ID = "agent_5901kvb1atsaf83txkq8tq62x77m"
BASE = "https://api.elevenlabs.io"
# colonne LISTA_PROSPECT (1-based) — header reale: 21=Esito, 22=Note, 23=Appuntamento
C_AZIENDA, C_ESITO, C_NOTE, C_APPUNT = 2, 21, 22, 23
HDR_ROW = 4

def key():
    return os.getenv("ELEVENLABS_API_KEY") or re.search(r'"(sk_[a-z0-9]{40,})"',
        open(os.path.join(ROOT, "culligan_batch_caller.py")).read()).group(1)
H = {"xi-api-key": key()}
def api(path, params=None):
    url = BASE + path
    if params:
        url += "?" + urllib.parse.urlencode(params)
    return json.loads(urllib.request.urlopen(urllib.request.Request(url, headers=H), timeout=30).read())

def norm(s):
    s = unicodedata.normalize("NFKD", str(s or "")).encode("ascii", "ignore").decode().lower()
    return re.sub(r"[^a-z0-9]+", " ", s).strip()
def gv(dc, k):
    v = (dc.get(k) or {})
    return v.get("value") if isinstance(v, dict) else v

def list_convs(hours):
    out, cursor = [], None
    after = int((datetime.datetime.now() - datetime.timedelta(hours=hours)).timestamp())
    while True:
        p = {"agent_id": LOOK_AGENT_ID, "page_size": 100}
        if cursor: p["cursor"] = cursor
        r = api("/v1/convai/conversations", p)
        for c in r.get("conversations", []):
            if c.get("start_time_unix_secs", 0) >= after:
                out.append(c)
        if r.get("has_more") and r.get("next_cursor"):
            cursor = r["next_cursor"]
        else:
            break
    return out

def build_nota_esito(conv):
    dc = (conv.get("analysis") or {}).get("data_collection_results", {}) or {}
    status = conv.get("status", "")
    transcript = conv.get("transcript", []) or []
    note_ai = (gv(dc, "note_ai") or "").strip()
    esito = (gv(dc, "esito") or "").strip()
    appun = gv(dc, "appuntamento_fissato")
    fattura = gv(dc, "fattura_promessa")
    decisore = gv(dc, "decisore_raggiunto")
    interest = (gv(dc, "interest_level") or "").strip()
    gestore = (gv(dc, "gestore_attuale") or "").strip()
    modalita = (gv(dc, "modalita_attuale") or "").strip()
    referente = (gv(dc, "nome_referente") or "").strip()
    data_app = (gv(dc, "data_appuntamento") or "").strip()
    optout = gv(dc, "opt_out_richiesto")

    failed = status in ("failed", "error") or not transcript
    # esito fallback
    if not esito:
        if failed: esito = "Non risposto"
        elif appun: esito = "Appuntamento fissato"
        elif fattura: esito = "Fattura richiesta"
        elif optout: esito = "Non interessato"
        elif interest == "high": esito = "Interessato"
        elif decisore is False: esito = "Da richiamare"
        else: esito = "Da richiamare"

    # nota
    if failed or esito == "Non risposto":
        nota = "Non risposto."
    elif esito == "Da richiamare" and (decisore is False or not note_ai):
        nota = note_ai or "Risponde il centralino, decisore non disponibile, da richiamare."
    else:
        if note_ai and len(note_ai) >= 35:
            nota = note_ai
        else:
            parts = []
            parts.append(f"Parlato con {referente}." if referente else "Parlato col referente.")
            if gestore: parts.append(f"Numero verde con {gestore}{(' (' + modalita + ')') if modalita and modalita!='non so' else ''}.")
            elif modalita and modalita != "non so": parts.append(f"Numero verde {modalita}.")
            if interest == "high": parts.append("Buon interesse.")
            elif interest == "medium": parts.append("Interesse tiepido.")
            if fattura: parts.append("Manda l'ultima fattura.")
            if appun and data_app: parts.append(f"Appuntamento {data_app}.")
            elif esito == "Referente individuato": parts.append("Richiamare per fissare.")
            nota = " ".join(parts)
    return esito, nota[:280], (data_app if appun else "")

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--hours", type=int, default=24)
    a = ap.parse_args()

    convs = list_convs(a.hours)
    print(f"Conversazioni Look ultime {a.hours}h: {len(convs)}")
    if not convs:
        return

    wb = openpyxl.load_workbook(MASTER)
    ws = wb["LISTA_PROSPECT"]
    rowmap = {}
    for r in range(HDR_ROW + 1, ws.max_row + 1):
        az = ws.cell(r, C_AZIENDA).value
        if az: rowmap.setdefault(norm(az), r)

    written = 0
    for c in convs:
        cid = c["conversation_id"]
        conv = api(f"/v1/convai/conversations/{cid}")
        dv = (conv.get("conversation_initiation_client_data") or {}).get("dynamic_variables", {}) or {}
        az = dv.get("nome_azienda", "")
        row = rowmap.get(norm(az))
        esito, nota, data_app = build_nota_esito(conv)
        tag = "" if row else " [NESSUN MATCH NEL MASTER]"
        print(f"  {az[:34]:34} -> {esito:24} | {nota}{tag}")
        if row and not a.dry_run:
            ws.cell(row, C_ESITO, esito)
            ws.cell(row, C_NOTE, nota)
            if data_app: ws.cell(row, C_APPUNT, data_app)
            written += 1

    if a.dry_run:
        print("\n[DRY-RUN] nulla scritto.")
        return
    if written:
        wb.save(MASTER)
        # aggiorna la scheda SETTER - DA CHIUDERE (caldi/tiepidi + colori) e ricarica su Drive (push unico)
        from look_build_setter_tab import build_setter_tab, push_drive
        n, caldi = build_setter_tab(MASTER)
        push_drive(MASTER)
        print(f"\nScritte {written} righe. Scheda SETTER aggiornata: {n} lead ({caldi} caldi). Master ricaricato su Drive.")

if __name__ == "__main__":
    import urllib.parse
    main()
