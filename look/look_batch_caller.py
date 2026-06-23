#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Look SRLS Batch Caller — collega il NOSTRO foglio all'agente AI Voice ElevenLabs
"Marco - Look SRLS Numero Verde". Per ogni riga del foglio costruisce le
dynamic_variables giuste e lancia la chiamata outbound (batch-calling API),
esattamente come il pipeline Culligan.

LOGICA VARIABILI (come richiesto):
- nome_azienda: ragione sociale
- categoria: settore
- citta: sede
- numero_verde: il numero verde 800 dell'azienda
- nome_contatto + decision_maker: CHI far chiedere a Marco
    * se conosciamo il titolare -> chiede del titolare per NOME ("il titolare")
    * altrimenti se conosciamo il resp. acquisti -> chiede di lui per nome ("il responsabile acquisti")
    * altrimenti chiede per RUOLO ("il titolare" / "il responsabile acquisti" / "il responsabile commerciale")
- numero da chiamare: il DIRETTO se presente e valido (raggiunge prima il decisore),
  altrimenti il centralino.

USO:
    python3 look/look_batch_caller.py --dry-run            # mostra cosa farebbe, NON chiama
    python3 look/look_batch_caller.py --limit 5            # lancia max 5 chiamate
    python3 look/look_batch_caller.py --source xlsx --dry-run   # legge il Master locale invece del foglio
"""
import os, re, sys, json, argparse, datetime

# ============================================================================
# CONFIGURAZIONE
# ============================================================================
ELEVENLABS_API_KEY = os.getenv("ELEVENLABS_API_KEY", "")  # se vuoto, lo carico da culligan_batch_caller.py
ELEVENLABS_BASE_URL = "https://api.elevenlabs.io"

LOOK_AGENT_ID = "agent_5901kvb1atsaf83txkq8tq62x77m"   # Marco - Look SRLS Numero Verde
# Numero outbound funzionante (+390554652406). Il secondo numero (+390574814928 / phnum_8001)
# NON e abilitato all'outbound: da SIP 404. Per non andare in contesa con Culligan sullo stesso
# trunk si usa CONCURRENCY=1 (footprint minimo, niente sovrapposizioni).
PHONE_NUMBER_ID = os.getenv("LOOK_PHONE_NUMBER_ID", "phnum_1501kr3sx76sfxeap503jqy1m7j9")
CONCURRENCY = int(os.getenv("LOOK_CONCURRENCY", "1"))   # 1 = nessuna sovrapposizione
EXCLUDE_PHONES = {re.sub(r"[^0-9+]", "", p) for p in os.getenv("LOOK_EXCLUDE_PHONES", "").split(",") if p.strip()}
# blocklist permanente numeri morti (SIP 404) — evita di richiamarli e generare SIP error
_bl = os.path.join(os.path.dirname(__file__), "blocklist_phones.txt")
if os.path.exists(_bl):
    EXCLUDE_PHONES |= {re.sub(r"[^0-9+]", "", p) for p in open(_bl).read().split(",") if p.strip()}

# registro chiamate: righe "data,slot,telefono" (vecchie "data,telefono"). Serve per i re-call mirati.
CALLED_LOG = os.path.join(os.path.dirname(__file__), "called_log.csv")
MAX_ATTEMPTS = 3   # numero massimo di tentativi AI su uno stesso numero no-contatto
def current_slot():
    import datetime
    return "AM" if datetime.datetime.now().hour < 14 else "PM"
def call_stats():
    import datetime
    today = datetime.date.today().isoformat()
    st = {}
    if not os.path.exists(CALLED_LOG):
        return st
    for line in open(CALLED_LOG):
        p = line.strip().split(",")
        if len(p) < 2:
            continue
        date = p[0]
        slot, ph = (p[1], p[2]) if len(p) >= 3 else ("?", p[1])
        ph = re.sub(r"[^0-9+]", "", ph)
        d = st.setdefault(ph, {"n": 0, "slots": set(), "today": False})
        d["n"] += 1
        d["slots"].add(slot)
        if date == today:
            d["today"] = True
    return st
def log_called(phones):
    import datetime
    d = datetime.date.today().isoformat(); slot = current_slot()
    with open(CALLED_LOG, "a") as f:
        for ph in phones:
            f.write(f"{d},{slot},{ph}\n")
RINGING_TIMEOUT = 75

# Foglio Google "creato da noi" — da popolare DOPO che lo condividi col service account:
#   claude-sheets@claude-telesales.iam.gserviceaccount.com
SHEET_ID = os.getenv("LOOK_SHEET_ID", "")   # <-- inserire l'ID del foglio Look
TAB_NAME = "CHIAMATE"
SA_PATH = os.path.join(os.path.dirname(os.path.dirname(__file__)), "service-account.json")
XLSX_PATH = os.path.join(os.path.dirname(__file__), "Master_Look_SRLS.xlsx")

# Colonne del tab CHIAMATE (0-based) — vedi look_push_to_sheet.py
C_AZIENDA, C_SETTORE, C_CITTA, C_VERDE, C_RUOLO = 0, 1, 2, 3, 4
C_TITOLARE, C_ACQUISTI, C_CENTRALINO, C_DIRETTO = 5, 6, 7, 8
C_ESITO, C_NOTE, C_APPUNT, C_DATA_CHIAMATA, C_CONV_ID = 9, 10, 11, 12, 13
NCOLS = 14

DEFAULT_LIMIT = 10

# ============================================================================
def get_key():
    global ELEVENLABS_API_KEY
    if ELEVENLABS_API_KEY:
        return ELEVENLABS_API_KEY
    src = open(os.path.join(os.path.dirname(__file__), "..", "culligan_batch_caller.py"), encoding="utf-8").read()
    ELEVENLABS_API_KEY = re.search(r'"(sk_[a-z0-9]{40,})"', src).group(1)
    return ELEVENLABS_API_KEY

def normalize_phone(p):
    if not p:
        return ""
    c = re.sub(r"[\s\-\.\/\(\)]", "", str(p).strip())
    if c.startswith("+39"):
        return c
    if c.startswith("0039"):
        return "+" + c[2:]
    if c.startswith("39") and len(c) >= 11:
        return "+" + c
    if c.startswith("0") or c.startswith("3"):
        return "+39" + c
    return c

def is_valid_phone(p):
    if not p or not p.startswith("+39"):
        return False
    return len(re.sub(r"\D", "", p)) >= 11  # +39 + >=9 cifre

def role_phrase(dm):
    d = (dm or "").lower()
    if "acquist" in d:
        return "il responsabile acquisti"
    if "commercial" in d:
        return "il responsabile commerciale"
    if "marketing" in d or "customer" in d:
        return "il responsabile marketing"
    if "direz" in d or "direttore" in d:
        return "la direzione"
    return "il titolare"

def decide_target(row):
    """Ritorna (phone_da_chiamare, nome_contatto, decision_maker_var)."""
    tit = (row[C_TITOLARE] or "").strip() if len(row) > C_TITOLARE else ""
    acq = (row[C_ACQUISTI] or "").strip() if len(row) > C_ACQUISTI else ""
    cen = normalize_phone(row[C_CENTRALINO] if len(row) > C_CENTRALINO else "")
    dir_ = normalize_phone(row[C_DIRETTO] if len(row) > C_DIRETTO else "")
    ruolo = row[C_RUOLO] if len(row) > C_RUOLO else ""
    phone = dir_ if is_valid_phone(dir_) else cen
    if tit:
        return phone, tit, "il titolare"
    if acq:
        return phone, acq, "il responsabile acquisti"
    rp = role_phrase(ruolo)
    return phone, rp, rp

def build_recipient(row):
    phone, nome_contatto, dm_var = decide_target(row)
    note = (row[C_NOTE] if len(row) > C_NOTE else "") or ""
    return {
        "phone_number": normalize_phone(phone),
        "conversation_initiation_client_data": {
            "dynamic_variables": {
                "nome_azienda": row[C_AZIENDA] if len(row) > C_AZIENDA else "",
                "categoria": (row[C_SETTORE] if len(row) > C_SETTORE else "").lower(),
                "citta": row[C_CITTA] if len(row) > C_CITTA else "",
                "numero_verde": row[C_VERDE] if len(row) > C_VERDE else "",
                "decision_maker": dm_var,
                "nome_contatto": nome_contatto,
                "giorni_consulenza": os.getenv("LOOK_GIORNI", "dal lunedi al venerdi"),
                "note": note[:180],
            }
        },
    }

# ---------- sorgenti ----------
def read_sheet():
    from google.oauth2.service_account import Credentials
    from googleapiclient.discovery import build
    if not SHEET_ID:
        raise SystemExit("LOOK_SHEET_ID non impostato. Crea il foglio, condividilo col service account "
                         "(claude-sheets@claude-telesales.iam.gserviceaccount.com) e imposta LOOK_SHEET_ID.")
    creds = Credentials.from_service_account_file(SA_PATH, scopes=["https://www.googleapis.com/auth/spreadsheets"])
    svc = build("sheets", "v4", credentials=creds)
    vals = svc.spreadsheets().values().get(spreadsheetId=SHEET_ID, range=f"'{TAB_NAME}'!A:N").execute().get("values", [])
    return svc, vals

def read_xlsx():
    import openpyxl
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb["LISTA_PROSPECT"]
    # mappa dalle colonne del Master a quelle del caller
    rows = [["Azienda","Settore","Citta","Numero verde","Ruolo","Titolare","Acquisti","Centralino","Diretto","Esito","Note","Appuntamento","DataChiamata","ConvID"]]
    import json as _j
    pros = {re.sub(r"[^a-z0-9]+"," ",str(p["azienda"]).lower()).strip(): p for p in _j.load(open(os.path.join(os.path.dirname(__file__),"prospects.json"),encoding="utf-8"))}
    for r in range(5, ws.max_row+1):
        az = ws.cell(r,2).value
        if not az: continue
        p = pros.get(re.sub(r"[^a-z0-9]+"," ",str(az).lower()).strip(), {})
        verde = re.search(r"800[\s\.]?\d[\d\s\.]{4,}", p.get("evidenza","") or "")
        rows.append([az, ws.cell(r,5).value or "", ws.cell(r,8).value or "",
                     (verde.group(0).strip() if verde else p.get("numero_verde","")),
                     ws.cell(r,12).value or "", p.get("titolare_nome",""), p.get("acquisti_nome",""),
                     ws.cell(r,10).value or "", ws.cell(r,11).value or "",
                     ws.cell(r,21).value or "", "", "", "", ""])  # col 21 = Esito
    return None, rows

def is_called(row):
    return len(row) > C_DATA_CHIAMATA and (row[C_DATA_CHIAMATA] or "").strip() != ""

def submit_batch(recipients, name):
    import requests
    good = [r for r in recipients if is_valid_phone(r["phone_number"])]
    bad = [r["phone_number"] for r in recipients if not is_valid_phone(r["phone_number"])]
    if bad:
        print(f"[submit] scartati {len(bad)} numeri non validi (non E.164): {bad[:8]}")
    if not good:
        raise SystemExit("Nessun numero valido da chiamare.")
    payload = {
        "call_name": name, "agent_id": LOOK_AGENT_ID, "agent_phone_number_id": PHONE_NUMBER_ID,
        "recipients": good, "target_concurrency_limit": CONCURRENCY,
        "telephony_call_config": {"ringing_timeout_secs": RINGING_TIMEOUT},
    }
    sched = os.getenv("LOOK_SCHEDULE_UNIX")
    if sched:
        payload["scheduled_time_unix"] = int(sched)
    resp = requests.post(f"{ELEVENLABS_BASE_URL}/v1/convai/batch-calling/submit",
                         headers={"xi-api-key": get_key(), "Content-Type": "application/json"},
                         json=payload, timeout=30)
    resp.raise_for_status()
    return resp.json()

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--limit", type=int, default=DEFAULT_LIMIT, help="0 = tutte")
    ap.add_argument("--source", choices=["sheet", "xlsx"], default="xlsx",
                    help="xlsx = Master creato da noi (default); sheet = foglio Google se condiviso col SA")
    a = ap.parse_args()

    svc, rows = read_sheet() if a.source == "sheet" else read_xlsx()
    if not rows:
        raise SystemExit("Foglio vuoto.")
    header, data = rows[0], rows[1:]

    # STRATEGIA call + re-call: l'AI RICHIAMA i no-contatto a una fascia oraria DIVERSA (aumenta la
    # conversione, cattura il decisore) + chiama NUOVI. CONCLUSI = mai. CALDI (Interessato/Referente/
    # Da ricontattare) = li chiude il SETTER, l'AI non li ritenta.
    DONE = {"Non interessato", "Non in target", "Numero/contatto errato", "Appuntamento fissato"}
    RECALL = {"Da richiamare", "Non risposto"}
    ESITO_RANK = {"Da richiamare": 0, "Non risposto": 1}
    slot = current_slot()
    stats = call_stats()
    recalls, fresh = [], []
    for i, row in enumerate(data, start=2):
        row = list(row) + [""] * (NCOLS - len(row))
        esito = (row[C_ESITO] or "").strip()
        if esito in DONE:
            continue
        phone, nome_contatto, dm = decide_target(row)
        ph = normalize_phone(phone)
        if not is_valid_phone(ph):
            continue
        key = re.sub(r"[^0-9+]", "", ph)
        if key in EXCLUDE_PHONES:
            continue
        stt = stats.get(key, {"n": 0, "slots": set(), "today": False})
        if stt["today"]:
            continue  # mai due volte nello stesso giorno
        rec = (i, row[C_AZIENDA], nome_contatto, dm, ph, row)
        if esito in RECALL and stt["n"] < MAX_ATTEMPTS:
            diff_slot = 0 if slot not in stt["slots"] else 1  # priorita a una fascia oraria non ancora provata
            recalls.append((diff_slot, ESITO_RANK.get(esito, 9), stt["n"], rec))
        elif esito in ("", "Da contattare") and stt["n"] == 0:
            fresh.append(rec)

    recalls.sort(key=lambda x: (x[0], x[1], x[2]))
    recall_recs = [x[3] for x in recalls]
    limit = a.limit or (len(recall_recs) + len(fresh))
    cap_recall = int(round(limit * 0.4))  # fino al 40% del blocco = re-call; resto NUOVI
    sel_recall = recall_recs[:cap_recall]
    sel_fresh = fresh[:limit - len(sel_recall)]
    rest = limit - len(sel_recall) - len(sel_fresh)
    if rest > 0:
        sel_recall += recall_recs[len(sel_recall):len(sel_recall) + rest]
    chosen = sel_recall + sel_fresh
    n_rc, n_fr = len(sel_recall), len(sel_fresh)
    recipients = [build_recipient(rec[5]) for rec in chosen]
    picked = [(rec[0], rec[1], rec[2], rec[3], rec[4]) for rec in chosen]

    print(f"Blocco {slot}: {len(chosen)} chiamate = {n_rc} RE-CALL (no-contatto, fascia diversa) + {n_fr} NUOVI")
    for i, az, nc, dm, ph in picked[:15]:
        print(f"  riga {i}: {az[:30]:30} -> {nc[:24]:24} | {ph}")
    if len(picked) > 15:
        print(f"  ... e altri {len(picked)-15}")

    if a.dry_run:
        print("\n[DRY-RUN] nessuna chiamata inviata.")
        return
    if not recipients:
        print("Niente da chiamare.")
        return

    name = "Look - " + datetime.datetime.now().strftime("%Y%m%d_%H%M")
    res = submit_batch(recipients, name)
    print("BATCH INVIATO:", json.dumps(res)[:300])
    # registra i numeri chiamati (per non richiamarli nei prossimi RECALL_DAYS giorni)
    log_called([p[4] for p in picked])
    # write-back data chiamata (solo sorgente sheet)
    if a.source == "sheet" and svc is not None:
        today = datetime.date.today().strftime("%d/%m/%Y")
        for (i, *_), in zip([(p,) for p in picked]):
            pass  # la marcatura puntuale avviene nel post-batch analyzer (modello Culligan)
        print("Ricorda: esiti/note dal post-batch analyzer (fetch transcript -> esito/nota/appuntamento sul foglio).")

if __name__ == "__main__":
    main()
