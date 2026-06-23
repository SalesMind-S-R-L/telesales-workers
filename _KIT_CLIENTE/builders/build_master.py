#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
BUILDER FISSO - Master Excel cliente (kit Telesales standardizzato).
Struttura identica per ogni cliente. La ricerca alimenta solo i dati via JSON.

USO:  python3 build_master.py /percorso/cartella_cliente
La cartella deve contenere: config.json, prospects.json
Opzionali: guida.json, kpi.json  (se assenti -> default ragionevoli)
Output:   Master_<CLIENTE>.xlsx nella stessa cartella.

CONTRATTO DATI -> vedi cartella esempi/ del kit.
"""
import sys, os, json, re, unicodedata, urllib.parse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

WD = sys.argv[1] if len(sys.argv) > 1 else '.'
def load(name, default=None):
    p = os.path.join(WD, name)
    if os.path.exists(p):
        return json.load(open(p, encoding='utf-8'))
    return default

config    = load('config.json', {})
prospects = load('prospects.json', [])
guida     = load('guida.json', {})
kpi       = load('kpi.json', {})

CLIENTE = config.get('cliente', 'CLIENTE')
SAFE = re.sub(r'[^A-Za-z0-9]+', '_', CLIENTE).strip('_')

# ---------- stile (fisso, schema Giglioli/IMEON) ----------
C_TITLE = PatternFill('solid', fgColor='1F3864')
C_HDR   = PatternFill('solid', fgColor='1F3864')
C_A     = PatternFill('solid', fgColor='C8E6C9')  # verde
C_B     = PatternFill('solid', fgColor='FFF59D')  # giallo
C_C     = PatternFill('solid', fgColor='F2F2F2')  # grigio
F_TITLE = Font(color='FFFFFF', bold=True, size=14)
F_HDRW  = Font(color='FFFFFF', bold=True)
F_SUB   = Font(italic=True, size=10, color='555555')
F_B     = Font(bold=True)
WRAP    = Alignment(wrap_text=True, vertical='top')
TOP     = Alignment(vertical='top')
fillmap = {'A': C_A, 'B': C_B, 'C': C_C}

wb = openpyxl.Workbook()
def title_block(ws, title, sub):
    ws['A1'] = title; ws['A1'].font = F_TITLE; ws['A1'].fill = C_TITLE
    ws['A2'] = sub;   ws['A2'].font = F_SUB
def fill_row_title(ws, ncol):
    for c in range(1, ncol+1): ws.cell(row=1, column=c).fill = C_TITLE

def norm(s):
    s = unicodedata.normalize('NFKD', str(s)).encode('ascii', 'ignore').decode().lower()
    return re.sub(r'\s+', ' ', re.sub(r'[^a-z0-9 ]', ' ', s)).strip()
def li_search(az):
    return 'https://www.google.com/search?q=' + urllib.parse.quote(f'{az} responsabile marketing site:linkedin.com/in')

# ---------- derivazione (fissa) ----------
ACC_OF  = {'Enterprise/multinazionale': 'Difficile', 'Grande nazionale': 'Media', 'Mid-large': 'Buona'}
INV_PTS = {'Alto': 60, 'Medio': 40, 'Basso': 20}
ACC_PTS = {'Buona': 8, 'Media': 4, 'Difficile': 0}
SECT_W  = config.get('settori_weight', {})  # opzionale: {"Fashion & Lifestyle":4,...}

def norm_dim(s):
    t = str(s or '').lower()
    if 'enter' in t or 'multinaz' in t: return 'Enterprise/multinazionale'
    if 'mid' in t or 'nicchia' in t:    return 'Mid-large'
    return 'Grande nazionale'
def clean_inv(v):
    v = str(v or '').strip().capitalize()
    return v if v in ('Alto', 'Medio', 'Basso') else 'Medio'

recs = []
for p in prospects:
    az = str(p.get('azienda', '')).strip()
    if not az: continue
    se = str(p.get('settore', '')).strip() or 'Altro'
    dim = norm_dim(p.get('dimensione'))
    acc = ACC_OF[dim]
    inv = clean_inv(p.get('investimento'))
    nuovo = str(p.get('nuovo_investimento', '') or '').strip()
    if nuovo.lower() in ('no', 'false', 'none'): nuovo = ''
    has_nuovo = bool(nuovo)
    if inv == 'Alto' and has_nuovo: prio = 'A'
    elif inv == 'Alto' or has_nuovo or (inv == 'Medio' and acc == 'Buona'): prio = 'B'
    else: prio = 'C'
    score = INV_PTS[inv] + (25 if has_nuovo else 0) + ACC_PTS[acc] + SECT_W.get(se, 0)
    recs.append(dict(
        az=az, se=se, sede=str(p.get('sede', '') or ''), sito=str(p.get('sito', '') or ''),
        dim=dim, acc=acc, inv=inv, nuovo=nuovo, prio=prio, score=score,
        quickwin='Quick win' if (acc == 'Buona' and inv in ('Alto', 'Medio')) else '',
        dm=str(p.get('decision_maker', '') or ''), li=li_search(az),
        nome=str(p.get('nome_decisore', '') or '').strip() or str(p.get('decision_maker', '') or ''),
        evid=str(p.get('evidenza', '') or '').strip()[:300] or 'Da verificare',
        perche=str(p.get('perche_in_target', '') or ''), hook=str(p.get('cosa_dire', '') or ''),
        telc=str(p.get('telefono_centralino', '') or '').strip(),
        teld=str(p.get('telefono_diretto', '') or '').strip(),
        telf=str(p.get('fonte_telefono', '') or '').strip(),
        origine=str(p.get('origine', '') or 'In lista'),
        canale=('Chiamata (centralino > marketing)' if prio == 'A' else ('Chiamata + LinkedIn' if prio == 'B' else 'LinkedIn + email')),
    ))
order = {'A': 0, 'B': 1, 'C': 2}
recs.sort(key=lambda x: (order[x['prio']], -x['score'], x['se'], x['az']))
TOT = len(recs)
nA = sum(1 for x in recs if x['prio'] == 'A'); nB = sum(1 for x in recs if x['prio'] == 'B'); nC = sum(1 for x in recs if x['prio'] == 'C')
n_qw = sum(1 for x in recs if x['quickwin'])
n_telc = sum(1 for x in recs if x['telc'])
n_giant = sum(1 for x in recs if x['dim'].startswith('Enterprise'))
n_grande = sum(1 for x in recs if x['dim'] == 'Grande nazionale')
n_mid = sum(1 for x in recs if x['dim'] == 'Mid-large')
sect_count = {}
for x in recs: sect_count[x['se']] = sect_count.get(x['se'], 0) + 1

DEF_ESITI = ['Da contattare', 'Non risposto', 'Da richiamare', 'Referente individuato', 'Interessato',
             'Appuntamento fissato', 'Da ricontattare piu avanti', 'Non in target', 'Non interessato', 'Numero/contatto errato']
esiti_list = [e[0] if isinstance(e, (list, tuple)) else e for e in guida.get('esiti', [])] or DEF_ESITI

# ============================================================ HOME
ws = wb.active; ws.title = 'HOME'
title_block(ws, config.get('titolo_master', f'MASTER {CLIENTE.upper()}'),
            config.get('sottotitolo', f'Strumento di lavoro Telesales per {CLIENTE} - aggiornato {config.get("data","")}'))
rows = [('',),
 (f'OBIETTIVO: {config.get("obiettivo","")}',),
 (f'SERVIZIO VENDUTO: {config.get("servizio","")}',),
 (f'MODALITA LISTA: {config.get("modalita_lista","scouting")}',),
 (f'CALENDARIO APPUNTAMENTI: {config.get("calendario_cliente","")}',) if config.get('calendario_cliente') else ('',),
 ('',),
 ('SE SEI UN SETTER: 1) Leggi GUIDA_SETTER  2) Lavora da LISTA_PROSPECT partendo dai VERDI in alto (gia ordinati per lavorabilita).',),
 ('',),
 ('COSA TROVI IN OGNI TAB', True),
 ('Tab', 'A chi serve', 'Cosa contiene'),
 ('HOME', 'Tutti', 'Indice, riepilogo numerico, regole d\'oro'),
 ('GUIDA_SETTER', 'Setter', 'Workflow, script, scenari, obiezioni, esiti, tono'),
 ('LISTA_PROSPECT', 'Setter', 'Lista di lavoro ordinata per priorita e lavorabilita, con telefoni ed evidenza'),
 ('ICP_e_INSIGHT', 'Telesales', 'Profilo cliente ideale + insight di mercato -> uso in chiamata'),
 ('KPI_TRACKING', 'Tutti', 'Target appuntamenti, review settimanale, prossimi passi'),
 ('DB_COMPLETO', 'Riferimento', 'Tutti i prospect + telefoni/fonti + legenda'),
 ('',),
 ('RIEPILOGO NUMERICO', True),
 (f'Prospect totali in lista', TOT),
 (f'  VERDE  A - investe Alto E segnale di nuovo investimento (di cui {n_qw} quick win raggiungibili)', nA),
 ('  GIALLO B - investe Alto, oppure ha un segnale di nuovo investimento', nB),
 ('  GRIGIO C - investe medio/basso senza segnali immediati', nC),
 (f'Telefoni centralino verificati', f'{n_telc} / {TOT}'),
 ('',),
 ('LAVORABILITA (dimensione -> accessibilita del decisore)', True),
 ('  Enterprise/multinazionale (difficile: procurement/gare, ciclo lungo)', n_giant),
 ('  Grande nazionale (media)', n_grande),
 ('  Mid-large (buona: decisore raggiungibile, appuntamento piu rapido)', n_mid),
 ('',),
 ('PROSPECT PER SETTORE', True)]
for s in sorted(sect_count): rows.append((f'  {s}', sect_count[s]))
rows += [('',), ('REGOLE D\'ORO IN PILLOLE (dettaglio in GUIDA_SETTER)', True)]
for g in (config.get('regole_oro') or ['L\'obiettivo della chiamata e l\'APPUNTAMENTO con il decisore.',
          'In chiamata NON si fanno prezzi.', 'Note brevi e professionali, mai gergo tecnico interno.']):
    rows.append((f'  - {g}',))
rows += [('',), (f'ULTIMO SYNC - {config.get("ultimo_sync", config.get("data",""))}', True)]
r = 3
for tup in rows:
    bold = len(tup) > 1 and tup[1] is True
    ws.cell(row=r, column=1, value=tup[0])
    if bold: ws.cell(row=r, column=1).font = F_B
    elif len(tup) > 1: ws.cell(row=r, column=2, value=tup[1])
    r += 1
# header riga "Tab | A chi serve | Cosa contiene"
for rr in range(3, r):
    if ws.cell(rr, 1).value == 'Tab':
        for c in range(1, 4):
            ws.cell(rr, c).fill = C_HDR; ws.cell(rr, c).font = F_HDRW
fill_row_title(ws, 3)
for k, v in {'A': 70, 'B': 22, 'C': 60}.items(): ws.column_dimensions[k].width = v

# ============================================================ GUIDA_SETTER
ws = wb.create_sheet('GUIDA_SETTER')
title_block(ws, 'GUIDA SETTER', guida.get('intro', f'Come chiamare i prospect di {CLIENTE} e fissare appuntamenti con il decisore.'))
G = []
G.append(('REGOLA NUMERO 1: il successo della chiamata e l\'APPUNTAMENTO con il decisore. Tutto il resto e dettaglio.', True))
G.append(('',))
G.append(('1. CHI STAI CHIAMANDO', True))
for p in guida.get('chi_chiami', ['Aziende in target che gia investono in comunicazione/contenuti. Obiettivo: parlare col decisore marketing e fissare un appuntamento conoscitivo.']):
    G.append((p,))
G.append(('',))
G.append(('2. WORKFLOW GIORNALIERO', True))
for row in guida.get('workflow', [['Mattina', 'Apri LISTA_PROSPECT, parti dal primo VERDE in alto.'],
        ['Prima della chiamata', 'Leggi colonna "Cosa dire" e l\'evidenza: sai gia perche e in target.'],
        ['Durante', 'Chiama il centralino e chiedi del marketing. Usa lo script (sez. 4).'],
        ['Subito dopo', 'Aggiorna ESITO, NOTE e se fissi APPUNTAMENTO.'],
        ['Fine giornata', 'Mini-report a Telesales.']]):
    G.append((row[0], row[1]))
G.append(('',))
G.append(('3. COSA SIGNIFICA OGNI COLORE', True))
for row in guida.get('colori', [['VERDE - A', 'Investe molto e si muove ora. Chiama questi PRIMA.'],
        ['GIALLO - B', 'Investe molto, oppure ha un segnale di nuovo investimento. Dopo i verdi.'],
        ['GRIGIO - C', 'Investe medio/basso senza segnali immediati. Solo dopo.']]):
    G.append((row[0], row[1]))
G.append(('',))
G.append(('4. SCRIPT', True))
sc = guida.get('script', {})
if sc.get('regola'): G.append((sc['regola'],))
for label, key in [('4.A - Apertura', 'apertura'), ('4.B - Sviluppo/qualifica', 'sviluppo'), ('4.C - Chiusura appuntamento', 'chiusura')]:
    if sc.get(key):
        G.append((label, True)); G.append((sc[key],))
G.append(('',))
G.append(('5. SCENARI', True))
for row in guida.get('scenari', []): G.append((row[0], row[1] if len(row) > 1 else ''))
G.append(('',))
G.append(('6. GESTIONE OBIEZIONI', True))
for row in guida.get('obiezioni', []): G.append((row[0], row[1] if len(row) > 1 else ''))
G.append(('',))
G.append(('7. REGOLE D\'ORO', True))
for row in guida.get('regole_oro_setter', [['1. L\'obiettivo e l\'APPUNTAMENTO.', ''], ['2. MAI prezzi al telefono.', ''], ['3. Non bruciare il contatto.', '']]):
    G.append((row[0], row[1] if len(row) > 1 else ''))
G.append(('',))
G.append(('8. ESITI (usa il menu a tendina)', True))
for e in guida.get('esiti', [[x, ''] for x in DEF_ESITI]):
    if isinstance(e, (list, tuple)): G.append((e[0], e[1] if len(e) > 1 else ''))
    else: G.append((e, ''))
G.append(('',))
G.append(('9. TONO DI VOCE', True))
for row in guida.get('tono', [['Calmo', 'Mai affannato.'], ['Concreto', 'Frasi corte, due opzioni di orario.'], ['Pari grado', 'Professionale, mai supplicante.']]):
    G.append((row[0], row[1] if len(row) > 1 else ''))
G.append(('',))
G.append(('10. FINE GIORNATA - mini-report', True))
G.append((guida.get('mini_report', 'Esempio: "X chiamate, Y centralini passati al marketing, Z appuntamenti fissati."'),))
r = 3
for tup in G:
    bold = len(tup) > 1 and tup[1] is True
    ws.cell(row=r, column=1, value=tup[0]); ws.cell(row=r, column=1).alignment = WRAP
    if bold: ws.cell(row=r, column=1).font = F_B
    elif len(tup) > 1:
        ws.cell(row=r, column=2, value=tup[1]); ws.cell(row=r, column=2).alignment = WRAP
    r += 1
fill_row_title(ws, 2)
ws.column_dimensions['A'].width = 60; ws.column_dimensions['B'].width = 80

# ============================================================ LISTA_PROSPECT
ws = wb.create_sheet('LISTA_PROSPECT')
cols = ['#', 'Azienda', 'Prio', 'Quick win', 'Settore', 'Dimensione', 'Accessibilita DM', 'Sede', 'Sito',
        'Telefono centralino', 'Telefono diretto', 'Decision Maker (ruolo)', 'Nome decisore (chi chiamare)', 'LinkedIn (ricerca)',
        'Investimento', 'Evidenza investimento adv/contenuti', 'Nuovo investimento',
        'Perche in target', 'Cosa dire (gancio)', 'Canale approccio', 'Esito', 'Note', 'Appuntamento (data/ora)']
title_block(ws, 'LISTA_PROSPECT - la tua lista di lavoro',
    f'{nA} verdi (A) - {nB} gialli (B) - {nC} grigi (C) su {TOT} prospect. Verificate per investimento adv/contenuti. In cima i piu lavorabili. In chiamata NON si fanno prezzi.')
ws['A3'] = 'VERDE = chiama PRIMA | GIALLO = dopo | GRIGIO = solo dopo | Telefono centralino verificato; diretto solo dove pubblico.'
ws['A3'].font = F_SUB
hdr = 4
for j, c in enumerate(cols, 1):
    cell = ws.cell(row=hdr, column=j, value=c); cell.fill = C_HDR; cell.font = F_HDRW; cell.alignment = WRAP
r = hdr + 1
for i, x in enumerate(recs, 1):
    vals = [i, x['az'], x['prio'], x['quickwin'], x['se'], x['dim'], x['acc'], x['sede'], x['sito'],
            x['telc'], x['teld'], x['dm'], x['nome'], x['li'], x['inv'], x['evid'], x['nuovo'],
            x['perche'], x['hook'], x['canale'], 'Da contattare', '', '']
    for j, v in enumerate(vals, 1):
        cell = ws.cell(row=r, column=j, value=v); cell.fill = fillmap[x['prio']]
        cell.alignment = WRAP if j in (13, 14, 16, 18, 19) else TOP
    ws.cell(row=r, column=10).number_format = '@'; ws.cell(row=r, column=11).number_format = '@'
    r += 1
last = r - 1
def add_dv(formula, rng):
    dv = DataValidation(type='list', formula1=formula, allow_blank=True)
    ws.add_data_validation(dv); dv.add(rng)
add_dv('"' + ','.join(esiti_list) + '"', f'U{hdr+1}:U{last}')
add_dv('"A,B,C"', f'C{hdr+1}:C{last}')
add_dv('"Alto,Medio,Basso"', f'O{hdr+1}:O{last}')
add_dv('"Difficile,Media,Buona"', f'G{hdr+1}:G{last}')
W = {'A':5,'B':28,'C':5,'D':10,'E':19,'F':22,'G':14,'H':20,'I':22,'J':18,'K':18,'L':24,'M':38,'N':34,'O':12,'P':58,'Q':30,'R':36,'S':44,'T':24,'U':22,'V':28,'W':18}
for k, v in W.items(): ws.column_dimensions[k].width = v
ws.freeze_panes = 'A5'; ws.auto_filter.ref = f'A{hdr}:W{last}'

# ============================================================ ICP_e_INSIGHT
ws = wb.create_sheet('ICP_e_INSIGHT')
title_block(ws, f'ICP e INSIGHT - {CLIENTE}', 'Profilo del cliente ideale + insight di mercato (dettaglio nel Dossier).')
r = 4
ws.cell(r, 1, 'PROFILO CLIENTE IDEALE (ICP)').font = F_B; r += 1
for c, t in enumerate(['Parametro', 'Valore'], 1):
    cell = ws.cell(r, c, t); cell.fill = C_HDR; cell.font = F_HDRW
r += 1
for row in config.get('icp', [['Settori', ', '.join(sorted(sect_count))], ['Profilo', 'Aziende che gia investono in adv/contenuti']]):
    ws.cell(r, 1, row[0]); cc = ws.cell(r, 2, row[1] if len(row) > 1 else ''); cc.alignment = WRAP; r += 1
r += 1
ws.cell(r, 1, 'INSIGHT DI MERCATO -> uso in chiamata').font = F_B; r += 1
for c, t in enumerate(['Area', 'Insight chiave (dato verificato)', 'Come usarlo in chiamata'], 1):
    cell = ws.cell(r, c, t); cell.fill = C_HDR; cell.font = F_HDRW; cell.alignment = WRAP
r += 1
for ins in config.get('insight', []):
    ws.cell(r, 1, ins[0] if len(ins) > 0 else '')
    a = ws.cell(r, 2, ins[1] if len(ins) > 1 else ''); a.alignment = WRAP
    b = ws.cell(r, 3, ins[2] if len(ins) > 2 else ''); b.alignment = WRAP
    r += 1
fill_row_title(ws, 3)
for k, v in {'A': 22, 'B': 58, 'C': 58}.items(): ws.column_dimensions[k].width = v

# ============================================================ KPI_TRACKING
ws = wb.create_sheet('KPI_TRACKING')
title_block(ws, 'KPI_TRACKING - target + review', kpi.get('intro', f'Obiettivo appuntamenti/mese: {config.get("obiettivo_appuntamenti","da definire")}'))
r = 4
ws.cell(r, 1, 'TARGET').font = F_B; r += 1
for c, t in enumerate(['Obiettivo', 'Target', 'Attuale', 'Note'], 1):
    cell = ws.cell(r, c, t); cell.fill = C_HDR; cell.font = F_HDRW
r += 1
for row in kpi.get('target', [['Appuntamenti in target fissati', config.get('obiettivo_appuntamenti', 'da definire'), 0, '']]):
    for c, v in enumerate(row, 1): ws.cell(r, c, v)
    r += 1
r += 1
ws.cell(r, 1, 'REVIEW SETTIMANALE').font = F_B; r += 1
for c, t in enumerate(['Settimana', 'Date', 'N. chiamate', 'Centralini passati', 'Appuntamenti fissati'], 1):
    cell = ws.cell(r, c, t); cell.fill = C_HDR; cell.font = F_HDRW
r += 1
for row in kpi.get('review_settimane', [['W1', ''], ['W2', ''], ['W3', ''], ['W4', '']]):
    for c, v in enumerate(row, 1): ws.cell(r, c, v)
    r += 1
r += 1
ws.cell(r, 1, 'PROSSIMI PASSI').font = F_B; r += 1
for s in kpi.get('prossimi_passi', ['1. Inviare la lista al cliente per approvazione.', '2. Avviare le chiamate dai verdi/quick win.']):
    ws.cell(r, 1, s); r += 1
fill_row_title(ws, 5)
for k, v in {'A': 42, 'B': 16, 'C': 14, 'D': 18, 'E': 20}.items(): ws.column_dimensions[k].width = v

# ============================================================ DB_COMPLETO
ws = wb.create_sheet('DB_COMPLETO')
title_block(ws, 'DB COMPLETO - tutti i prospect (backup) + legenda',
    f'{TOT} prospect. Telefoni verificati da fonte ufficiale. Vuoto = non verificabile (da non inventare).')
cols = ['#', 'Prio', 'Azienda', 'Settore', 'Dimensione', 'Accessibilita DM', 'Sede', 'Sito',
        'Telefono centralino', 'Telefono diretto', 'Fonte telefono', 'Investimento', 'Nuovo investimento',
        'Evidenza investimento adv/contenuti', 'Decision Maker (ruolo)', 'Origine', 'Score']
hdr = 4
for j, c in enumerate(cols, 1):
    cell = ws.cell(row=hdr, column=j, value=c); cell.fill = C_HDR; cell.font = F_HDRW; cell.alignment = WRAP
r = hdr + 1
for i, x in enumerate(recs, 1):
    vals = [i, x['prio'], x['az'], x['se'], x['dim'], x['acc'], x['sede'], x['sito'],
            x['telc'], x['teld'], x['telf'], x['inv'], x['nuovo'], x['evid'], x['dm'], x['origine'], x['score']]
    for j, v in enumerate(vals, 1):
        cell = ws.cell(row=r, column=j, value=v); cell.fill = fillmap[x['prio']]
        cell.alignment = WRAP if j in (13, 14) else TOP
    ws.cell(row=r, column=9).number_format = '@'; ws.cell(row=r, column=10).number_format = '@'
    r += 1
dblast = r - 1
ws.freeze_panes = 'A5'; ws.auto_filter.ref = f'A{hdr}:Q{dblast}'
W = {'A':5,'B':6,'C':28,'D':19,'E':24,'F':14,'G':20,'H':20,'I':18,'J':18,'K':40,'L':12,'M':30,'N':58,'O':24,'P':20,'Q':7}
for k, v in W.items(): ws.column_dimensions[k].width = v
r = dblast + 3
ws.cell(r, 1, 'LEGENDA').font = F_B; r += 1
leg = [('Prio A (verde)', 'Investe Alto e ha un segnale di nuovo investimento ora.'),
       ('Prio B (giallo)', 'Investe Alto, oppure ha un segnale di nuovo investimento, oppure mid-large accessibile che investe.'),
       ('Prio C (grigio)', 'Investe medio/basso senza segnali immediati.'),
       ('Quick win', 'Accessibilita Buona + investe Alto/Medio: i primi da chiamare.'),
       ('Dimensione', 'Enterprise/multinazionale (difficile) | Grande nazionale (media) | Mid-large (buona). Stima per fascia.'),
       ('Accessibilita DM', 'Quanto e facile arrivare al decisore e fissare. Inversamente legata alla dimensione.'),
       ('Telefono centralino', 'Numero ufficiale verificato (sito Contatti, registro, elenco verificato). Formato nazionale, niente +39. Vuoto = non verificabile.'),
       ('Telefono diretto', 'Linea diretta reparto/decisore SOLO se pubblicata su fonte ufficiale. Vuoto = non disponibile pubblicamente (da non inventare).'),
       ('Fonte telefono', 'URL della fonte da cui e stato letto il numero (tracciabilita anti-errore).'),
       ('Investimento/Evidenza/Nuovo', 'Prova concreta che gia investe in adv/contenuti + segnale che vuole investire ora.'),
       ('Score', 'Investimento + segnale nuovo + accessibilita + peso settore.')]
for k, v in leg:
    ws.cell(r, 1, k).font = F_B; cc = ws.cell(r, 3, v); cc.alignment = WRAP; r += 1
fill_row_title(ws, len(cols))

out = os.path.join(WD, f'Master_{SAFE}.xlsx')
wb.save(out)
print('SALVATO:', out)
print('Tab:', wb.sheetnames)
print(f'Prospect: {TOT} | A {nA} B {nB} C {nC} | quick win {n_qw} | tel centralino {n_telc}/{TOT}')
print('Dimensione:', n_giant, 'enterprise /', n_grande, 'grandi /', n_mid, 'mid-large')
